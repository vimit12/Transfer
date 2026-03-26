# ======================
# core/excel_generator.py — Excel report generation
# ======================
import re
import sys
import os
import numpy as np
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from core.utils import (
    ROUNDED_VALUES, get_month_details, date_calculation,
    get_details_for_name, preprocess_name, sanitize_sheet_name
)

TOTAL_WORKING_DAY = 0
range_limit = 0


def generate_excel(month, year, output_file_name, selected_row, holiday_list,
                   name_mapping, name_order_list, progress_bar):
    global TOTAL_WORKING_DAY
    sheets_name = []
    try:
        user_data = []
        user_leave_record = []
        non_complaince_user = []
        month_name = month
        year = int(year)

        month_details, month_number = get_month_details(month_name, year)
        month_number = f"{month_number:02}" if month_number < 10 else month_number
        month_holiday_list = [x for x in holiday_list if re.findall(f"\\d+-{month_number}-\\d+", x)]
        month_day_holiday_list = [k.split("-")[0] for k in month_holiday_list]
        df_sheets = {}
        excel_file_path = output_file_name

        order_map = {preprocess_name(name): idx for idx, name in enumerate(name_order_list)}
        selected_row = sorted(
            selected_row,
            key=lambda x: order_map.get(preprocess_name(x["Rsname"]), float('inf'))
        )

        attendance_len = len(selected_row)
        progress_step = 0
        step = 100 / attendance_len

        for new_data in selected_row:
            leave_dates = []
            start_date = end_date = sd = sm = sy = ed = em = ey = sm_flag = em_flag = None
            billable_days = weekends = total_working_days = leave_taken = public_holiday = 0
            data_model = []
            mismatch_date = []
            msg = None

            name = preprocess_name(new_data.get("Rsname"))
            details = get_details_for_name(name, name_mapping)
            start_date, end_date = details[2], details[3]

            if start_date:
                sd, sm, sy = date_calculation(start_date)
                sm = f"{sm:02}" if sm < 10 else sm
                sm_flag = sm == month_number

            if end_date:
                ed, em, ey = date_calculation(end_date)
                em = f"{em:02}" if em < 10 else em
                em_flag = em == month_number

            for week in month_details:
                for day in week:
                    if not day:
                        continue
                    date = day["day"]
                    is_weekend_or_leave = "Weekend" if day["is_weekend"] else ""
                    if day["is_weekend"]:
                        weekends += 1
                    else:
                        total_working_days += 1

                    dt = f"{date}-{month_name[:3]}"
                    day_name = day['day_name']

                    # On-boarding / off-boarding guards
                    if sm_flag or em_flag:
                        if sm_flag and not em_flag:
                            if date < sd:
                                data_model.append((dt, day_name[:3], "Not On Boarded", "", "", "", "", ""))
                                continue
                        elif not sm_flag and em_flag:
                            if date > ed:
                                data_model.append((dt, day_name[:3], "Off Boarded", "", "", "", "", ""))
                                continue
                        else:
                            if date < sd:
                                data_model.append((dt, day_name[:3], "Not On Boarded", "", "", "", "", ""))
                                continue
                            elif date > ed:
                                data_model.append((dt, day_name[:3], "Off Boarded", "", "", "", "", ""))
                                continue
                            else:
                                msg = "On Board"

                    key = (f"{day_name[:3]}, {f'{date:02}' if date < 10 else date}-{month_name[:3].title()}")
                    calculated_date = f"{date:02}" if date < 10 else f"{date}"

                    if day["is_weekend"]:
                        dt_status = 0
                    else:
                        match new_data.get(key):
                            case 8:
                                dt_status = 1
                                if calculated_date in month_day_holiday_list:
                                    mismatch_date.append(calculated_date)
                                    public_holiday += 1
                                    dt_status = 0
                                    is_weekend_or_leave = "Holiday"
                            case 4:
                                dt_status = 0.5
                                leave_taken += 0.5
                                leave_dates.append(f"{date}(0.5)")
                            case value if value in ROUNDED_VALUES:
                                dt_status = 0.25
                                leave_taken += 0.25
                            case 0:
                                if calculated_date in month_day_holiday_list:
                                    public_holiday += 1
                                    dt_status = 0
                                    is_weekend_or_leave = "Holiday"
                                else:
                                    leave_taken += 1
                                    dt_status = 0
                                    leave_dates.append(date)
                                    is_weekend_or_leave = "Leave"
                            case _:
                                leave_taken += 1
                                dt_status = 0

                    billable_days += dt_status

                    if sm_flag and not em_flag and date < sd:
                        dt_status = 0
                        total_working_days -= 1
                    if em_flag and not sm_flag and date > ed:
                        dt_status = 0
                        total_working_days -= 1
                    if em_flag and sm_flag and (date > ed or date < sd):
                        dt_status = 0
                        total_working_days -= 1

                    if is_weekend_or_leave == "Holiday" and dt_status == 0:
                        data_model.append((dt, day_name[:3], "Holiday", "", "", "", "", ""))
                    elif is_weekend_or_leave == "Leave" and dt_status == 0:
                        data_model.append((dt, day_name[:3], "Leave", "", "", "", "", ""))
                    elif is_weekend_or_leave == "Weekend" and dt_status == 0:
                        data_model.append((dt, day_name[:3], dt_status, is_weekend_or_leave, "", "", "", ""))
                    else:
                        if dt_status == 1:
                            data_model.append((dt, day_name[:3], 8, is_weekend_or_leave, "", "", "", ""))
                        else:
                            data_model.append((dt, day_name[:3], 4, is_weekend_or_leave, "", "", "", ""))

            billable_days = total_working_days - leave_taken
            point_of_contact = details[1] if details else "xxxxxxx"
            ID_521 = details[0] if details else "xxxxxxx"

            if mismatch_date:
                non_complaince_user.append({
                    "Name": new_data.get("Rsname"), "521_ID": details[0], "Year": year,
                    "Month": month, "Listed Month Holiday": month_day_holiday_list,
                    "Attendance Marked on Holiday": mismatch_date,
                })

            data = {
                "Vendor Organization": ["Resource Name", "Month", "Date"],
                "Hitachi Digital Service": [f"{new_data.get('Rsname')}", f"{month_name}", "Day"],
                "Point of Contact": ["5-2-1", "Working Days", "Working Status"],
                f"{point_of_contact}": [f"{ID_521}", total_working_days, "Remarks"],
                "Adjustments from Last Month": ["", "", ""], "0": ["", "", ""], "": ["", "", ""],
                "Week Off": ["Personal/Sick Leave", "", ""],
            }
            df = pd.DataFrame(data)
            sheet_name = sanitize_sheet_name(new_data.get("Rsname"))

            billable_days = 0
            user_days_cal = 0
            for row in data_model:
                value = row[2]
                is_weekend_flag = row[3] == 'Weekend'
                if isinstance(value, (int, float)):
                    if value == 8:
                        billable_days += 1
                    elif value == 4:
                        billable_days += 0.5
                valid_status = ['Not On Boarded', 'Off Boarded']
                if value not in valid_status and not is_weekend_flag:
                    user_days_cal += 1
                df.loc[len(df)] = row

            df.loc[len(df)] = ["Leaves Taken", leave_taken, "Billable Days", billable_days, "", "", "", ""]
            df.loc[len(df)] = ["Weekends", weekends, "", "", "", "", "", ""]
            df.loc[len(df)] = ["Public Holidays", public_holiday, "", "", "", "", "", ""]
            df_sheets[sheet_name] = df

            if sm_flag or em_flag:
                user_data.append({
                    "Name": sheet_name,
                    "Billable Time (Hours)": (user_days_cal - leave_taken) * 8,
                    "Total Number of Billable Days": user_days_cal - leave_taken,
                    "Service Credit Pool Days": leave_taken,
                })
            else:
                user_data.append({
                    "Name": sheet_name,
                    "Billable Time (Hours)": (total_working_days - leave_taken) * 8,
                    "Total Number of Billable Days": total_working_days - leave_taken,
                    "Service Credit Pool Days": leave_taken,
                })

            user_leave_record.append({
                "name": sheet_name, "id_521": details[0],
                "year": year, "month": month, "leave_days": leave_dates,
            })
            progress_step += int(step)
            progress_bar.setValue(progress_step)
        else:
            TOTAL_WORKING_DAY = total_working_days

            with pd.ExcelWriter(excel_file_path, engine="xlsxwriter") as writer:
                for key, value in df_sheets.items():
                    value.to_excel(writer, sheet_name=key, index=False)
                workbook = writer.book
                worksheets = writer.sheets
                for sheet_name, worksheet in worksheets.items():
                    worksheet.set_column(0, 0, 20)
                    worksheet.set_column(1, 1, 20)
                    worksheet.set_column(2, 2, 20)
                    worksheet.set_column(3, 3, 15)
                    worksheet.set_column(4, 4, 30)
                    worksheet.set_column(7, 7, 17)
                    sheets_name.append(sheet_name)

            wb_style = load_workbook(excel_file_path)
            border = Border(
                left=Side(border_style="thin"), right=Side(border_style="thin"),
                top=Side(border_style="thin"), bottom=Side(border_style="thin"),
            )
            for i in sheets_name:
                sheet = wb_style[i]
                for row in sheet.iter_rows():
                    for cell in row:
                        cell.border = border
                        if cell.value == "Weekend":
                            numbers = re.findall(r"\d+", cell.coordinate)[0]
                            for k in [f"B{numbers}", f"C{numbers}", f"D{numbers}"]:
                                wb_style[i][k].fill = PatternFill(start_color="b6bbbf", end_color="b6bbbf", fill_type="solid")
                        if cell.value == "Leave":
                            cell.fill = PatternFill(start_color="fce1dc", end_color="fce1dc", fill_type="solid")
                        if cell.value == "Holiday":
                            cell.fill = PatternFill(start_color="cffccf", end_color="cffccf", fill_type="solid")
                        if cell.value in ["Leaves Taken", "Weekends", "Public Holidays", "Billable Days"]:
                            cell.font = Font(bold=True, color="FFFFFF")
                            cell.fill = PatternFill(start_color="4d6c82", end_color="4d6c82", fill_type="solid")

                for j in ["B1", "D1", "H1", "F1"]:
                    sheet[j].font = Font(bold=False)

                for j in ["A1", "C1", "E1", "A2", "C2", "A3", "C3", "A4", "B4", "C4", "D4", "G1", "G2"]:
                    cell_bold = sheet[j]
                    cell_bold.font = Font(bold=True)
                    if j == "G2":
                        cell_bold.fill = PatternFill(start_color="fce1dc", end_color="fce1dc", fill_type="solid")
                    else:
                        cell_bold.fill = PatternFill(start_color="b6bbbf", end_color="b6bbbf", fill_type="solid")

                for j in ["A4", "B4", "C4", "D4", "E4", "F4"]:
                    cell_bold = sheet[j]
                    cell_bold.font = Font(bold=True, color="FFFFFF")
                    cell_bold.fill = PatternFill(start_color="4d6c82", end_color="4d6c82", fill_type="solid")

                for k in [chr(i) + f"{j}" for i in range(65, 73) for j in range(1, 5)]:
                    sheet[k].alignment = Alignment(horizontal="left")

                global range_limit
                for k in [f"A{j}" for j in range(5, 40)]:
                    sheet[k].alignment = Alignment(horizontal="center")

                for k in [f"C{j}" for j in range(5, 40)]:
                    cell = sheet[k]
                    if cell.value == "Billable Days":
                        range_limit = int(re.findall(r"\d+", cell.coordinate)[0])
                        break
                    cell.alignment = Alignment(horizontal="center")

                for k in [i + f"{j}" for i in ["B", "D"] for j in range(5, 40)]:
                    cell = sheet[k]
                    if cell.coordinate == f"D{range_limit}":
                        cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="center")

            wb_style.save(excel_file_path)

        return [200, "Report Generated Successfully.", user_data, non_complaince_user, user_leave_record]

    except Exception as e:
        print(f"An error occurred: {str(e)}")
        exc_type, exc_obj, exc_tb = sys.exc_info()
        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        print(exc_type, fname, exc_tb.tb_lineno)
        return [500, str(e), None, None, None]


def non_compliance_resources(data, filename="non_complaint_user.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Data"
    headers = ["Name", "Month", "Listed Month Holiday", "Attendance Marked on Holiday"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        width = 40 if header == "Name" else max(len(header), 20)
        ws.column_dimensions[get_column_letter(col_num)].width = width

    for row_num, entry in enumerate(data, start=2):
        ws.cell(row=row_num, column=1, value=entry["Name"])
        ws.cell(row=row_num, column=2, value=entry["Month"])
        ws.cell(row=row_num, column=3, value=", ".join(entry["Listed Month Holiday"]))
        ws.cell(row=row_num, column=4, value=", ".join(entry["Attendance Marked on Holiday"]))
        for col_num in range(1, 5):
            ws.cell(row=row_num, column=col_num).alignment = Alignment(horizontal="center", vertical="center")

    wb.save(filename)


def add_summary_page(data, filename="my_workbook.xlsx"):
    """Add a 'Summary' sheet at index 0 with hyperlinked names and SUM totals."""
    wb = load_workbook(filename)
    sheet = wb.create_sheet("Summary", 0)
    sheet.sheet_properties.tabColor = "34b1eb"

    thin_side = Side(border_style="thin", color="000000")
    cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    headers = {2: "Name", 3: "Total Number of Billable Days", 4: "Leave Days"}
    for col_idx, header_text in headers.items():
        cell = sheet.cell(row=4, column=col_idx, value=header_text)
        cell.font = Font(bold=True, color="111212")
        cell.fill = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
        cell.alignment = Alignment(
            horizontal="left" if col_idx == 2 else "center", vertical="center"
        )
        cell.border = cell_border

    start_row = 5
    for idx, entry in enumerate(data, start=start_row):
        sheet_name = entry.get("Name", "")
        cell_B = sheet.cell(row=idx, column=2, value=sheet_name)
        if sheet_name:
            cell_B.hyperlink = f"#'{sheet_name}'!A1"
            cell_B.style = "Hyperlink"
            cell_B.font = Font(color="000000")
        cell_B.alignment = Alignment(horizontal="left", vertical="center")
        cell_B.border = cell_border

        cell_C = sheet.cell(row=idx, column=3, value=entry.get("Total Number of Billable Days", 0))
        cell_C.alignment = Alignment(horizontal="center", vertical="center")
        cell_C.border = cell_border

        cell_D = sheet.cell(row=idx, column=4, value=entry.get("Service Credit Pool Days", 0))
        cell_D.alignment = Alignment(horizontal="center", vertical="center")
        cell_D.border = cell_border

    last_data_row = start_row + len(data) - 1
    total_row = last_data_row + 1
    total_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    for col_num in range(2, 5):
        cell = sheet.cell(row=total_row, column=col_num)
        cell.fill = total_fill
        cell.border = cell_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    sheet.cell(row=total_row, column=2, value="Total")
    sheet.cell(row=total_row, column=3,
               value=f"=SUM({get_column_letter(3)}{start_row}:{get_column_letter(3)}{last_data_row})")
    sheet.cell(row=total_row, column=4,
               value=f"=SUM({get_column_letter(4)}{start_row}:{get_column_letter(4)}{last_data_row})")

    # Auto-fit column widths
    max_lengths = {col_idx: 0 for col_idx in headers}
    for col_idx, header_text in headers.items():
        max_lengths[col_idx] = max(max_lengths[col_idx], len(str(header_text)))
    for row_idx in range(start_row, last_data_row + 1):
        for col_idx in range(2, 5):
            val = sheet.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_lengths[col_idx] = max(max_lengths[col_idx], len(str(val)))

    for col_idx in range(2, 5):
        sheet.column_dimensions[get_column_letter(col_idx)].width = max_lengths[col_idx] + 2

    wb.save(filename)
