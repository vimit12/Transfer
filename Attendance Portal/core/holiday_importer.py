# ======================
# core/holiday_importer.py — Holiday file import logic
# ======================
import os
import json
import sqlite3
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook


def import_holidays_from_excel(file_path: str) -> tuple:
    """
    Parse an Excel / Numbers holiday file.
    Returns (excel_year: str, holidays: list[str]) or raises ValueError.

    File format:
      Row 1 col A : 4-digit year (int)
      Row 2+      : date cells (datetime, str in various formats)
    """
    if not file_path:
        raise ValueError("No file path provided.")

    file_extension = os.path.splitext(file_path)[1].lower()
    holidays = []
    excel_year = None
    DATE_FORMATS = ["%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d"]

    def _parse_date(val):
        if isinstance(val, datetime):
            return val
        if isinstance(val, str):
            for fmt in DATE_FORMATS:
                try:
                    return datetime.strptime(val, fmt)
                except ValueError:
                    continue
            raise ValueError(f"Unable to parse date: {val}")
        return datetime.strptime(str(val), "%Y-%m-%d %H:%M:%S")

    if file_extension in ('.xlsx', '.xls'):
        wb = load_workbook(filename=file_path)
        sheet = wb.active
        year_cell = sheet['A1'].value
        if not isinstance(year_cell, int) or len(str(year_cell)) != 4:
            raise ValueError("First cell must contain a 4-digit year (e.g., 2025)")
        excel_year = str(year_cell)

        for row in sheet.iter_rows(min_row=2, values_only=True):
            cell_value = row[0]
            if not cell_value:
                break
            date_obj = _parse_date(cell_value)
            if str(date_obj.year) != excel_year:
                raise ValueError(f"Date {date_obj.date()} doesn't match file year {excel_year}")
            formatted = date_obj.strftime("%d-%m-%Y")
            if formatted not in holidays:
                holidays.append(formatted)

    elif file_extension == '.numbers':
        try:
            try:
                df = pd.read_excel(file_path, header=None, engine='openpyxl')
            except Exception:
                df = pd.read_csv(file_path, header=None)

            if df.empty:
                raise ValueError("File appears to be empty")

            year_cell = df.iloc[0, 0]
            if not str(year_cell).isdigit() or len(str(year_cell)) != 4:
                raise ValueError("First cell must contain a 4-digit year")
            excel_year = str(year_cell)

            for idx in range(1, len(df)):
                cell_value = df.iloc[idx, 0]
                if pd.isna(cell_value):
                    continue
                if isinstance(cell_value, pd.Timestamp):
                    date_obj = cell_value.to_pydatetime()
                elif hasattr(cell_value, 'date'):
                    date_obj = cell_value
                else:
                    date_obj = _parse_date(cell_value)

                if str(date_obj.year) != excel_year:
                    raise ValueError(f"Date {date_obj.date()} doesn't match file year {excel_year}")
                formatted = date_obj.strftime("%d-%m-%Y")
                if formatted not in holidays:
                    holidays.append(formatted)
        except Exception as numbers_error:
            raise ValueError(
                f"Unable to read Numbers file directly. "
                f"Please export as .xlsx first.\nTechnical error: {numbers_error}"
            )
    else:
        raise ValueError(f"Unsupported file format: {file_extension}")

    if not holidays:
        raise ValueError("No valid holiday dates found in the file.")

    return excel_year, holidays


def save_holidays_to_db(conn, excel_year: str, holidays: list) -> bool:
    """
    Insert or replace the holiday list for *excel_year* in the DB.
    Returns True on success.
    """
    try:
        cursor = conn.cursor()
        cursor.execute(
            "INSERT OR REPLACE INTO holiday (year, holidays) VALUES (?, ?)",
            (excel_year, json.dumps(holidays))
        )
        conn.commit()
        cursor.close()
        return True
    except sqlite3.Error as e:
        print(f"Error saving holidays: {e}")
        conn.rollback()
        return False


def year_has_holidays(conn, year: str) -> bool:
    """Return True if the DB already has holidays for *year*."""
    try:
        cursor = conn.cursor()
        cursor.execute("SELECT year FROM holiday WHERE year = ?", (str(year),))
        result = cursor.fetchone()
        cursor.close()
        return result is not None
    except sqlite3.Error:
        return False
