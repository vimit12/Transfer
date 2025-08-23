import copy
import sys
from collections import Counter
from PyQt6.QtWidgets import (QApplication, QMainWindow, QWidget, QStackedWidget, QTableWidget, QGridLayout, QPushButton,
                             QLabel, QVBoxLayout, QTableWidgetItem, QFileDialog, QTextEdit, QGraphicsDropShadowEffect,
                             QFrame, QLineEdit, QComboBox, QFormLayout, QHeaderView, QDialog, QProgressBar, QMessageBox,
                             QSizePolicy, QHBoxLayout, QSpacerItem, QGroupBox, QPlainTextEdit, QScrollArea,
                             QAbstractItemView)
from PyQt6.QtGui import QFont, QIcon, QColor, QPalette, QAction
from PyQt6 import QtCore
from PyQt6.QtCore import Qt, QDate, QDateTime, QTimer
import sqlite3
import json
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
import pandas as pd
import numpy as np
import os
import re
from dateutil.parser import parse
import calendar
import itertools
from typing import Dict, List, Tuple, Optional, Any, Union
from pandas._libs.tslibs.timestamps import Timestamp
from pandas._libs.tslibs.nattype import NaTType
from openpyxl.utils import get_column_letter
from datetime import datetime

# ======================
# THEME DEFINITIONS
# ======================
DARK_THEME = """
QWidget {
    background-color: #2b2b2b;
    color: #e0e0e0;
    font-family: 'Segoe UI', Arial;
}

QMainWindow {
    background-color: #1e1e1e;
}

/* Titles and Labels */
QLabel#title {
    font-size: 20px;
    font-weight: 600;
    color: #4fc3f7;
    padding-bottom: 15px;
}

QLabel#subtitle {
    font-size: 14px;
    color: #90a4ae;
}

/* Buttons */
QPushButton {
    background-color: #37474f;
    border: 1px solid #455a64;
    border-radius: 6px;
    padding: 8px 16px;
    min-width: 100px;
    font-size: 12px;
    color: #ffffff;
    transition: all 0.2s ease;
}

QPushButton:hover {
    background-color: #455a64;
    border-color: #546e7a;
}

QPushButton:pressed {
    background-color: #263238;
}

QPushButton:disabled {
    background-color: #37474f;
    color: #90a4ae;
}

/* Special Buttons */
#generate_button {
    background-color: #0288d1;
    font-weight: 600;
    font-size: 14px;
}

#generate_button:hover {
    background-color: #039be5;
}

#generate_button:pressed {
    background-color: #0277bd;
}

/* Input Fields */
QLineEdit, QComboBox, QPlainTextEdit {
    background-color: #263238;
    border: 1px solid #37474f;
    border-radius: 6px;
    padding: 8px;
    color: #ffffff;
    selection-background-color: #4fc3f7;
    font-size: 12px;
}

QComboBox QAbstractItemView {
    background-color: #263238;
    color: #ffffff;
    selection-background-color: #0288d1;
    border: 1px solid #37474f;
}

/* Group Boxes */
QGroupBox {
    border: 1px solid #37474f;
    border-radius: 8px;
    margin-top: 16px;
    padding-top: 24px;
    font-size: 14px;
    font-weight: 500;
    color: #bb86fc;
    background-color: #1e1e1e;
}

QGroupBox::title {
    subcontrol-origin: margin;
    subcontrol-position: top left;
    left: 12px;
    padding: 0 8px;
}

/* Tables */
QTableWidget {
    background-color: #263238;
    color: #e0e0e0;
    border: 1px solid #37474f;
    gridline-color: #455a64;
    border-radius: 6px;
    font-size: 12px;
}

QHeaderView::section {
    background-color: #37474f;
    color: #ffffff;
    padding: 8px;
    font-weight: bold;
    border: none;
}

QTableWidget::item {
    color: #e0e0e0;
    padding: 8px;
}

QTableWidget::item:selected {
    background-color: #0288d1;
    color: #ffffff;
}

/* Progress Bar */
QProgressBar {
    border: 1px solid #37474f;
    border-radius: 6px;
    background-color: #263238;
    text-align: center;
    font-size: 12px;
    height: 20px;
}

QProgressBar::chunk {
    background-color: #0288d1;
    border-radius: 5px;
}

/* Cards */
.card {
    background-color: #263238;
    border: 1px solid #37474f;
    border-radius: 8px;
    padding: 20px;
    font-size: 16px;
    font-weight: 500;
    text-align: center;
    transition: all 0.3s ease;
}

.card:hover {
    background-color: #37474f;
    transform: translateY(-3px);
}

.card:pressed {
    background-color: #1e1e1e;
}

/* Status Messages */
#statusMsg[messageType="error"] {
    color: #ff5252;
    background-color: #2d1d1d;
    border: 1px solid #5c2b2b;
}

#statusMsg[messageType="success"] {
    color: #69f0ae;
    background-color: #1d2d24;
    border: 1px solid #2b5c40;
}

#statusMsg[messageType="info"] {
    color: #40c4ff;
    background-color: #1d2a2d;
    border: 1px solid #2b4d5c;
}

/* Sidebar */
#sidebar {
    background-color: #1e1e1e;
    border-right: 1px solid #37474f;
}

#sidebar QPushButton {
    text-align: left;
    padding: 12px 16px;
    border-radius: 0;
    border: none;
    border-left: 4px solid transparent;
}

#sidebar QPushButton:checked {
    background-color: #263238;
    border-left: 4px solid #0288d1;
}
"""

LIGHT_THEME = """
QWidget {
    background-color: #f5f7fa;
    color: #333333;
    font-family: 'Segoe UI', Arial;
}

QMainWindow {
    background-color: #ffffff;
}

/* Titles and Labels */
QLabel#title {
    font-size: 20px;
    font-weight: 600;
    color: #1565c0;
    padding-bottom: 15px;
}

QLabel#subtitle {
    font-size: 14px;
    color: #666666;
}

/* Buttons */
QPushButton {
    background-color: #e3f2fd;
    border: 1px solid #bbdefb;
    border-radius: 6px;
    padding: 8px 16px;
    min-width: 100px;
    font-size: 12px;
    color: #1976d2;
    transition: all 0.2s ease;
}

QPushButton:hover {
    background-color: #bbdefb;
    border-color: #90caf9;
}

QPushButton:pressed {
    background-color: #90caf9;
}

QPushButton:disabled {
    background-color: #e3f2fd;
    color: #90a4ae;
}

/* Special Buttons */
#generate_button {
    background-color: #1976d2;
    color: white;
    font-weight: 600;
    font-size: 14px;
}

#generate_button:hover {
    background-color: #1e88e5;
}

#generate_button:pressed {
    background-color: #1565c0;
}

/* Input Fields */
QLineEdit, QComboBox, QPlainTextEdit {
    background-color: #ffffff;
    border: 1px solid #bbdefb;
    border-radius: 6px;
    padding: 8px;
    color: #333333;
    selection-background-color: #4fc3f7;
    font-size: 12px;
}

QComboBox QAbstractItemView {
    background-color: #ffffff;
    color: #333333;
    selection-background-color: #bbdefb;
    border: 1px solid #bbdefb;
}

/* Group Boxes */
QGroupBox {
    border: 1px solid #bbdefb;
    border-radius: 8px;
    margin-top: 16px;
    padding-top: 24px;
    font-size: 14px;
    font-weight: 500;
    color: #1565c0;
    background-color: #ffffff;
}

QGroupBox::title {
    subcontrol-origin: margin;
    subcontrol-position: top left;
    left: 12px;
    padding: 0 8px;
}

/* Tables */
QTableWidget {
    background-color: #ffffff;
    color: #333333;
    border: 1px solid #bbdefb;
    gridline-color: #e3f2fd;
    border-radius: 6px;
    font-size: 12px;
}

QHeaderView::section {
    background-color: #e3f2fd;
    color: #1565c0;
    padding: 8px;
    font-weight: bold;
    border: none;
}

QTableWidget::item {
    color: #333333;
    padding: 8px;
}

QTableWidget::item:selected {
    background-color: #bbdefb;
    color: #333333;
}

/* Progress Bar */
QProgressBar {
    border: 1px solid #bbdefb;
    border-radius: 6px;
    background-color: #ffffff;
    text-align: center;
    font-size: 12px;
    height: 20px;
}

QProgressBar::chunk {
    background-color: #1976d2;
    border-radius: 5px;
}

/* Cards */
.card {
    background-color: #ffffff;
    border: 1px solid #bbdefb;
    border-radius: 8px;
    padding: 20px;
    font-size: 16px;
    font-weight: 500;
    text-align: center;
    transition: all 0.3s ease;
    color: #1976d2;
}

.card:hover {
    background-color: #e3f2fd;
    transform: translateY(-3px);
    box-shadow: 0 4px 8px rgba(0,0,0,0.1);
}

.card:pressed {
    background-color: #bbdefb;
}

/* Status Messages */
#statusMsg[messageType="error"] {
    color: #d32f2f;
    background-color: #ffebee;
    border: 1px solid #ffcdd2;
}

#statusMsg[messageType="success"] {
    color: #388e3c;
    background-color: #e8f5e9;
    border: 1px solid #c8e6c9;
}

#statusMsg[messageType="info"] {
    color: #0288d1;
    background-color: #e3f2fd;
    border: 1px solid #bbdefb;
}

/* Sidebar */
#sidebar {
    background-color: #f5f7fa;
    border-right: 1px solid #bbdefb;
}

#sidebar QPushButton {
    text-align: left;
    padding: 12px 16px;
    border-radius: 0;
    border: none;
    border-left: 4px solid transparent;
}

#sidebar QPushButton:checked {
    background-color: #e3f2fd;
    border-left: 4px solid #1976d2;
}
"""


# ======================
# UTILITY FUNCTIONS
# ======================
def clean_date(value: Any) -> Optional[str]:
    """Convert NaT to None and Timestamps to string format for database insertion"""
    if isinstance(value, NaTType):
        return None
    if isinstance(value, pd.Timestamp):
        return value.strftime("%d-%m-%Y")
    return str(value) if value is not None else None


def clean_string(s: str) -> str:
    """Clean and normalize strings for comparison"""
    return re.sub(r"\[.*?\]|[, -]", " ", s).lower().strip()


def coverage_percentage(str1: str, str2: str) -> float:
    """Calculate word coverage percentage between two strings"""
    words1 = set(clean_string(str1).split())
    words2 = set(clean_string(str2).split())
    common_words = words1 & words2
    max_len = max(len(words1), len(words2))
    return len(common_words) / max_len * 100 if max_len > 0 else 0


def read_file(file_path: str) -> Optional[List[Dict]]:
    """Read CSV or Excel file into a list of dictionaries"""
    try:
        if file_path.endswith(".csv"):
            df = pd.read_csv(file_path)
        elif file_path.endswith((".xlsx", ".xls")):
            df = pd.read_excel(file_path)
        else:
            raise ValueError("Unsupported file format. Please provide a CSV or Excel file.")
        return df.to_dict("records")
    except Exception as e:
        print(f"Error reading file: {e}")
        return None


def sort_list_of_dicts(data: List[Dict]) -> List[Dict]:
    """Sort list of dictionaries with 'Total' at the end"""
    total_dicts = [d for d in data if d.get('Name') == 'Total']
    sorted_data = sorted([d for d in data if d.get('Name') != 'Total'], key=lambda x: x['Name'])
    return sorted_data + total_dicts


def get_month_details(month_name: str, year: int) -> Tuple[List[List[Dict]], int]:
    """Get detailed calendar information for a specific month"""
    month_number = list(calendar.month_name).index(month_name.capitalize())
    cal = calendar.monthcalendar(year, month_number)
    weekdays = {0: "Monday", 1: "Tuesday", 2: "Wednesday", 3: "Thursday", 4: "Friday", 5: "Saturday", 6: "Sunday"}

    month_details = []
    for week in cal:
        week_details = []
        for day in week:
            if day == 0:
                week_details.append(None)
            else:
                day_name = weekdays[calendar.weekday(year, month_number, day)]
                is_weekend = day_name in ["Saturday", "Sunday"]
                week_details.append({"day": day, "day_name": day_name, "is_weekend": is_weekend})
        month_details.append(week_details)
    return month_details, month_number


def date_calculation(date: Union[str, datetime]) -> Tuple[int, int, int]:
    """Extract day, month, and year from a date"""
    if isinstance(date, datetime):
        return date.day, date.month, date.year
    date_obj = datetime.strptime(date, "%d-%m-%Y")
    return date_obj.day, date_obj.month, date_obj.year


def preprocess_name(input_str: str) -> str:
    """Preprocess names for consistent comparison"""
    return "".join(sorted(input_str.replace(",", "").lower().split()))


def get_details_for_name(name: str, name_mapping: Dict) -> Optional[Tuple]:
    """Get details for a name from the mapping dictionary"""
    clean_name = preprocess_name(name)
    for key in name_mapping:
        if coverage_percentage(name, preprocess_name(key)) == 100:
            return name_mapping[key]
    return None


def sanitize_sheet_name(name: str, default: str = "Sheet1") -> str:
    """Sanitize Excel sheet names"""
    if not name:
        return default
    for char in ['\\', '/', '*', '?', ':', '[', ']']:
        name = name.replace(char, '')
    return name[:31]


def format_date(date_str: Union[str, Timestamp]) -> str:
    """Format dates consistently"""
    if isinstance(date_str, Timestamp):
        date_str = date_str.strftime("%Y-%m-%d")
    try:
        date_obj = parse(date_str)
        return date_obj.strftime("%d-%m-%Y")
    except (ValueError, TypeError):
        return str(date_str)


def sizeof_fmt(num: int, suffix: str = 'B') -> str:
    """Convert bytes to human-readable format"""
    for unit in ['', 'K', 'M', 'G']:
        if abs(num) < 1024.0:
            return f"{num:3.1f}{unit}{suffix}"
        num /= 1024.0
    return f"{num:.1f}Yi{suffix}"


# ======================
# CUSTOM WIDGETS
# ======================
class CardButton(QPushButton):
    """Custom card-style button widget"""

    def __init__(self, icon: str, text: str, parent=None):
        super().__init__(parent)
        self.setProperty("class", "card")
        layout = QVBoxLayout()
        layout.setAlignment(Qt.AlignmentFlag.AlignCenter)

        icon_label = QLabel(icon)
        icon_label.setFont(QFont("Arial", 24))
        text_label = QLabel(text)
        text_label.setFont(QFont("Segoe UI", 10))
        text_label.setWordWrap(True)

        layout.addWidget(icon_label)
        layout.addWidget(text_label)
        self.setLayout(layout)

    def sizeHint(self):
        return QtCore.QSize(180, 140)


class ModernTableWidget(QTableWidget):
    """Enhanced table widget with better styling and features"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setAlternatingRowColors(True)
        self.setShowGrid(False)
        self.verticalHeader().setVisible(False)
        self.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.horizontalHeader().setStretchLastSection(True)
        self.setSortingEnabled(True)
        self.setWordWrap(False)

        # Add subtle shadow effect
        shadow = QGraphicsDropShadowEffect()
        shadow.setBlurRadius(10)
        shadow.setOffset(2, 2)
        shadow.setColor(QColor(0, 0, 0, 50))
        self.setGraphicsEffect(shadow)

    def load_data(self, data: List[Dict], headers: List[str]):
        """Load data into the table"""
        self.clear()
        self.setColumnCount(len(headers))
        self.setHorizontalHeaderLabels(headers)
        self.setRowCount(len(data))

        for row_idx, row_data in enumerate(data):
            for col_idx, header in enumerate(headers):
                value = str(row_data.get(header, ""))
                item = QTableWidgetItem(value)
                self.setItem(row_idx, col_idx, item)

        # Adjust column widths
        for col in range(self.columnCount()):
            self.resizeColumnToContents(col)

        # Apply zebra striping
        for row in range(self.rowCount()):
            if row % 2 == 0:
                for col in range(self.columnCount()):
                    item = self.item(row, col)
                    if item:
                        item.setBackground(
                            QColor(240, 240, 240) if self.parent().current_theme == "Light" else QColor(45, 45, 45))


# ======================
# MAIN APPLICATION
# ======================
class MainWindow(QMainWindow):
    """Main application window for the Billing Report Generator"""

    def __init__(self):
        super().__init__()
        self.setWindowTitle("Billing Report Generator")
        self.setMinimumSize(1200, 700)
        self.current_theme = "Light"
        self.themes = ["Dark", "Light"]
        self.db_connection = None
        self.current_table = None
        self.current_year = str(datetime.now().year)
        self.raw_category_list = []
        self.name_order_list = []
        self.categories = {}
        self.name_mapping = {}
        self.HOLIDAY_LIST = []
        self.df = None
        self.total_working_days = 0

        self.init_ui()
        self.initialize_database()

    def init_ui(self):
        """Initialize the user interface"""
        self.setup_theme()
        self.setup_menu()
        self.setup_main_layout()

    def setup_theme(self):
        """Set initial theme"""
        app = QApplication.instance()
        app.setFont(QFont("Segoe UI", 10))
        app.setStyleSheet(LIGHT_THEME)

    def setup_menu(self):
        """Create the application menu"""
        menubar = self.menuBar()

        # File menu
        file_menu = menubar.addMenu("&File")

        new_action = QAction("New", self)
        open_action = QAction("Open", self)
        save_action = QAction("Save", self)
        exit_action = QAction("Exit", self)
        exit_action.triggered.connect(self.close)

        file_menu.addAction(new_action)
        file_menu.addAction(open_action)
        file_menu.addAction(save_action)
        file_menu.addSeparator()
        file_menu.addAction(exit_action)

        # Edit menu
        edit_menu = menubar.addMenu("&Edit")
        edit_menu.addAction("Preferences")

        # Help menu
        help_menu = menubar.addMenu("&Help")
        help_menu.addAction("Documentation")
        about_action = QAction("About", self)
        about_action.triggered.connect(lambda: self.switch_page(4))
        help_menu.addAction(about_action)

    def setup_main_layout(self):
        """Set up the main layout with sidebar and content area"""
        # Central widget and main layout
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QHBoxLayout(central_widget)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)

        # Sidebar
        self.setup_sidebar()
        main_layout.addWidget(self.sidebar)

        # Content area
        content_area = QFrame()
        content_layout = QVBoxLayout(content_area)
        content_layout.setContentsMargins(0, 0, 0, 0)

        # Title bar
        title_bar = QWidget()
        title_bar.setFixedHeight(40)
        title_bar_layout = QHBoxLayout(title_bar)
        title_bar_layout.setContentsMargins(20, 0, 20, 0)

        # Theme toggle
        theme_container = QWidget()
        theme_layout = QHBoxLayout(theme_container)
        theme_layout.setContentsMargins(0, 0, 0, 0)

        self.theme_btn = QPushButton()
        self.theme_btn.setFixedHeight(32)
        self.theme_btn.clicked.connect(self.cycle_theme)
        self.update_theme_button()

        theme_layout.addWidget(self.theme_btn)
        title_bar_layout.addWidget(QLabel("Billing Report Generator"))
        title_bar_layout.addStretch()
        title_bar_layout.addWidget(theme_container)

        # Stacked widget for pages
        self.stacked_widget = QStackedWidget()

        content_layout.addWidget(title_bar)
        content_layout.addWidget(self.stacked_widget)

        # Initialize pages
        self.init_pages()

        main_layout.addWidget(content_area, 1)

    def setup_sidebar(self):
        """Set up the sidebar navigation"""
        self.sidebar = QFrame()
        self.sidebar.setObjectName("sidebar")
        self.sidebar.setFixedWidth(200)
        sidebar_layout = QVBoxLayout(self.sidebar)
        sidebar_layout.setContentsMargins(0, 20, 0, 20)
        sidebar_layout.setSpacing(2)

        # Navigation buttons
        nav_buttons = [("🏠 Home", "Home"), ("📁 Database", "Database"), ("📊 Load Data", "Load Data"),
            ("📈 Spreadsheet", "Spreadsheet"), ("ℹ️ About", "About")]

        self.nav_buttons = []
        for icon, text in nav_buttons:
            btn = QPushButton(f"{icon} {text}")
            btn.setCheckable(True)
            btn.setCursor(Qt.CursorShape.PointingHandCursor)
            sidebar_layout.addWidget(btn)
            self.nav_buttons.append(btn)

        # Connect navigation
        self.nav_buttons[0].clicked.connect(lambda: self.switch_page(0))
        self.nav_buttons[1].clicked.connect(lambda: self.switch_page(1))
        self.nav_buttons[2].clicked.connect(lambda: self.switch_page(2))
        self.nav_buttons[3].clicked.connect(lambda: self.switch_page(3))
        self.nav_buttons[4].clicked.connect(lambda: self.switch_page(4))

        sidebar_layout.addStretch()

        # App info
        app_info = QLabel("Billing Report Generator\nv1.3.0")
        app_info.setAlignment(Qt.AlignmentFlag.AlignCenter)
        app_info.setStyleSheet("color: #90a4ae; font-size: 10px;")
        sidebar_layout.addWidget(app_info)

    def init_pages(self):
        """Initialize all application pages"""
        self.stacked_widget.addWidget(self.create_home_page())
        self.stacked_widget.addWidget(self.create_database_page())
        self.stacked_widget.addWidget(self.create_load_data_page())
        self.stacked_widget.addWidget(self.create_spreadsheet_page())
        self.stacked_widget.addWidget(self.create_about_page())

        # Set home page as default
        self.switch_page(0)
        self.nav_buttons[0].setChecked(True)

    def show_format_guide(self):
        """Show Excel format requirements in a modern dialog"""
        guide_text = """
            <div style="font-family: 'Segoe UI', Arial, sans-serif; max-width: 600px;">
                <h2 style="color: #1976d2; margin-top: 0;">Excel Format Requirements</h2>

                <div style="background-color: #e3f2fd; border-radius: 8px; padding: 15px; margin-bottom: 20px;">
                    <p style="margin-top: 0;">Follow these guidelines for holiday data files:</p>
                    <ol style="margin-bottom: 0;">
                        <li>First cell must contain the <b>4-digit year</b> (e.g., 2025)</li>
                        <li>Subsequent cells should contain dates in any valid Excel date format</li>
                        <li>Dates must belong to the specified year</li>
                        <li>Save file as .xlsx or .xls format</li>
                    </ol>
                </div>

                <h3 style="color: #1976d2;">Example Format</h3>
                <div style="overflow-x: auto;">
                    <table style="border-collapse: collapse; width: 100%;">
                        <thead>
                            <tr style="background-color: #bbdefb;">
                                <th style="padding: 10px; text-align: left; border: 1px solid #90caf9;">Year</th>
                            </tr>
                        </thead>
                        <tbody>
                            <tr>
                                <td style="padding: 10px; border: 1px solid #e0e0e0;">2025</td>
                            </tr>
                            <tr style="background-color: #f5f5f5;">
                                <td style="padding: 10px; border: 1px solid #e0e0e0;">01-Jan-2025</td>
                            </tr>
                            <tr>
                                <td style="padding: 10px; border: 1px solid #e0e0e0;">15-Jan-2025</td>
                            </tr>
                            <tr style="background-color: #f5f5f5;">
                                <td style="padding: 10px; border: 1px solid #e0e0e0;">26-Jan-2025</td>
                            </tr>
                            <tr>
                                <td style="padding: 10px; border: 1px solid #e0e0e0;">... etc ...</td>
                            </tr>
                        </tbody>
                    </table>
                </div>

                <div style="margin-top: 20px; background-color: #fff8e1; border-left: 4px solid #ffc107; padding: 10px;">
                    <p style="margin: 0;"><b>Tip:</b> You can use any date format recognized by Excel (DD-MM-YYYY, MM/DD/YYYY, etc.)</p>
                </div>
            </div>
        """

        msg_box = QMessageBox(self)
        msg_box.setWindowTitle("Holiday Format Guide")
        msg_box.setTextFormat(Qt.TextFormat.RichText)
        msg_box.setText(guide_text)

        # Add a download example button
        download_btn = msg_box.addButton("Download Example", QMessageBox.ButtonRole.ActionRole)
        download_btn.setIcon(QIcon.fromTheme("document-save"))
        download_btn.clicked.connect(self.download_example_format)

        msg_box.setStandardButtons(QMessageBox.StandardButton.Close)
        msg_box.exec()

    def download_example_format(self):
        """Download an example holiday format file"""
        file_path, _ = QFileDialog.getSaveFileName(self, "Save Example File", "holiday_format_example.xlsx",
            "Excel Files (*.xlsx)")

        if not file_path:
            return

        try:
            # Create example workbook
            wb = Workbook()
            ws = wb.active

            # Add sample data
            ws['A1'] = "2025"
            ws['A2'] = "01-01-2025"
            ws['A3'] = "15-01-2025"
            ws['A4'] = "26-01-2025"
            ws['A5'] = "15-08-2025"
            ws['A6'] = "02-10-2025"
            ws['A7'] = "25-12-2025"

            # Save file
            wb.save(file_path)
            self.show_message("Example file saved successfully!", "success", 3000)

        except Exception as e:
            self.show_message(f"Error saving example: {str(e)}", "error", 5000)

    def create_home_page(self) -> QWidget:
        """Create the home page with report generation controls"""
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(30, 20, 30, 20)
        layout.setSpacing(20)

        # Header
        header = QLabel("Generate Billing Reports")
        header.setObjectName("title")
        header.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(header)

        # Configuration container
        config_container = QWidget()
        config_layout = QVBoxLayout(config_container)
        config_layout.setContentsMargins(0, 0, 0, 0)
        config_layout.setSpacing(20)

        # Date selection
        date_group = self.create_date_selection_group()
        config_layout.addWidget(date_group)

        # Holiday management
        holiday_group = self.create_holiday_management_group()
        config_layout.addWidget(holiday_group)

        # Category management
        category_group = self.create_category_management_group()
        config_layout.addWidget(category_group)

        # File upload
        file_group = self.create_file_upload_group()
        config_layout.addWidget(file_group)

        layout.addWidget(config_container, 1)

        # Status and progress
        status_group = self.create_status_progress_group()
        layout.addWidget(status_group)

        return page

    def create_date_selection_group(self) -> QGroupBox:
        """Create date selection group"""
        group = QGroupBox("Select Month & Year")
        layout = QHBoxLayout(group)
        layout.setContentsMargins(15, 25, 15, 15)
        layout.setSpacing(15)

        # Month selection
        self.month_combo = QComboBox()
        self.month_combo.addItems(
            ["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November",
                "December"])
        current_month = QDate.currentDate().month()
        self.month_combo.setCurrentIndex(current_month - 1)

        # Year selection
        self.year_combo = QComboBox()
        current_year = QDate.currentDate().year()
        self.year_combo.addItems([str(y) for y in range(2022, 2051)])
        self.year_combo.setCurrentText(str(current_year))

        # Disable future years
        for idx in range(self.year_combo.count()):
            year = int(self.year_combo.itemText(idx))
            self.year_combo.model().item(idx).setEnabled(year <= current_year)

        # Add to layout
        layout.addWidget(QLabel("Month:"), 1)
        layout.addWidget(self.month_combo, 3)
        layout.addWidget(QLabel("Year:"), 1)
        layout.addWidget(self.year_combo, 2)

        return group

    def load_holidays_to_db(self):
        """Handle holiday loading with improved error handling"""
        file_path, _ = QFileDialog.getOpenFileName(self, "Select Holiday File", "",
            "Spreadsheet Files (*.xlsx *.xls *.numbers);;All Files (*)")

        if not file_path:
            return

        try:
            # Update UI with file info
            file_name = os.path.basename(file_path)
            file_size = os.path.getsize(file_path)
            self.holiday_input.setPlainText(f"{file_name} ({sizeof_fmt(file_size)})")

            # Read file
            if file_path.endswith(".csv"):
                df = pd.read_csv(file_path, header=None)
            else:
                df = pd.read_excel(file_path, header=None)

            # Extract year
            excel_year = str(df.iloc[0, 0])
            if not excel_year.isdigit() or len(excel_year) != 4:
                raise ValueError("First cell must contain a 4-digit year")

            # Extract holidays
            holidays = []
            for idx in range(1, len(df)):
                date_val = df.iloc[idx, 0]
                if pd.isna(date_val):
                    continue

                if isinstance(date_val, str):
                    date_obj = parse(date_val)
                elif isinstance(date_val, pd.Timestamp):
                    date_obj = date_val.to_pydatetime()
                else:
                    date_obj = pd.to_datetime(date_val).to_pydatetime()

                if str(date_obj.year) != excel_year:
                    raise ValueError(f"Date {date_obj.date()} doesn't match file year {excel_year}")

                formatted_date = date_obj.strftime("%d-%m-%Y")
                if formatted_date not in holidays:
                    holidays.append(formatted_date)

            # Save to database
            self.save_holidays_to_db(excel_year, holidays)
            self.show_message(f"Loaded {len(holidays)} holidays for {excel_year}", "success", 5000)

        except Exception as e:
            self.show_message(f"Error: {str(e)}", "error", 5000)

    def show_holiday_viewer(self):
        """Show holiday viewer dialog with table display"""
        dialog = QDialog(self)
        dialog.setWindowTitle("Holiday List")
        dialog.setMinimumSize(500, 400)  # Increased width and height for better visibility

        layout = QVBoxLayout(dialog)

        # Year selection
        year_layout = QHBoxLayout()
        year_label = QLabel("Select Year:")

        self.viewer_year_combo = QComboBox()
        self.viewer_year_combo.setFixedWidth(100)  # Increased width for better readability
        self.viewer_year_combo.addItems([str(y) for y in range(2024, 2051)])
        # Set the current year as the default selection
        current_year = str(QDate.currentDate().year())
        self.viewer_year_combo.setCurrentText(current_year)

        year_layout.addWidget(year_label)
        year_layout.addWidget(self.viewer_year_combo)
        year_layout.addStretch()

        # Table setup with scroll area
        self.holiday_table = QTableWidget()
        self.holiday_table.setColumnCount(2)
        self.holiday_table.setHorizontalHeaderLabels(["Date", "Day"])
        self.holiday_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)

        # Make table cells uneditable
        self.holiday_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)

        # Wrap the table in a scroll area
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)

        table_container = QWidget()
        table_layout = QVBoxLayout(table_container)
        table_layout.addWidget(self.holiday_table)
        scroll_area.setWidget(table_container)

        # Refresh button
        refresh_btn = QPushButton("Refresh")
        refresh_btn.setFixedHeight(35)  # Adjust button height for a better look
        refresh_btn.clicked.connect(lambda: self.populate_holiday_table(self.viewer_year_combo.currentText()))

        layout.addLayout(year_layout)
        layout.addWidget(scroll_area)  # Add scrollable table
        layout.addWidget(refresh_btn)

        # Initial population
        self.viewer_year_combo.currentTextChanged.connect(self.populate_holiday_table)
        self.populate_holiday_table(self.viewer_year_combo.currentText())

        dialog.exec()
    def create_holiday_management_group(self) -> QGroupBox:
        """Create holiday management group"""
        group = QGroupBox("Holiday Management")
        layout = QVBoxLayout(group)
        layout.setContentsMargins(15, 25, 15, 15)
        layout.setSpacing(10)

        # Input field
        self.holiday_input = QPlainTextEdit()
        self.holiday_input.setFixedHeight(40)
        self.holiday_input.setReadOnly(True)

        # Button row
        button_layout = QHBoxLayout()
        button_layout.setSpacing(10)

        self.format_info_btn = QPushButton("Format Guide")
        self.format_info_btn.setObjectName("format_info_btn")
        self.format_info_btn.setIcon(QIcon.fromTheme("help"))
        self.format_info_btn.clicked.connect(self.show_format_guide)

        self.load_holiday_btn = QPushButton("Load Holidays")
        self.load_holiday_btn.setObjectName("load_holiday_btn")
        self.load_holiday_btn.setIcon(QIcon.fromTheme("document-open"))
        self.load_holiday_btn.clicked.connect(self.load_holidays_to_db)

        self.view_holiday_btn = QPushButton("View Holidays")
        self.view_holiday_btn.setObjectName("view_holiday_btn")
        self.view_holiday_btn.setIcon(QIcon.fromTheme("view-list"))
        self.view_holiday_btn.clicked.connect(self.show_holiday_viewer)

        button_layout.addWidget(self.format_info_btn)
        button_layout.addWidget(self.load_holiday_btn)
        button_layout.addWidget(self.view_holiday_btn)

        # Error label
        self.holiday_error_label = QLabel()
        self.holiday_error_label.setStyleSheet("color: #ff5252; font-size: 11px;")
        self.holiday_error_label.setWordWrap(True)
        self.holiday_error_label.hide()

        # Add to layout
        layout.addWidget(QLabel("Holiday File:"))
        layout.addWidget(self.holiday_input)
        layout.addLayout(button_layout)
        layout.addWidget(self.holiday_error_label)

        return group

    def select_category(self):
        file_dialog = QFileDialog(self)
        filepath, _ = file_dialog.getOpenFileName(self, "Open Excel File", "", "Excel Files (*.xlsx *.xls, *.csv)")

        self.raw_category_list, self.name_order_list = ([],) * 2
        self.categories = dict()
        self.name_mapping = dict()

        try:
            if filepath:
                fileInfo = QtCore.QFileInfo(filepath)
                file_name = fileInfo.fileName()
                file_size = fileInfo.size()  # in bytes
                # Convert file size to kilobytes
                file_size_kb = file_size / 1024.0
                print(f"File Name: {file_name}, File Size: {file_size_kb:.2f} KB")

                self.category_input.setPlainText(f"{file_name} ({file_size_kb:.2f} KB)")

                sheet_name = "PublicCloudResourceList"
                # Read the Excel file into a DataFrame
                df = pd.read_excel(filepath, sheet_name=sheet_name)
                df.replace({np.nan: None}, inplace=True)
                self.add_data_resource_tab(df)
                # Convert the DataFrame to a list of dictionaries
                self.raw_category_list = df.to_dict(orient="records")

                # Standardize the "Full Name" field by removing commas and spaces
                for item in self.raw_category_list:
                    name = clean_string(item['Full Name'])
                    team = item["Team"]

                    # Check if the team is already in the dictionary
                    if team in self.categories:
                        self.categories[team].append(name)
                    else:
                        self.categories[team] = [name]

                    self.name_mapping.update(
                        {name: [item["521 ID"], item["Point of Contact"], item["Start Date"], item["End Date"]]})

                # Create a mapping from Full Name to 521 ID for quick lookup
                # self.name_mapping = {
                #     preprocess_name(item["Full Name"]): [item["521 ID"], item["Point of Contact"], item["Start Date"],
                #         item["End Date"], ] for item in self.raw_category_list}

                for k, v in self.categories.items():
                    temp_list = sorted(v)
                    self.categories[k] = temp_list
                    self.name_order_list.extend(temp_list)
                # QMessageBox.information(self, "Success", "Category data successfully created!",
                #     QMessageBox.StandardButton.Ok)
                self.show_message("Category data successfully created!", "success", 3000)

        except Exception as e:
            self.raw_category_list, self.name_order_list = [], []  # ✅ Separate lists
            self.categories, self.name_mapping = {}, {}  # ✅ Separate dictionaries
            self.category_input.setPlainText(f"")
            self.show_message(f"Error: Not a valid category file", "error", 5000)

    def upload_file(self):
        file_dialog = QFileDialog(self)
        filepath, _ = file_dialog.getOpenFileName(self, "Open Excel File", "", "Excel Files (*.xlsx *.xls, *.csv)")

        if filepath:
            fileInfo = QtCore.QFileInfo(filepath)
            file_name = fileInfo.fileName()
            file_size = fileInfo.size()  # in bytes
            # Convert file size to kilobytes
            file_size_kb = file_size / 1024.0
            print(f"File Name: {file_name}, File Size: {file_size_kb:.2f} KB")
            self.main_input.setPlainText(f"{file_name} ({file_size_kb:.2f} KB)")

            # Create DataFrame with Pandas
            self.df = read_file(filepath)

            # Adding Validation to file upload
            try:
                dict_value = dict(Counter(list(itertools.chain.from_iterable(
                    [[item.split("-")[-1] for item in j] for j in [list(i.keys())[4:-2] for i in self.df]]))))
                value = max(dict_value, key=dict_value.get)

                self.selected_month = self.month_combo.currentText()

                if value == self.selected_month[:3]:
                    self.show_message("Valid Raw Excel Loaded", "info", 4000)
                else:
                    self.show_message("Invalid Raw excel, please check the file or selected month.", "error", 5000)
                    self.main_input.setPlainText("")
                    self.df = None  # # Update progress  # self.progress_bar.setValue(0)  # 0-100%  #  # # Show/hide when needed  # self.progress_bar.setVisible(True)

                # Reset on completion  # self.progress_bar.reset()
            except Exception as e:
                print(e)
                self.show_message("Invalid Raw excel, please check the file.", "error", 5000)
                self.df = None
    def show_table_contents(self, table_name):
        self.current_table = table_name
        """Display contents of selected table from dropdown with enhanced UI"""
        if not table_name:
            return

        try:
            cursor = self.db_connection.cursor()
            cursor.execute(f"PRAGMA table_info({table_name})")
            columns = [col[1] for col in cursor.fetchall()]

            cursor.execute(f"SELECT * FROM {table_name}")
            rows = cursor.fetchall()
            self.table_view.clear()  #reset the view in every table selected
            # Configure table view with modern styling
            self.table_view.setRowCount(len(rows))
            self.table_view.setColumnCount(len(columns) + 1)
            self.table_view.setHorizontalHeaderLabels(columns + ["Actions"])

            # Make table cells uneditable
            self.table_view.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)

            # Enhanced table styling
            self.table_view.setStyleSheet("""
                QTableWidget {
                    background-color: #ffffff;
                    border: 1px solid #e0e0e0;
                    border-radius: 8px;
                    gridline-color: #f0f0f0;
                    selection-background-color: #e3f2fd;
                    font-size: 12px;
                    font-family: 'Segoe UI', Arial, sans-serif;
                }
                QTableWidget::item {
                    padding: 12px 8px;
                    border-bottom: 1px solid #f5f5f5;
                }
                QTableWidget::item:selected {
                    background-color: #e3f2fd;
                    color: #1976d2;
                }
                QTableWidget::item:hover {
                    background-color: #f8f9fa;
                }
                QHeaderView::section {
                    background-color: #f8f9fa;
                    color: #424242;
                    font-weight: bold;
                    font-size: 11px;
                    text-transform: uppercase;
                    letter-spacing: 0.5px;
                    padding: 12px 8px;
                    border: none;
                    border-bottom: 2px solid #e0e0e0;
                    border-right: 1px solid #e0e0e0;
                }
                QHeaderView::section:first {
                    border-top-left-radius: 8px;
                }
                QHeaderView::section:last {
                    border-top-right-radius: 8px;
                    border-right: none;
                }
                QScrollBar:vertical {
                    border: none;
                    background: #f5f5f5;
                    width: 12px;
                    border-radius: 6px;
                }
                QScrollBar::handle:vertical {
                    background: #c0c0c0;
                    border-radius: 6px;
                    min-height: 20px;
                }
                QScrollBar::handle:vertical:hover {
                    background: #a0a0a0;
                }
                QScrollBar:horizontal {
                    border: none;
                    background: #f5f5f5;
                    height: 12px;
                    border-radius: 6px;
                }
                QScrollBar::handle:horizontal {
                    background: #c0c0c0;
                    border-radius: 6px;
                    min-width: 20px;
                }
                QScrollBar::handle:horizontal:hover {
                    background: #a0a0a0;
                }
                QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal {
                    width: 0;
                }
            """)

            # Enable word wrapping for all items
            self.table_view.setWordWrap(True)

            # Set row height to auto-adjust based on content
            self.table_view.verticalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
            self.table_view.verticalHeader().setDefaultSectionSize(80)  # Minimum row height
            # self.table_view.verticalHeader().hide()  # Hide row numbers for cleaner look

            # Set alternating row colors
            self.table_view.setAlternatingRowColors(True)

            # Clear existing filters with animation-like effect
            while self.filter_layout.count():
                if child := self.filter_layout.takeAt(0):
                    if widget := child.widget():
                        widget.deleteLater()

            # Enhanced filter section with proper alignment
            filter_container = QWidget()
            filter_container.setStyleSheet("""
                QWidget {
                    background-color: #f8f9fa;
                    border-radius: 8px;
                    margin: 4px;
                    padding: 8px;
                }
            """)

            # Create a grid layout for filters to align with table columns
            filter_grid_layout = QGridLayout(filter_container)
            filter_grid_layout.setSpacing(4)
            filter_grid_layout.setContentsMargins(12, 8, 12, 8)

            # Add filter label
            filter_label = QLabel("🔍 Filters:")
            filter_label.setStyleSheet("""
                QLabel {
                    color: #424242;
                    font-weight: bold;
                    font-size: 12px;
                    background: none;
                    padding: 0;
                }
            """)
            filter_grid_layout.addWidget(filter_label, 0, 0, 1, len(columns) + 1)

            # Add enhanced filter inputs aligned with columns
            self.filter_inputs = []
            for col_idx, col_name in enumerate(columns):
                filter_edit = QLineEdit()
                filter_edit.setPlaceholderText(col_name)
                filter_edit.setStyleSheet("""
                    QLineEdit {
                        background-color: white;
                        border: 2px solid #e0e0e0;
                        border-radius: 6px;
                        padding: 6px 8px;
                        font-size: 10px;
                        min-height: 20px;
                    }
                    QLineEdit:focus {
                        border-color: #2196f3;
                        background-color: #fafafa;
                    }
                    QLineEdit:hover {
                        border-color: #bdbdbd;
                    }
                """)
                filter_edit.textChanged.connect(self.apply_filters)
                filter_grid_layout.addWidget(filter_edit, 1, col_idx)
                self.filter_inputs.append(filter_edit)

            # Add clear filters button in the actions column
            clear_filters_btn = QPushButton("✖️")
            clear_filters_btn.setFixedSize(38, 38)
            clear_filters_btn.setStyleSheet("""
                QPushButton {
                    background-color: #ff9800;
                    color: white;
                    border: none;
                    border-radius: 14px;
                    font-weight: bold;
                    font-size: 12px;
                }
                QPushButton:hover {
                    background-color: #f57c00;
                }
                QPushButton:pressed {
                    background-color: #ef6c00;
                }
            """)
            clear_filters_btn.setToolTip("Clear All Filters")
            clear_filters_btn.clicked.connect(self.clear_all_filters)
            filter_grid_layout.addWidget(clear_filters_btn, 1, len(columns))

            # Add the filter container to the main layout
            self.filter_layout.addWidget(filter_container)

            # Configure columns with better spacing and horizontal scrolling
            header = self.table_view.horizontalHeader()

            # Enable horizontal scrolling when needed
            self.table_view.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
            self.table_view.setHorizontalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)

            # Action column with proper width calculation
            if len(columns) <= 3:
                # Set all columns to have max width and word wrap
                for col in range(len(columns)):
                    # Set max width to 250 (adjust as needed) with word wrap
                    header.setSectionResizeMode(col, QHeaderView.ResizeMode.Interactive)
                    self.table_view.setColumnWidth(col, 370)  # Reduced from 250 to make room for actions
                    header.setMaximumSectionSize(400)  # Max width for content columns
                action_column_width = 80  # Width for two buttons + spacing
            else:
                # Set all columns to have max width and word wrap
                for col in range(len(columns)):
                    # Set max width to 250 (adjust as needed) with word wrap
                    header.setSectionResizeMode(col, QHeaderView.ResizeMode.Interactive)
                    self.table_view.setColumnWidth(col, 200)  # Reduced from 250 to make room for actions
                    header.setMaximumSectionSize(300)  # Max width for content columns
                action_column_width = 210  # Width for two buttons + spacing

            header.setSectionResizeMode(len(columns), QHeaderView.ResizeMode.Fixed)
            self.table_view.setColumnWidth(len(columns), action_column_width)
            # Set minimum width for the action column header
            header.setMinimumSectionSize(action_column_width)

            # Populate data with enhanced styling
            for row_idx, row in enumerate(rows):
                for col_idx, value in enumerate(row):
                    item = QTableWidgetItem(str(value) if value is not None else "")
                    # Set alignment: left horizontal, center vertical
                    item.setTextAlignment(Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft)

                    # Add subtle styling based on data type
                    if isinstance(value, (int, float)) and value != 0:
                        item.setForeground(QColor("#1976d2"))  # Blue for numbers
                    elif str(value).lower() in ['true', 'false', 'yes', 'no']:
                        item.setForeground(QColor("#4caf50" if str(value).lower() in ['true', 'yes'] else "#f44336"))

                    self.table_view.setItem(row_idx, col_idx, item)

                # Enhanced action buttons - properly sized
                action_widget = QWidget()
                action_widget.setStyleSheet("background-color: transparent;")
                action_layout = QHBoxLayout(action_widget)
                action_layout.setContentsMargins(2, 2, 2, 2)  # Small margins
                action_layout.setSpacing(6)  # Space between buttons

                # Better sized edit button
                edit_btn = QPushButton("🖋️Edit")
                edit_btn.setFixedSize(28, 28)  # Increased from 24x24
                edit_btn.setStyleSheet("""
                    QPushButton {
                        background-color: #FFB300; /* Amber */
                        color: white;
                        border: none;
                        border-radius: 14px;
                        font-weight: bold;
                        font-size: 12px;
                        padding: 6px 12px;
                    }
                    QPushButton:hover {
                        background-color: #FFA000; /* Darker Amber */
                    }
                    QPushButton:pressed {
                        background-color: #FF8F00; /* Even deeper Amber */
                    }
                """)
                edit_btn.setToolTip("Edit Record")
                edit_btn.setCursor(Qt.CursorShape.PointingHandCursor)
                edit_btn.clicked.connect(lambda _, r=row, tn=table_name: self.open_edit_dialog(tn, r))

                # Better sized delete button
                delete_btn = QPushButton("🗑️Delete")
                delete_btn.setFixedSize(28, 28)  # Increased from 24x24
                delete_btn.setStyleSheet("""
                    QPushButton {
                        background-color: #D32F2F;
                        color: white;
                        border: none;
                        border-radius: 14px;
                        font-weight: bold;
                        font-size: 12px;
                        padding: 6px 12px;
                    }
                    QPushButton:hover {
                        background-color: #E53935;
                    }
                    QPushButton:pressed {
                        background-color: #C62828;
                    }
                """)
                delete_btn.setToolTip("Delete Record")
                delete_btn.setCursor(Qt.CursorShape.PointingHandCursor)
                delete_btn.clicked.connect(lambda _, r=row, tn=table_name: self.delete_row(tn, r))

                action_layout.addWidget(edit_btn)
                action_layout.addWidget(delete_btn)

                # action_layout.addWidget(edit_btn)
                # action_layout.addWidget(delete_btn)
                # action_layout.addStretch()

                self.table_view.setCellWidget(row_idx, len(columns), action_widget)

            # Apply initial filters
            self.apply_filters()

            # Add a subtle drop shadow effect to the table
            shadow_effect = QGraphicsDropShadowEffect()
            shadow_effect.setBlurRadius(15)
            shadow_effect.setColor(QColor(0, 0, 0, 30))
            shadow_effect.setOffset(0, 2)
            self.table_view.setGraphicsEffect(shadow_effect)

        except sqlite3.Error as e:
            QMessageBox.critical(self, "Database Error", f"Failed to load table:\n{str(e)}")
        finally:
            if cursor:
                cursor.close()

    def clear_all_filters(self):
        """Clear all filter inputs"""
        if hasattr(self, 'filter_inputs'):
            for filter_input in self.filter_inputs:
                filter_input.clear()

    def apply_filters(self):
        """Apply case-insensitive partial matching filters to table rows"""
        try:
            # Get lowercase filter texts
            filters = [edit.text().strip().lower() for edit in self.filter_inputs]

            # Check each row
            for row in range(self.table_view.rowCount()):
                should_show = True
                for col in range(len(filters)):
                    if filters[col]:  # Only check non-empty filters
                        item = self.table_view.item(row, col)
                        cell_text = item.text().lower() if item else ""

                        # Partial match check
                        if filters[col] not in cell_text:
                            should_show = False
                            break

                # Show/hide row based on filter match
                self.table_view.setRowHidden(row, not should_show)

        except Exception as e:
            print(f"Filter error: {str(e)}")
    def export_record(self):
        ...
    def delete_current_table(self):
        """Delete currently selected table with user confirmation"""
        if not self.current_table:
            QMessageBox.warning(self, "No Table Selected", "Please select a table to delete.")
            return

        confirm = QMessageBox.question(self, "Confirm Delete",
                                       f"Are you sure you want to permanently delete the table '{self.current_table}'?",
                                       QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)

        if confirm == QMessageBox.StandardButton.Yes:
            try:
                cursor = self.db_connection.cursor()
                cursor.execute(f"DROP TABLE IF EXISTS '{self.current_table}'")
                self.db_connection.commit()

                QMessageBox.information(self, "Table Deleted", f"Table '{self.current_table}' was deleted successfully.")

                # Refresh UI
                self.initialize_database()  # Refresh dropdown/list
                self.table_view.clear()  # Clear the current view
                self.current_table = None  # Reset current table
                self.delete_table_btn.setEnabled(True)
                self.export_btn.setEnabled(True)
                self.initialize_database()

                self.show_table_contents(self.all_tables_name[0])
                self.db_table_combo.setPlaceholderText(self.all_tables_name[0])
            except sqlite3.Error as e:
                QMessageBox.critical(self, "Database Error", f"Failed to delete table:\n{str(e)}")

            finally:
                if cursor:
                    cursor.close()

    def create_category_management_group(self) -> QGroupBox:
        """Create category management group"""
        group = QGroupBox("Resource Categories")
        layout = QVBoxLayout(group)
        layout.setContentsMargins(15, 25, 15, 15)
        layout.setSpacing(10)

        # Input field
        self.category_input = QPlainTextEdit()
        self.category_input.setFixedHeight(40)
        self.category_input.setReadOnly(True)

        # Button
        button_layout = QHBoxLayout()
        self.category_btn = QPushButton("Load Categories")
        self.category_btn.setObjectName("category_btn")
        self.category_btn.setIcon(QIcon.fromTheme("document-open"))
        self.category_btn.clicked.connect(self.select_category)
        button_layout.addWidget(self.category_btn)

        # Add to layout
        layout.addWidget(QLabel("Category File:"))
        layout.addWidget(self.category_input)
        layout.addLayout(button_layout)

        return group

    def create_file_upload_group(self) -> QGroupBox:
        """Create file upload group"""
        group = QGroupBox("Upload Attendance Data")
        layout = QVBoxLayout(group)
        layout.setContentsMargins(15, 25, 15, 15)
        layout.setSpacing(10)

        # Input field
        self.main_input = QPlainTextEdit()
        self.main_input.setFixedHeight(40)
        self.main_input.setReadOnly(True)

        # Button
        button_layout = QHBoxLayout()
        self.upload_button = QPushButton("Select File")
        self.upload_button.setObjectName("upload_button")
        self.upload_button.setIcon(QIcon.fromTheme("document-open"))
        self.upload_button.clicked.connect(self.upload_file)
        button_layout.addWidget(self.upload_button)

        # Add to layout
        layout.addWidget(QLabel("Attendance File:"))
        layout.addWidget(self.main_input)
        layout.addLayout(button_layout)

        return group

    def create_status_progress_group(self) -> QWidget:
        """Create status and progress group"""
        container = QWidget()
        layout = QVBoxLayout(container)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(15)

        # Progress bar
        self.progress_bar = QProgressBar()
        self.progress_bar.setObjectName("progress_bar")
        self.progress_bar.setValue(0)
        self.progress_bar.setTextVisible(True)
        self.progress_bar.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.progress_bar.setFixedHeight(25)

        # Status message
        self.msg_label = QLabel()
        self.msg_label.setObjectName("statusMsg")
        self.msg_label.setWordWrap(True)
        self.msg_label.setMinimumHeight(60)
        self.msg_label.hide()

        # Generate button
        self.generate_button = QPushButton("Generate Report")
        self.generate_button.setObjectName("generate_button")
        self.generate_button.setFixedHeight(45)
        self.generate_button.setIcon(QIcon.fromTheme("document-export"))
        self.generate_button.clicked.connect(self.generate_report)

        # Add to layout
        layout.addWidget(self.progress_bar)
        layout.addWidget(self.msg_label)
        layout.addWidget(self.generate_button)

        return container

    def create_database_page(self) -> QWidget:
        """Create the database management page"""
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(30, 20, 30, 20)
        layout.setSpacing(20)

        # Header
        header = QLabel("Database Management")
        header.setObjectName("title")
        header.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(header)

        # Controls
        controls_layout = QHBoxLayout()
        controls_layout.setSpacing(15)

        # Table selection
        table_layout = QHBoxLayout()
        table_layout.addWidget(QLabel("Select Table:"))

        self.db_table_combo = QComboBox()
        self.db_table_combo.currentTextChanged.connect(self.show_table_contents)
        table_layout.addWidget(self.db_table_combo, 1)

        # Buttons
        self.export_btn = QPushButton("Export Records")
        self.export_btn.setIcon(QIcon.fromTheme("document-export"))
        self.export_btn.clicked.connect(self.export_record)

        self.delete_table_btn = QPushButton("Delete Table")
        self.delete_table_btn.setIcon(QIcon.fromTheme("edit-delete"))
        self.delete_table_btn.clicked.connect(self.delete_current_table)

        # Add to controls
        controls_layout.addLayout(table_layout, 4)
        controls_layout.addWidget(self.export_btn)
        controls_layout.addWidget(self.delete_table_btn)

        layout.addLayout(controls_layout)

        # Filter row
        self.filter_row = QWidget()
        self.filter_layout = QHBoxLayout(self.filter_row)
        self.filter_layout.setContentsMargins(0, 0, 0, 0)
        self.filter_layout.setSpacing(5)
        layout.addWidget(self.filter_row)

        # Table view
        self.table_view = ModernTableWidget()
        layout.addWidget(self.table_view, 1)

        return page

    def open_resource_popup(self):
        """Opens a dialog for uploading resource mapping"""
        dialog = QDialog(self)
        dialog.setWindowTitle("Upload Resource Mapping")
        dialog.setFixedSize(400, 250)

        layout = QVBoxLayout(dialog)
        label = QLabel("Upload your resource mapping file (.csv or .xlsx)")
        layout.addWidget(label)

        file_btn = QPushButton("Choose File")
        file_btn.clicked.connect(lambda: self.choose_file(dialog))
        layout.addWidget(file_btn)

        save_btn = QPushButton("Close")
        save_btn.clicked.connect(dialog.accept)
        layout.addWidget(save_btn)

        dialog.exec()

    def handle_custom_file_upload(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "Select Excel or CSV File", "",
                                                   "Excel Files (*.xlsx *.xls);;CSV Files (*.csv)")
        if file_path:
            try:
                # Load the file
                if file_path.endswith(".csv"):
                    df = pd.read_csv(file_path)
                else:
                    df = pd.read_excel(file_path)

                # Validate the schema
                validation_result = self.validate_excel_schema(df, file_path)

                # Show appropriate message based on validation result
                if validation_result["status"] == "success":
                    # Show success message and proceed on OK
                    msg_box = QMessageBox(self)
                    msg_box.setIcon(QMessageBox.Icon.Information)
                    msg_box.setWindowTitle("File Validation Success")
                    msg_box.setText(validation_result["message"])
                    msg_box.setStandardButtons(QMessageBox.StandardButton.Ok)

                    if msg_box.exec() == QMessageBox.StandardButton.Ok:
                        # Proceed to table creation form
                        headers = df.columns.tolist()
                        self.show_table_creation_form(headers, df)

                elif validation_result["status"] == "warning":
                    # Show warning message and let user decide
                    msg_box = QMessageBox(self)
                    msg_box.setIcon(QMessageBox.Icon.Warning)
                    msg_box.setWindowTitle("File Validation Warning")
                    msg_box.setText(validation_result["message"])
                    msg_box.setStandardButtons(QMessageBox.StandardButton.Ok | QMessageBox.StandardButton.Cancel)
                    msg_box.setDefaultButton(QMessageBox.StandardButton.Ok)

                    if msg_box.exec() == QMessageBox.StandardButton.Ok:
                        # User chose to proceed despite warnings
                        headers = df.columns.tolist()
                        self.show_table_creation_form(headers, df)  # If Cancel clicked, do nothing (don't proceed)

                else:  # error status
                    # Show error message - only OK button, don't proceed
                    msg_box = QMessageBox(self)
                    msg_box.setIcon(QMessageBox.Icon.Critical)
                    msg_box.setWindowTitle("File Validation Error")
                    msg_box.setText(validation_result["message"])
                    msg_box.setStandardButtons(QMessageBox.StandardButton.Ok)
                    msg_box.exec()  # Do not proceed regardless of OK click

            except Exception as e:
                # Show critical error for file loading issues
                msg_box = QMessageBox(self)
                msg_box.setIcon(QMessageBox.Icon.Critical)
                msg_box.setWindowTitle("File Loading Error")
                msg_box.setText(f"❌ Error loading file:\n\n{str(e)}")
                msg_box.setStandardButtons(QMessageBox.StandardButton.Ok)
                msg_box.exec()
    def handle_analysis(self):
        # Logic to analyze the uploaded DataFrame
        pass

    def handle_save_to_db(self):
        # Logic to save the table to database
        pass

    def create_load_data_page(self) -> QWidget:
        """Create the data loading page"""
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(30, 20, 30, 20)
        layout.setSpacing(30)

        # Header
        header = QLabel("Load Data")
        header.setObjectName("title")
        header.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(header)

        # Card layout
        card_layout = QHBoxLayout()
        card_layout.setContentsMargins(20, 10, 20, 20)
        card_layout.setSpacing(30)
        card_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)

        # Create cards
        cards = [("📥", "Import Holiday Data", self.import_holidays_from_excel),
            ("📋", "Upload Resource Mapping", self.open_resource_popup),
            ("🗂️", "Upload Custom File", self.handle_custom_file_upload)]

        for icon, text, callback in cards:
            card = CardButton(icon, text)
            card.clicked.connect(callback)
            card_layout.addWidget(card)

        layout.addLayout(card_layout)
        layout.addStretch()

        return page

    def create_spreadsheet_page(self) -> QWidget:
        """Create the spreadsheet analysis page"""
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(30, 20, 30, 20)
        layout.setSpacing(20)

        # Header
        header_layout = QHBoxLayout()

        title = QLabel("Spreadsheet Analysis")
        title.setObjectName("title")

        # Buttons
        button_layout = QHBoxLayout()
        button_layout.setSpacing(15)

        self.load_data_card = QPushButton("Load Excel/CSV")
        self.load_data_card.setIcon(QIcon.fromTheme("document-open"))
        self.load_data_card.clicked.connect(self.handle_custom_file_upload)

        self.analyze_button = QPushButton("Analyze Data")
        self.analyze_button.setIcon(QIcon.fromTheme("edit-find"))
        self.analyze_button.clicked.connect(self.handle_analysis)

        self.save_button = QPushButton("Save to Database")
        self.save_button.setIcon(QIcon.fromTheme("document-save"))
        self.save_button.clicked.connect(self.handle_save_to_db)

        button_layout.addWidget(self.load_data_card)
        button_layout.addWidget(self.analyze_button)
        button_layout.addWidget(self.save_button)

        # Add to header
        header_layout.addWidget(title)
        header_layout.addStretch()
        header_layout.addLayout(button_layout)

        layout.addLayout(header_layout)

        # Table view
        self.excel_table_view = ModernTableWidget()
        layout.addWidget(self.excel_table_view, 1)

        return page

    def create_about_page(self) -> QWidget:
        """Create the about page"""
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(50, 50, 50, 50)
        layout.setSpacing(30)
        layout.setAlignment(Qt.AlignmentFlag.AlignCenter)

        # App icon
        icon_label = QLabel("📊")
        icon_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        icon_label.setFont(QFont("Arial", 72))
        layout.addWidget(icon_label)

        # Title
        title = QLabel("Billing Report Generator")
        title.setObjectName("title")
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(title)

        # Version
        version = QLabel("Version 1.3.0")
        version.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(version)

        # Description
        description = QLabel("A comprehensive tool for generating billing reports\n"
                             "and managing attendance data with Excel integration")
        description.setAlignment(Qt.AlignmentFlag.AlignCenter)
        description.setWordWrap(True)
        layout.addWidget(description)

        # Features
        features = QLabel("• Excel report generation\n"
                          "• Database management\n"
                          "• Attendance analysis\n"
                          "• Custom data import\n"
                          "• Theme customization")
        features.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(features)

        # Footer
        footer = QLabel("© 2025 Hitachi Digital Services")
        footer.setAlignment(Qt.AlignmentFlag.AlignCenter)
        footer.setStyleSheet("color: #90a4ae; margin-top: 30px;")
        layout.addWidget(footer)

        layout.addStretch()

        return page

    # ======================
    # CORE FUNCTIONALITY
    # ======================
    def switch_page(self, index: int):
        """Switch between application pages"""
        self.stacked_widget.setCurrentIndex(index)
        for i, btn in enumerate(self.nav_buttons):
            btn.setChecked(i == index)

    def cycle_theme(self):
        """Cycle through available themes"""
        current_index = self.themes.index(self.current_theme)
        new_index = (current_index + 1) % len(self.themes)
        self.current_theme = self.themes[new_index]
        self.update_theme_button()
        self.change_theme(self.current_theme)

    def update_theme_button(self):
        """Update the theme toggle button"""
        theme_data = {"Dark": {"icon": "🌙", "text": "Dark Theme"}, "Light": {"icon": "☀️", "text": "Light Theme"}}

        current_data = theme_data[self.current_theme]
        self.theme_btn.setText(f"{current_data['icon']} {current_data['text']}")
        self.theme_btn.setMinimumWidth(150)

    def change_theme(self, theme_name: str):
        """Apply the selected theme"""
        app = QApplication.instance()
        if theme_name == "Dark":
            app.setStyleSheet(DARK_THEME)
        else:
            app.setStyleSheet(LIGHT_THEME)
        self.current_theme = theme_name

    def show_message(self, text: str, msg_type: str = "info", timeout: int = 5000):
        """Show a status message"""
        self.msg_label.setProperty("messageType", msg_type)
        self.msg_label.setText(text)
        self.msg_label.setStyleSheet("")  # Force style refresh
        self.msg_label.show()

        # Auto-hide after timeout
        if timeout > 0:
            QTimer.singleShot(timeout, self.clear_message)

    def clear_message(self):
        """Clear the status message"""
        self.msg_label.hide()
        self.msg_label.setText("")

    def initialize_database(self):
        """Initialize the database connection and tables"""
        try:
            self.db_connection = sqlite3.connect('billing.db')
            cursor = self.db_connection.cursor()

            # Create tables if they don't exist
            tables = {'holiday': """
                    CREATE TABLE IF NOT EXISTS holiday (
                        year TEXT PRIMARY KEY,
                        holidays TEXT
                    )
                """, 'user': """
                    CREATE TABLE IF NOT EXISTS user (
                        name TEXT,
                        id_521 TEXT,
                        month TEXT,
                        year TEXT,
                        attendance_report TEXT,
                        PRIMARY KEY (name, month, year)
                    )
                """, 'user_leave': """
                    CREATE TABLE IF NOT EXISTS user_leave (
                        name TEXT,
                        id_521 TEXT,
                        year TEXT,
                        month TEXT,
                        leave_days TEXT,
                        PRIMARY KEY (name, year, month)
                    )
                """, 'resource_mapping': """
                    CREATE TABLE IF NOT EXISTS resource_mapping (
                        full_name TEXT,
                        id_521 TEXT PRIMARY KEY,
                        point_of_contact TEXT,
                        team TEXT,
                        start_date TEXT,
                        end_date TEXT
                    )
                """, 'non_complaint_user': """
                    CREATE TABLE IF NOT EXISTS non_complaint_user (
                        name TEXT,
                        id_521 TEXT,
                        year TEXT,
                        month TEXT,
                        observed_leave_count TEXT,
                        observed_leave_dates TEXT,
                        month_holiday_count TEXT,
                        month_holiday_dates TEXT,
                        PRIMARY KEY (name, year, month)
                    )
                """}

            # Create tables
            for table_name, query in tables.items():
                cursor.execute(query)

            self.db_connection.commit()

            # Load tables into dropdown
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' ORDER BY name;")
            all_tables = [table[0] for table in cursor.fetchall() if table[0].lower() != 'sqlite_sequence']
            self.db_table_combo.clear()
            self.db_table_combo.addItems(all_tables)

        except sqlite3.Error as e:
            print(f"Database initialization error: {e}")
        finally:
            if cursor:
                cursor.close()

    # ======================
    # REPORT GENERATION
    # ======================
    def generate_report(self):
        """Generate the billing report"""
        if not all([self.raw_category_list, self.categories, self.name_mapping, self.name_order_list]):
            if not self.fetch_all_resource_mappings():
                self.show_message("Error: Please load resource categories first!", "error", 5000)
                return

        # Get selected date
        self.selected_month = self.month_combo.currentText()
        self.selected_year = self.year_combo.currentText()

        # Load holidays
        self.HOLIDAY_LIST = self.get_holidays_for_year(self.selected_year)
        if not self.HOLIDAY_LIST:
            self.show_message(f"Error: Please load holidays for {self.selected_year}", "error", 5000)
            return

        # Check if attendance data is loaded
        if not self.df:
            self.show_message("Error: Please provide attendance data!", "error", 5000)
            return

        # Clean and validate data
        self.df = self.clean_keys(self.df)
        self.validate_attendance_data()

        # Generate reports for each category
        for category, valid_rsnames in self.categories.items():
            self.generate_category_report(category, valid_rsnames)

        self.show_message("Report generation completed successfully!", "success", 10000)
        self.progress_bar.setValue(100)

    def generate_category_report(self, category: str, valid_rsnames: List[str]):
        """Generate report for a specific category"""
        # Filter data for category
        filtered_df = [record for record in self.df if any(
            coverage_percentage(clean_string(record["Rsname"]), clean_string(valid_rs)) >= 60 for valid_rs in valid_rsnames)]

        if not filtered_df:
            print(f"No data for category: {category}")
            return

        # Generate Excel report
        filename = f"{category}_Timesheet_{self.selected_month} {self.selected_year}.xlsx"
        status, response, user_data, non_compliance, user_leave = generate_excel(self.selected_month,
            int(self.selected_year), filename, filtered_df, self.HOLIDAY_LIST, self.name_mapping, self.name_order_list,
            self.progress_bar)

        if status == 200:
            # Add summary and handle results
            self.add_summary_page(user_data, filename)
            if non_compliance:
                self.non_compliance_resources(non_compliance,
                                              f"{category}_non_compliance_{self.selected_month} {self.selected_year}.xlsx")
                self.update_non_complaint_user(non_compliance)
            if user_leave:
                self.update_user_leave(user_leave)
        else:
            self.show_message(f"Error: {response}", "error", 5000)

    # ======================
    # HELPER METHODS
    # ======================
    def get_holidays_for_year(self, year: str) -> List[str]:
        """Get holidays for a specific year"""
        try:
            cursor = self.db_connection.cursor()
            cursor.execute("SELECT holidays FROM holiday WHERE year = ?", (year,))
            result = cursor.fetchone()
            return json.loads(result[0]) if result else []
        except sqlite3.Error as e:
            print(f"Error fetching holidays: {e}")
            return []
        finally:
            if cursor:
                cursor.close()

    def fetch_all_resource_mappings(self) -> bool:
        """Fetch all resource mappings from the database"""
        try:
            cursor = self.db_connection.cursor()
            cursor.execute("SELECT * FROM resource_mapping")
            records = cursor.fetchall()

            if not records:
                return False

            # Get column names
            columns = [desc[0] for desc in cursor.description]
            self.raw_category_list = [dict(zip(columns, row)) for row in records]

            # Process mappings
            for item in self.raw_category_list:
                name = clean_string(item['full_name'])
                team = item["team"]

                if team in self.categories:
                    self.categories[team].append(name)
                else:
                    self.categories[team] = [name]

                self.name_mapping[name] = (item['id_521'], item['point_of_contact'], item['start_date'], item['end_date'])

            # Sort categories
            for k, v in self.categories.items():
                self.categories[k] = sorted(v)
                self.name_order_list.extend(self.categories[k])

            return True
        except sqlite3.Error as e:
            print(f"Database error: {e}")
            return False
        finally:
            if cursor:
                cursor.close()

    def validate_attendance_data(self):
        """Validate the loaded attendance data"""
        try:
            # Extract month abbreviations from column names
            month_abbrs = set()
            for record in self.df:
                for key in record.keys():
                    if re.match(r"^\w{3}, \d{1,2}-\w{3}$", key):
                        month_abbrs.add(key.split('-')[-1])

            # Find most common month
            if not month_abbrs:
                raise ValueError("No valid date columns found")

            common_month = max(month_abbrs, key=list(month_abbrs).count)
            if common_month != self.selected_month[:3]:
                self.show_message("Warning: Attendance data doesn't match selected month. "
                                  "Proceed with caution.", "info", 5000)
        except Exception as e:
            self.show_message(f"Validation error: {str(e)}", "error", 5000)

    # ======================
    # UI HANDLERS
    # ======================
    def import_holidays_from_excel(self):
        """Import holidays from Excel file"""
        file_path, _ = QFileDialog.getOpenFileName(self, "Select Holiday File", "",
            "Spreadsheet Files (*.xlsx *.xls *.numbers);;All Files (*)")

        if not file_path:
            return

        # Update UI with file info
        file_name = os.path.basename(file_path)
        file_size = os.path.getsize(file_path)
        self.holiday_input.setPlainText(f"{file_name} ({sizeof_fmt(file_size)})")

        # Process file
        try:
            # Read file
            if file_path.endswith(".csv"):
                df = pd.read_csv(file_path, header=None)
            else:
                df = pd.read_excel(file_path, header=None)

            # Extract year
            excel_year = str(df.iloc[0, 0])
            if not excel_year.isdigit() or len(excel_year) != 4:
                raise ValueError("First cell must contain a 4-digit year")

            # Extract holidays
            holidays = []
            for idx in range(1, len(df)):
                date_val = df.iloc[idx, 0]
                if pd.isna(date_val):
                    continue

                if isinstance(date_val, str):
                    date_obj = parse(date_val)
                elif isinstance(date_val, pd.Timestamp):
                    date_obj = date_val.to_pydatetime()
                else:
                    date_obj = pd.to_datetime(date_val).to_pydatetime()

                if str(date_obj.year) != excel_year:
                    raise ValueError(f"Date {date_obj.date()} doesn't match file year {excel_year}")

                formatted_date = date_obj.strftime("%d-%m-%Y")
                if formatted_date not in holidays:
                    holidays.append(formatted_date)

            # Save to database
            self.save_holidays_to_db(excel_year, holidays)
            self.show_message(f"Loaded {len(holidays)} holidays for {excel_year}", "success", 5000)

        except Exception as e:
            self.show_message(f"Error: {str(e)}", "error", 5000)

    def save_holidays_to_db(self, year: str, holidays: List[str]):
        """Save holidays to database"""
        try:
            cursor = self.db_connection.cursor()

            # Check if year exists
            cursor.execute("SELECT year FROM holiday WHERE year = ?", (year,))
            exists = cursor.fetchone()

            # Update or insert
            if exists:
                cursor.execute("UPDATE holiday SET holidays = ? WHERE year = ?", (json.dumps(holidays), year))
            else:
                cursor.execute("INSERT INTO holiday (year, holidays) VALUES (?, ?)", (year, json.dumps(holidays)))

            self.db_connection.commit()
            self.HOLIDAY_LIST = holidays

            # Update UI
            if self.year_combo.findText(year) == -1:
                self.year_combo.addItem(year)
            self.year_combo.setCurrentText(year)

        except sqlite3.Error as e:
            self.show_message(f"Database error: {str(e)}", "error", 5000)
        finally:
            if cursor:
                cursor.close()

    # ======================
    # EXCEL GENERATION (Optimized)
    # ======================
    def generate_excel(month: str, year: int, output_file: str, selected_rows: List[Dict], holiday_list: List[str],
                       name_mapping: Dict, name_order: List[str], progress_bar: QProgressBar) -> Tuple[
        int, str, Optional[List], Optional[List], Optional[List]]:
        """Generate Excel report with optimized performance"""
        try:
            # Prepare data structures
            sheets_name = []
            user_data = []
            non_compliance = []
            user_leave_records = []

            # Get month details
            month_details, month_number = get_month_details(month, year)
            month_num_str = f"{month_number:02d}"
            month_holidays = [h for h in holiday_list if re.search(rf"\b\d{{2}}-{month_num_str}-\d{{4}}\b", h)]
            holiday_days = [h.split("-")[0] for h in month_holidays]

            # Sort rows by custom order
            order_map = {preprocess_name(name): idx for idx, name in enumerate(name_order)}
            selected_rows = sorted(selected_rows, key=lambda x: order_map.get(preprocess_name(x["Rsname"]), float('inf')))

            # Setup progress tracking
            total_rows = len(selected_rows)
            progress_step = 100 / total_rows if total_rows else 100
            current_progress = 0

            # Process each row
            for row in selected_rows:
                # Prepare data for current row
                data_model = []
                leave_dates = []
                leave_taken = 0
                public_holiday = 0
                mismatch_dates = []
                billable_days = 0
                weekends = 0
                total_working = 0

                # Get resource details
                name = preprocess_name(row.get("Rsname"))
                details = get_details_for_name(name, name_mapping)
                start_date, end_date = (details[2], details[3]) if details else (None, None)

                # Process each day in month
                for week in month_details:
                    for day in week:
                        if not day:
                            continue

                        date = day["day"]
                        day_str = f"{date}-{month[:3]}"
                        day_name = day['day_name'][:3]
                        is_weekend = day["is_weekend"]

                        # Calculate day status
                        dt_status, is_holiday = self.calculate_day_status(row, date, day_str, day_name, is_weekend,
                            holiday_days, details, month_num_str)

                        # Update counters
                        if is_weekend:
                            weekends += 1
                        else:
                            total_working += 1

                        if is_holiday:
                            public_holiday += 1

                        billable_days += dt_status

                        # Add to data model
                        data_model.append(
                            (day_str, day_name, dt_status, "Holiday" if is_holiday else ("Weekend" if is_weekend else "")))

                # Create DataFrame for sheet
                sheet_name = sanitize_sheet_name(row.get("Rsname"))
                df = self.create_data_frame(row, data_model, leave_taken, total_working, public_holiday, weekends, details)

                # Save sheet
                sheets_name.append(sheet_name)
                self.save_sheet(df, sheet_name, output_file)

                # Update progress
                current_progress += progress_step
                progress_bar.setValue(int(current_progress))

            # Apply styling to all sheets
            self.apply_excel_styling(output_file, sheets_name)

            return 200, "Report generated successfully", user_data, non_compliance, user_leave_records

        except Exception as e:
            print(f"Error generating report: {e}")
            return 500, str(e), None, None, None

    def calculate_day_status(self, row: Dict, date: int, day_str: str, day_name: str, is_weekend: bool,
                             holiday_days: List[str], details: Tuple, month_num: str) -> Tuple[float, bool]:
        """Calculate the status for a specific day"""
        # Default values
        dt_status = 0.0
        is_holiday = False

        # Check if date is a holiday
        date_str = f"{date:02d}"
        if date_str in holiday_days:
            is_holiday = True

        # Get attendance value
        key = f"{day_name}, {date_str}-{month_num}"
        attendance_val = row.get(key, 0)

        # Calculate status based on attendance
        if is_weekend:
            dt_status = 0.0
        elif attendance_val == 8:
            dt_status = 1.0
            if is_holiday:
                is_holiday = False  # Override if attendance marked on holiday
        elif attendance_val == 4:
            dt_status = 0.5
        elif 2.5 <= attendance_val < 4:
            dt_status = 0.25
        else:  # attendance_val == 0
            dt_status = 0.0
            if not is_holiday:
                # Leave day
                pass

        return dt_status, is_holiday

    # ======================  # MAIN EXECUTION  # ======================


if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setStyle("Fusion")

    # Set application styles
    palette = QPalette()
    palette.setColor(QPalette.ColorRole.Window, QColor(240, 240, 240))
    palette.setColor(QPalette.ColorRole.WindowText, QColor(0, 0, 0))
    app.setPalette(palette)

    window = MainWindow()
    window.show()
    sys.exit(app.exec())