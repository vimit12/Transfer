import copy
import sys
from collections import Counter
from PyQt6.QtWidgets import (QApplication, QMainWindow, QWidget, QStackedWidget, QTableWidget, QGridLayout, QPushButton,
                             QLabel, QVBoxLayout, QTableWidgetItem, QFileDialog, QTextEdit, QGraphicsDropShadowEffect,
                             QFrame, QLineEdit, QComboBox, QFormLayout, QHeaderView, QDialog, QProgressBar, QMessageBox,
                             QSizePolicy, QHBoxLayout, QSpacerItem, QGroupBox, QPlainTextEdit, QScrollArea,
                             QAbstractItemView)
from PyQt6.QtGui import QFont, QIcon, QColor
from PyQt6 import QtCore
from PyQt6.QtCore import Qt, QDate, QDateTime, QTimer
import sqlite3
import json
from openpyxl import Workbook
import pandas as pd
from pandas._libs.tslibs.timestamps import Timestamp
from openpyxl.utils import get_column_letter
from datetime import datetime
import calendar
import itertools
import os
import re
from dateutil.parser import parse
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
import numpy as np
from pandas._libs.tslibs.nattype import NaTType

# ======================
# THEME DEFINITIONS
# ======================
DARK_THEME = """
QWidget {
    background-color: #2b2b2b;
    color: #ffffff;
    selection-background-color: #3a3a3a;
}

/* Buttons */
QPushButton {
    background-color: #353535;
    border: 1px solid #454545;
    border-radius: 4px;
    padding: 6px 10px;
    min-width: 70px;
    font-size: 10px;
    color: #ffffff; /* Ensure white text */
}

QPushButton:hover {
    background-color: #454545;
}

QPushButton:checked {
    background-color: #007acc;
    border-color: #006ab3;
}

/* Inputs */
QLineEdit, QComboBox {
    background-color: #353535;
    border: 1px solid #454545;
    border-radius: 4px;
    padding: 6px 8px;
    min-height: 28px;
    color: #ffffff; /* White text */
}

/* Dropdown Items */
QComboBox QAbstractItemView {
    background-color: #353535;
    color: #ffffff;
    selection-background-color: #007acc;
}

/* Sidebar */
QFrame#sidebar {
    background-color: #252526;
    border-right: 1px solid #353535;
}

/* Titles */
QLabel#title {
    font-size: 20px;
    font-weight: 600;
    color: #007acc;
    padding-bottom: 15px;
}

QLabel#subtitle {
    font-size: 14px;
    color: #888;
}

QFormLayout QLabel {
    font-weight: 500;
    padding-bottom: 3px;
    color: #ffffff;
}

/* Table */
QTableWidget {
    background-color: #353535;
    color: #ffffff;
    border: 1px solid #454545;
    gridline-color: #555555;
}

QHeaderView::section {
    background-color: #454545;
    color: #ffffff;
    padding: 6px;
    font-weight: bold;
}

QTableWidget::item {
    color: #ffffff;
    padding: 6px;
}

QTableWidget::item:selected {
    background-color: #007acc;
    color: #ffffff;
}

/* Holiday Group Box Buttons */
#format_info_btn {
    background-color: #6c5ce7;
    color: white;
    border: 1px solid #5d4ec9;
}

#load_holiday_btn {
    background-color: #00b894;
    color: white;
    border: 1px solid #00a383;
}

#view_holiday_btn {
    background-color: #e66767;
    color: white;
    border: 1px solid #d35454;
}

/* Button Hover Effects */
QPushButton:hover {
    background-color: rgba(108, 92, 231, 0.8);
}

#load_holiday_btn:hover {
    background-color: rgba(0, 184, 148, 0.8);
}

#view_holiday_btn:hover {
    background-color: rgba(230, 103, 103, 0.8);
}
QComboBox {
    background-color: #353535;
    border: 1px solid #454545;
    border-radius: 4px;
    color: #ffffff;
    padding: 6px;
    min-height: 28px;
}
QComboBox QAbstractItemView {
    background-color: #3a3a3a;
    color: #ffffff;
    selection-background-color: #007acc;
}
"""

LIGHT_THEME = """
QWidget {
    background-color: #ffffff;
    color: #333333;
    selection-background-color: #e0e0e0;
}

/* Buttons */
QPushButton {
    background-color: #f5f5f5;
    border: 1px solid #cccccc;
    border-radius: 4px;
    padding: 6px 10px;
    min-width: 70px;
    font-size: 10px;
    color: #333333;
}

QPushButton:hover {
    background-color: #e8e8e8;
}

QPushButton:checked {
    background-color: #007acc;
    color: white;
    border-color: #006ab3;
}

/* Inputs */
QLineEdit, QComboBox {
    background-color: #ffffff;
    border: 1px solid #cccccc;
    border-radius: 4px;
    padding: 6px 8px;
    min-height: 28px;
    color: #333333;
}

/* Dropdown Items */
QComboBox QAbstractItemView {
    background-color: #ffffff;
    color: #333333;
    selection-background-color: #007acc;
}

/* Sidebar */
QFrame#sidebar {
    background-color: #f8f8f8;
    border-right: 1px solid #e0e0e0;
}

/* Titles */
QLabel#title {
    font-size: 20px;
    font-weight: 600;
    color: #007acc;
    padding-bottom: 15px;
}

QLabel#subtitle {
    font-size: 14px;
    color: #666;
}

QFormLayout QLabel {
    font-weight: 500;
    padding-bottom: 3px;
    color: #333333;
}

/* Table */
QTableWidget {
    background-color: #ffffff;
    color: #333333;
    border: 1px solid #cccccc;
    gridline-color: #dddddd;
}

QHeaderView::section {
    background-color: #f0f0f0;
    color: #333333;
    padding: 6px;
    font-weight: bold;
}

QTableWidget::item {
    color: #333333;
    padding: 6px;
}

QTableWidget::item:selected {
    background-color: #007acc;
    color: #ffffff;
}

/* Holiday Group Box Buttons */
#format_info_btn {
    background-color: #a8a4e6;
    color: #2d3436;
    border: 1px solid #8f8bd9;
}

#load_holiday_btn {
    background-color: #55efc4;
    color: #2d3436;
    border: 1px solid #48cfad;
}

#view_holiday_btn {
    background-color: #ff7675;
    color: #2d3436;
    border: 1px solid #ff6564;
}

/* Button Hover Effects */
QPushButton:hover {
    background-color: rgba(168, 164, 230, 0.8);
}

#load_holiday_btn:hover {
    background-color: rgba(85, 239, 196, 0.8);
}

#view_holiday_btn:hover {
    background-color: rgba(255, 118, 117, 0.8);
}
"""

# =======================
#  GENERIC FUNCTION
#========================

TOTAL_WORKING_DAY = 0


def clean_date(value):
    """Convert NaT to None and Timestamps to string format for database insertion"""
    if isinstance(value, NaTType):
        return None  # Convert NaT to None
    if isinstance(value, pd.Timestamp):
        return value.strftime("%d-%m-%Y")  # Convert Timestamp to 'DD-MM-YYYY' format
    return value  # Return as is if it's already a string


def clean_string(s):
    # Remove text inside brackets (e.g., "[C]", "[c]", "[text]"), commas, and hyphens
    cleaned = re.sub(r"\[.*?\]|[, -]", " ", s)
    return cleaned.lower().strip()


def coverage_percentage(str1, str2):
    words1 = set(clean_string(str1).split())
    words2 = set(clean_string(str2).split())
    common_words = words1 & words2
    coverage = (len(common_words) / max(len(words1), len(words2))) * 100 if max(len(words1), len(words2)) > 0 else 0
    return coverage


def read_file(file_path):
    try:
        # Check file extension to determine how to read the file
        if file_path.endswith(".csv"):
            df = pd.read_csv(file_path)
        elif file_path.endswith(".xlsx"):
            df = pd.read_excel(file_path)
        else:
            raise ValueError("Unsupported file format. Please provide a CSV or Excel file.")

        # Return the DataFrame if reading is successful
        return df.to_dict("records")

    except FileNotFoundError:
        print(f"Error: File not found at '{file_path}'")
    except Exception as e:
        print(f"Error reading file '{file_path}': {e}")

    # Return None if there is an error
    return None


def sort_list_of_dicts(data):
    # Separate the dict with 'Name' equal to 'Total'
    total_dict = [d for d in data if d.get('Name') == 'Total']

    # Sort the remaining dicts by 'Name'
    sorted_data = sorted([d for d in data if d.get('Name') != 'Total'], key=lambda x: x['Name'])

    # Append the 'Total' dict at the end
    if total_dict:
        sorted_data.extend(total_dict)

    return sorted_data


def get_month_details(month_name, year):
    # Get the month number from the month name
    month_number = list(calendar.month_name).index(month_name.capitalize())

    # Get the calendar for the specified month and year
    cal = calendar.monthcalendar(year, month_number)

    # Dictionary to map weekday number to weekday name
    weekdays = {0: "Monday", 1: "Tuesday", 2: "Wednesday", 3: "Thursday", 4: "Friday", 5: "Saturday", 6: "Sunday", }

    month_details = []

    # Iterate through each week in the month
    for week in cal:
        week_details = []
        # Iterate through each day in the week
        for day in week:
            # If the day is zero, it means it's part of the previous or next month
            if day == 0:
                week_details.append(None)
            else:
                # Get the day name using the weekdays dictionary
                day_name = weekdays[calendar.weekday(year, month_number, day)]
                # Check if it's a weekend (Saturday or Sunday)
                is_weekend = day_name in ["Saturday", "Sunday"]
                # Append the day details to the week_details list
                week_details.append({"day": day, "day_name": day_name, "is_weekend": is_weekend})
        # Append the week_details list to the month_details list
        month_details.append(week_details)

    return month_details, month_number


def date_calculation(date):
    # Check if the date is already a datetime object
    if isinstance(date, datetime):
        date_obj = date
    else:
        # Parse the date string using the format DD-MM-YYYY
        date_obj = datetime.strptime(date, "%d-%m-%Y")

    # Extract the day, month, and year
    return date_obj.day, date_obj.month, date_obj.year


def get_details_for_name(name, name_mapping):
    """
    Given an input name and a mapping of names to details,
    returns the (start_date, end_date) if a key is found with 100% coverage.
    """
    for key in name_mapping:
        # Assuming clean_string is defined elsewhere
        if coverage_percentage(name, preprocess_name(key)) == 100:
            return name_mapping[key]
    # If no exact 100% match is found, return (None, None)
    return None

def sanitize_sheet_name(name: str, default: str = "Sheet1") -> str:
    """
    Returns a safe sheet name, max 31 characters, removing illegal characters.
    """
    if not name:
        return default

    # Remove any characters not allowed in Excel sheet names
    for char in ['\\', '/', '*', '?', ':', '[', ']']:
        name = name.replace(char, '')

    # Truncate to 31 characters
    return name[:31]

def generate_excel(month, year, output_file_name, selected_row, holiday_list, name_mapping, name_order_list, progress_bar):
    global TOTAL_WORKING_DAY
    sheets_name = []
    try:
        user_data = list()

        user_leave_record = list()
        non_complaince_user = []
        month_name = month
        holiday_list = holiday_list

        year = int(year)
        month_details, month_number = get_month_details(month_name, year)
        month_number = f"{month_number:02}" if month_number < 10 else month_number
        month_holiday_list = [x for x in holiday_list if re.findall(f"\d+-{month_number}-\d+", x)]
        # print("month_holiday_list ===>", month_holiday_list)
        month_day_holiday_list = [k.split("-")[0] for k in month_holiday_list]
        # print("month_day_holiday_list ===>", month_day_holiday_list)
        df_sheets = dict()
        excel_file_path = output_file_name

        # Creating a mapping from Rsname to their positions in the custom order
        order_map = {preprocess_name(name): index for index, name in enumerate(name_order_list)}

        # Sorting the list of dictionaries by the custom order
        selected_row = sorted(selected_row, key=lambda x: order_map.get(preprocess_name(x["Rsname"]), float('inf')))

        attendance_len = len(selected_row)
        progress_step = 0
        step = 100 / attendance_len
        for new_data in selected_row:
            leave_dates = []
            start_date, end_date, sd, sm, sy, ed, em, ey, sm_flag, em_flag = (None,) * 10
            billable_days = 0
            weekends = 0
            total_working_days = 0
            data_model = []
            leave_taken = 0
            public_holiday = 0
            mismatch_date = []
            msg = None
            name = preprocess_name(new_data.get("Rsname"))

            details = get_details_for_name(name, name_mapping)
            start_date, end_date = details[2], details[3]
            # start_date, end_date = name_mapping[name][2:] if name_mapping.get(name) else (None,) * 2

            if start_date:
                sd, sm, sy = date_calculation(start_date)

                sm = f"{sm:02}" if sm < 10 else sm

                sm_flag = sm == month_number

            if end_date:
                ed, em, ey = date_calculation(end_date)
                em = f"{em:02}" if em < 10 else em
                em_flag = em == month_number

            for week in month_details:
                for day in week:
                    if day:
                        date = day["day"]
                        is_weekend_or_leave = "Weekend" if day["is_weekend"] else ""
                        if day["is_weekend"]:
                            weekends += 1
                        else:
                            total_working_days += 1
                        dt = f"{date}-{month_name[:3]}"
                        day_name = f"{day['day_name']}"

                        if sm_flag or em_flag:
                            if sm_flag and not em_flag:
                                if date < sd:
                                    data_model.append((dt, day_name[:3], "Not On Boarded", "", "", "", "", "",))
                                    continue
                            elif not sm_flag and em_flag:
                                if date > ed:
                                    data_model.append((dt, day_name[:3], "Off Boarded", "", "", "", "", "",))
                                    continue
                            else:
                                if date < sd:
                                    data_model.append((dt, day_name[:3], "Not On Boarded", "", "", "", "", "",))
                                    continue
                                elif date > ed:
                                    data_model.append((dt, day_name[:3], "Off Boarded", "", "", "", "", "",))
                                    continue
                                else:
                                    msg = "On Board"

                        key = (f"{day_name[:3]}, {f'{date:02}' if date < 10 else date}-"
                               f"{month_name[:3].title()}")
                        # print("DAY ===>", f"{date:02}" if date < 10 else date)
                        # holiday_list = [j.split("-")[0] for j in holiday_list ]
                        calculated_date = f"{date:02}" if date < 10 else f"{date}"
                        if day["is_weekend"]:
                            dt_status = 0
                        else:
                            rounded_values = list(np.around(np.arange(2.5, 4, 0.1), decimals=1))

                            match new_data.get(key):
                                case 8:
                                    dt_status = 1
                                    # As some user can mark attendace on holiday date
                                    if calculated_date in month_day_holiday_list:
                                        mismatch_date.append(calculated_date)
                                        public_holiday += 1
                                        dt_status = 0

                                        # for Holiday keyword to be added
                                        is_weekend_or_leave = "Holiday"
                                case 4:
                                    dt_status = 0.5
                                    leave_taken += 0.5
                                    leave_dates.append(f"{date}(0.5)")
                                case value if value in rounded_values:
                                    dt_status = 0.25
                                    leave_taken += 0.25
                                case 0:
                                    if calculated_date in month_day_holiday_list:
                                        public_holiday += 1
                                        dt_status = 0

                                        # for Holiday keyword to be added
                                        is_weekend_or_leave = "Holiday"
                                    else:
                                        leave_taken += 1
                                        dt_status = 0
                                        leave_dates.append(date)
                                        # for leave keyword to be added
                                        is_weekend_or_leave = "Leave"
                                case _:
                                    leave_taken += 1
                                    dt_status = 0
                        billable_days += dt_status

                        """
                            (dt, day_name[:3], dt_status, is_weekend_or_leave, "", "", "", "")
                            ('1-Jun', 'Thu', 1, '', '', '', '', '')
                        """
                        if sm_flag and not em_flag:
                            if date < sd:
                                dt_status = 0
                                total_working_days -= 1

                        if em_flag and not sm_flag:
                            if date > ed:
                                dt_status = 0
                                total_working_days -= 1

                        if em_flag and sm_flag:
                            if date > ed or date < sd:
                                dt_status = 0
                                total_working_days -= 1

                        if is_weekend_or_leave == "Holiday" and dt_status == 0:
                            data_model.append((dt, day_name[:3], "Holiday", "", "", "", "", "",))
                        elif is_weekend_or_leave == "Leave" and dt_status == 0:
                            data_model.append((dt, day_name[:3], "Leave", "", "", "", "", "",))
                        elif is_weekend_or_leave == "Weekend" and dt_status == 0:
                            data_model.append((dt, day_name[:3], dt_status, is_weekend_or_leave, "", "", "", "",))
                        else:
                            if dt_status == 1:
                                data_model.append((dt, day_name[:3], 8, is_weekend_or_leave, "", "", "", "",))
                            else:
                                data_model.append((dt, day_name[:3], 4, is_weekend_or_leave, "", "", "", "",))
            billable_days = total_working_days - leave_taken
            # billable_days = total_working_days
            # print("USER =====>", new_data.get("Rsname"))
            # print("BILLABLE ====>", billable_days)
            # print("WEEKENDS ====>", weekends)
            # print("TOTAL WORKING DAYS ====>", total_working_days)
            # print("LEAVE TAKEN ====>", leave_taken)
            # print("-"*20)
            # print("PUBLIC HOLIDAY ====>", public_holiday)
            point_of_contact = (details[1] if details else "xxxxxxx")
            ID_521 = details[0] if details else "xxxxxxx"
            if mismatch_date:
                non_complaince_user.append(
                    {"Name": new_data.get("Rsname"), "521_ID": details[0], "Year": year, "Month": month,
                     "Listed Month Holiday": month_day_holiday_list, "Attendance Marked on Holiday": mismatch_date, })
            data = {"Vendor Organization": ["Resource Name", "Month", "Date"],
                    "Hitachi Digital Service": [f"{new_data.get('Rsname')}", f"{month_name}", "Day", ],
                    "Point of Contact": ["5-2-1", "Working Days", "Working Status"],
                    f"{point_of_contact}": [f"{ID_521}", total_working_days, "Remarks", ],
                    "Adjustments from Last Month": ["", "", ""], "0": ["", "", ""], "": ["", "", ""],
                    "Week Off": ["Personal/Sick Leave", "", ""], }
            df = pd.DataFrame(data)

            # Create a new sheet or get the existing one
            sheet_name = sanitize_sheet_name(new_data.get("Rsname"))
            billable_days = 0
            user_days_cal = 0
            for row in data_model:
                value = row[2]  # third element
                is_weekend_flag = row[3] == 'Weekend'
                if isinstance(value, (int, float)):
                    if value == 8:
                        billable_days += 1
                    elif value == 4:
                        billable_days += 0.5
                # you Leave and holiday to be added then include 'Holiday' and 'Leave'
                valid_status = ['Not On Boarded', 'Off Boarded']
                if value not in valid_status and not is_weekend_flag:
                    user_days_cal += 1
                df.loc[len(df)] = row

            df.loc[len(df)] = ["Leaves Taken", leave_taken, "Billable Days", billable_days, "", "", "", "", ]

            df.loc[len(df)] = ["Weekends", weekends, "", "", "", "", "", ""]
            df.loc[len(df)] = ["Public Holidays", public_holiday, "", "", "", "", "", "", ]

            # print(df)
            df_sheets.update({sheet_name: df})
            if sm_flag or em_flag:
                user_data.append(
                    {"Name": sheet_name,  # "Total Billable Time": (total_working_days-leave_taken-public_holiday) * 8 ,
                     "Billable Time (Hours)": (user_days_cal - leave_taken) * 8,
                     #    "Weekends": weekends, "Public Holidays": public_holiday,
                     "Total Number of Billable Days": user_days_cal - leave_taken,
                     "Service Credit Pool Days": leave_taken, })
            else:
                user_data.append(
                    {"Name": sheet_name,  #   "Total Billable Time": (total_working_days-leave_taken-public_holiday) * 8 ,
                     "Billable Time (Hours)": (total_working_days - leave_taken) * 8,
                     #    "Weekends": weekends, "Public Holidays": public_holiday,
                     "Total Number of Billable Days": total_working_days - leave_taken,
                     "Service Credit Pool Days": leave_taken, })
            user_leave_record.append(
                {"name": sheet_name, "id_521": details[0], "year": year, "month": month, "leave_days": leave_dates})
            progress_step += int(step)
            progress_bar.setValue(progress_step)
        else:
            TOTAL_WORKING_DAY = total_working_days

            # Create a Pandas Excel writer using XlsxWriter as the engine
            with pd.ExcelWriter(excel_file_path, engine="xlsxwriter") as writer:

                for key, value in df_sheets.items():
                    # Write each dataframe to a different sheet
                    value.to_excel(writer, sheet_name=key, index=False)

                # Access the XlsxWriter workbook and worksheet objects
                workbook = writer.book
                worksheets = writer.sheets

                # Access each worksheet and modify the formatting if needed
                for sheet_name, worksheet in worksheets.items():
                    # Example: set column width of the first column to 20
                    worksheet.set_column(0, 0, 20)
                    worksheet.set_column(1, 1, 20)
                    worksheet.set_column(2, 2, 20)
                    worksheet.set_column(3, 3, 15)
                    worksheet.set_column(4, 4, 30)
                    worksheet.set_column(7, 7, 17)

                    sheets_name.append(sheet_name)

            wb_style = load_workbook(excel_file_path)
            border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"),
                            bottom=Side(border_style="thin"), )
            for i in sheets_name:
                sheet = wb_style[i]

                for row in sheet.iter_rows():
                    for cell in row:
                        cell.border = border
                        if cell.value == "Weekend":
                            numbers = re.findall(r"\d+", cell.coordinate)[0]
                            cell_list = [f"B{numbers}", f"C{numbers}", f"D{numbers}", ]
                            for k in cell_list:
                                cell_bold = sheet[k]
                                cell_bold.fill = PatternFill(start_color="b6bbbf", end_color="b6bbbf",
                                                             fill_type="solid", )  # grey color
                        if cell.value == "Leave":
                            cell_bold = sheet[cell.coordinate]
                            cell_bold.fill = PatternFill(start_color="fce1dc", end_color="fce1dc",
                                                         fill_type="solid", )  # red color

                        if cell.value == "Holiday":
                            cell_bold = sheet[cell.coordinate]
                            cell_bold.fill = PatternFill(start_color="cffccf", end_color="cffccf",
                                                         fill_type="solid", )  # green color

                        if cell.value in ["Leaves Taken", "Weekends", "Public Holidays", "Billable Days", ]:
                            cell_bold = sheet[cell.coordinate]
                            cell_bold.font = Font(bold=True, color="FFFFFF")
                            cell_bold.fill = PatternFill(start_color="4d6c82", end_color="4d6c82", fill_type="solid", )

                for j in ["B1", "D1", "H1", "F1"]:
                    cell_bold = sheet[j]
                    cell_bold.font = Font(bold=False)

                for j in ["A1", "C1", "E1", "A2", "C2", "A3", "C3", "A4", "B4", "C4", "D4", "G1", "G2", ]:
                    cell_bold = sheet[j]
                    cell_bold.font = Font(bold=True)
                    if j == "G2":
                        cell_bold.fill = PatternFill(start_color="fce1dc", end_color="fce1dc",
                                                     fill_type="solid", )  # red color
                    else:
                        cell_bold.fill = PatternFill(start_color="b6bbbf", end_color="b6bbbf",
                                                     fill_type="solid", )  # grey color

                for j in ["A4", "B4", "C4", "D4", "E4", "F4"]:
                    cell_bold = sheet[j]
                    cell_bold.font = Font(bold=True, color="FFFFFF")
                    cell_bold.fill = PatternFill(start_color="4d6c82", end_color="4d6c82", fill_type="solid")

                for k in [chr(i) + f"{j}" for i in range(65, 73) for j in range(1, 5)]:
                    cell = sheet[k]
                    cell.alignment = Alignment(horizontal="left")

                for k in [f"A{j}" for j in range(5, 40)]:
                    cell = sheet[k]

                    cell.alignment = Alignment(horizontal="center")

                for k in [f"C{j}" for j in range(5, 40)]:
                    cell = sheet[k]

                    if cell.value == "Billable Days":
                        global range_limit
                        range_limit = int(re.findall(r"\d+", cell.coordinate)[0])
                        break

                    cell.alignment = Alignment(horizontal="center")

                for k in [i + f"{j}" for i in ["B", "D"] for j in range(5, 40)]:
                    cell = sheet[k]

                    if cell.coordinate == f"D{range_limit}":
                        cell.font = Font(bold=True)

                    cell.alignment = Alignment(horizontal="center")

            else:
                wb_style.save(excel_file_path)

        return [200, "Report Generated Successfully.", user_data, non_complaince_user, user_leave_record]

    except Exception as e:
        # Log the error
        print(f"An error occurred: {str(e)}")
        exc_type, exc_obj, exc_tb = sys.exc_info()
        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        print(exc_type, fname, exc_tb.tb_lineno)

        return [500, str(e), None, None, None]


def format_date(date_str):
    if isinstance(date_str, Timestamp):
        date_str = date_str.strftime("%Y-%m-%d")  # Convert Timestamp to string
    try:
        # Attempt to parse the date string with various formats
        date_obj = parse(date_str)
        formatted_date = date_obj.strftime("%d-%m-%Y")
        return formatted_date
    except (ValueError, TypeError):
        # If parsing fails or the input is not a string, return the original value
        return str(date_str)


# Preprocess the name and item['Name'] strings
def preprocess_name(input_str):
    # Split the string into words, remove spaces, convert to lowercase, and sort
    return "".join(sorted(input_str.replace(",", "").lower().split()))


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Billing Report Generator")
        self.setGeometry(100, 100, 1200, 700)
        self.setFixedSize(1200, 700)
        self.current_theme = "Light"
        self.themes = ["Dark", "Light", "System"]
        self.db_connection = None
        self.current_table = None
        self.current_year = str(datetime.now().year)  # Add this line
        # print(f"Current year : {self.current_year}")
        self.raw_category_list, self.name_order_list = [], []  # ✅ Separate lists
        self.categories, self.name_mapping = {}, {}  # ✅ Separate dictionaries
        self.HOLIDAY_LIST = []
        self.df = None

        self.init_ui()
        self.initialize_database()

    def init_ui(self):
        # Create custom title bar controls
        title_bar_widget = QWidget()
        title_bar_layout = QHBoxLayout(title_bar_widget)
        title_bar_layout.setContentsMargins(0, 0, 10, 0)

        # Theme toggle button
        self.theme_btn = QPushButton()
        self.theme_btn.setFixedHeight(32)  # Fixed height only
        self.theme_btn.setMinimumWidth(120)  # Minimum width to prevent cutting
        self.theme_btn.clicked.connect(self.cycle_theme)

        # Add vertical spacing around the button
        btn_container = QVBoxLayout()
        btn_container.setContentsMargins(0, 10, 0, 0)  # Top margin for the button
        btn_container.addWidget(self.theme_btn)

        # Set button styling
        self.theme_btn.setStyleSheet("""
                QPushButton {
                    margin-top: 20px; /* Increase top margin */
                    padding: 5px 10px;
                    border-radius: 10px;
                    border: 1px solid #606060;
                    font-weight: 500;
                }
            """)
        self.update_theme_button()

        title_bar_layout.addSpacerItem(QSpacerItem(40, 20, QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Minimum))
        title_bar_layout.addLayout(btn_container)  # Add the container instead of direct button

        # Add title bar widget to main window
        self.setMenuWidget(title_bar_widget)

        # Main layout
        main_widget = QWidget()
        main_layout = QHBoxLayout(main_widget)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)

        # Sidebar (updated to 3 buttons)
        self.sidebar = QFrame()
        self.sidebar.setObjectName("sidebar")
        self.sidebar.setFixedWidth(150)

        sidebar_layout = QVBoxLayout()
        sidebar_layout.setContentsMargins(8, 20, 8, 20)
        sidebar_layout.setSpacing(8)

        self.btn_home = QPushButton("🏠 Home")
        self.btn_database = QPushButton("📁 Database")
        self.btn_load_data = QPushButton("📑 Load Dataset")
        self.btn_load_spreadsheet = QPushButton("📊 Load Spreadsheet")
        self.btn_about = QPushButton("ℹ️ About")

        for btn in [self.btn_home, self.btn_database, self.btn_load_data, self.btn_load_spreadsheet, self.btn_about]:
            btn.setCheckable(True)
            btn.setFixedSize(130, 40)
            btn.setFont(QFont("Segoe UI", 10))
            btn.setCursor(Qt.CursorShape.PointingHandCursor)
            sidebar_layout.addWidget(btn)

        sidebar_layout.addStretch()
        self.sidebar.setLayout(sidebar_layout)

        # Main content area
        self.stacked_widget = QStackedWidget()
        self.init_pages()

        main_layout.addWidget(self.sidebar)
        main_layout.addWidget(self.stacked_widget)
        self.setCentralWidget(main_widget)

        # Connect navigation
        self.btn_home.clicked.connect(lambda: self.switch_page(0))
        self.btn_database.clicked.connect(lambda: self.switch_page(1))
        self.btn_load_data.clicked.connect(lambda: self.switch_page(2))  # ✅ New Page
        self.btn_load_spreadsheet.clicked.connect(lambda: self.switch_page(3))
        self.btn_about.clicked.connect(lambda: self.switch_page(4))

    def init_pages(self):
        # ======================
        # PAGE CREATION
        # ======================
        self.stacked_widget.addWidget(self.create_home_page())
        self.stacked_widget.addWidget(self.create_database_page())
        self.stacked_widget.addWidget(self.create_load_data_page())
        self.stacked_widget.addWidget(self.create_spreadsheet_page())
        self.stacked_widget.addWidget(self.create_about_page())

    def initialize_database(self):
        """Initialize database connection and required tables"""
        try:
            self.db_connection = sqlite3.connect('billing.db')
            cursor = self.db_connection.cursor()

            # Check existing tables
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
            existing_tables = [table[0].lower() for table in cursor.fetchall() if table[0].lower() != 'sqlite_sequence']

            # Table creation queries with corrected syntax
            tables = {'holiday': '''
                            CREATE TABLE IF NOT EXISTS holiday (
                                year TEXT PRIMARY KEY,
                                holidays TEXT
                            )
                        ''', 'user': '''
                            CREATE TABLE IF NOT EXISTS user (
                                name TEXT,
                                id_521 TEXT,
                                month TEXT,
                                year TEXT,
                                attendance_report TEXT,
                                PRIMARY KEY (name, month, year)
                            )
                        ''', 'user_leave': '''
                            CREATE TABLE IF NOT EXISTS user_leave (
                                name TEXT,
                                id_521 TEXT,
                                year TEXT,
                                month TEXT,
                                leave_days TEXT,
                                PRIMARY KEY (name, year, month)
                            )
                        ''', 'resource_mapping': '''
                            CREATE TABLE IF NOT EXISTS resource_mapping (
                                full_name TEXT,
                                id_521 TEXT PRIMARY KEY,
                                point_of_contact TEXT,
                                team TEXT,
                                start_date TEXT,
                                end_date TEXT
                            )
                        ''', 'non_complaint_user': '''
                            CREATE TABLE IF NOT EXISTS non_complaint_user (
                                name TEXT,
                                id_521 TEXT,
                                year TEXT,
                                month TEXT,
                                observed_leave_count TEXT,
                                observed_leave_dates TEXT,
                                month_holiday_count TEXT,
                                month_holiday_dates TEXT,
                                PRIMARY KEY (name, year, month)
                            )
                        '''}

            # Create missing tables
            created_tables = []
            for table_name, query in tables.items():
                if table_name not in existing_tables:
                    try:
                        cursor.execute(query)
                        created_tables.append(table_name)
                    except sqlite3.Error as e:
                        print(f"Error creating table {table_name}: {e}")

            self.db_connection.commit()

            # Update database page with table list
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' ORDER BY name;")
            all_tables = [table[0] for table in cursor.fetchall() if table[0].lower() != 'sqlite_sequence']
            self.all_tables_name = copy.deepcopy(all_tables)
            # Update dropdown
            self.db_table_combo.clear()
            self.db_table_combo.addItems(all_tables)

            # # Check for current year holidays  # cursor.execute("SELECT year FROM holiday WHERE year = ?", (self.current_year,))  # if not cursor.fetchone():  #     self.show_holiday_import_dialog()

        except sqlite3.Error as e:
            error_msg = f"Database error: {str(e)}"
            print(error_msg)
            if hasattr(self, 'db_status_label'):
                self.db_status_label.setText(error_msg)
        finally:
            if cursor:
                cursor.close()

    def create_load_data_page(self):
        """Creates the Load Data page with two card-like buttons"""
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(20)

        label = QLabel("Load Data")
        label.setFont(QFont("Segoe UI", 14, QFont.Weight.Bold))
        layout.addWidget(label)

        button_layout = QHBoxLayout()

        # Card 1
        self.card1 = QPushButton("📥 Import Holiday Data")
        self.card1.setFixedSize(250, 120)
        self.card1.setStyleSheet(self.get_card_style())
        self.card1.clicked.connect(self.import_holidays_from_excel)

        # Card 2
        self.card2 = QPushButton("📋 Upload Resource Mapping")
        self.card2.setFixedSize(250, 120)
        self.card2.setStyleSheet(self.get_card_style())
        self.card2.clicked.connect(self.open_resource_popup)

        # Card 3
        self.card3 = QPushButton("🗂️ Upload Custom File")
        self.card3.setFixedSize(250, 120)
        self.card3.setStyleSheet(self.get_card_style())
        self.card3.clicked.connect(self.handle_custom_file_upload)

        button_layout.addWidget(self.card1)
        button_layout.addWidget(self.card2)
        button_layout.addWidget(self.card3)

        layout.addLayout(button_layout)
        layout.addStretch()

        return page

    def create_spreadsheet_page(self):
        page = QWidget()

        # Outer layout
        outer_layout = QVBoxLayout(page)
        outer_layout.setContentsMargins(20, 20, 20, 20)
        outer_layout.setSpacing(20)

        # Header layout - title on left, buttons on right
        header_layout = QHBoxLayout()

        # Title on extreme left
        title = QLabel("📊 Spreadsheet Loader")
        title.setStyleSheet("font-size: 18px; font-weight: bold;")
        title.setAlignment(Qt.AlignmentFlag.AlignLeft)

        # Shared button style
        button_style = """
            QPushButton {
                background-color: #ffffff;
                border: 2px solid #d0d0d0;
                border-radius: 16px;
                font-size: 14pt;
                font-weight: 500;
                padding: 10px 20px;
                color: #333;
            }
            QPushButton:hover {
                background-color: #f2f9ff;
                border: 2px solid #7cbfff;
                color: #005999;
            }
            QPushButton:pressed {
                background-color: #e6f0ff;
                border: 2px solid #5aa0ff;
            }
        """

        # Load Data button
        self.load_data_card = QPushButton("📂 Load Excel/CSV File")
        self.load_data_card.setStyleSheet(button_style)
        self.load_data_card.clicked.connect(self.handle_custom_file_upload)

        # Analyze button
        self.analyze_button = QPushButton("📊 Analyze")
        self.analyze_button.setStyleSheet(button_style)
        self.analyze_button.clicked.connect(self.handle_analysis)

        # Save to DB button
        self.save_button = QPushButton("💾 Save to DB")
        self.save_button.setStyleSheet(button_style)
        self.save_button.clicked.connect(self.handle_save_to_db)

        # Add title and buttons to header
        header_layout.addWidget(title)
        header_layout.addStretch()
        header_layout.addWidget(self.load_data_card)
        header_layout.addWidget(self.analyze_button)
        header_layout.addWidget(self.save_button)

        # Table view section
        self.excel_table_view = QTableWidget()
        self.excel_table_view.setSortingEnabled(True)
        self.excel_table_view.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.excel_table_view.horizontalHeader().setStretchLastSection(True)
        self.excel_table_view.verticalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        self.excel_table_view.setAlternatingRowColors(True)
        self.excel_table_view.verticalHeader().setDefaultSectionSize(30)
        self.excel_table_view.horizontalHeader().setSectionsMovable(False)
        self.excel_table_view.verticalHeader().setVisible(False)

        self.excel_table_view.setStyleSheet("""
            QTableWidget {
                background-color: #ffffff;
                border: 1px solid #d0d0d0;
                border-radius: 12px;
                font-size: 12pt;
                gridline-color: #e0e0e0;
                selection-background-color: #cce6ff;
                selection-color: #000000;
                alternate-background-color: #f9f9f9;
            }

            QHeaderView::section {
                background-color: #f2f9ff;
                color: #005999;
                padding: 10px;
                border: 1px solid #d0d0d0;
                font-weight: bold;
                font-size: 13pt;
            }

            QTableWidget::item {
                padding: 6px;
            }

            QTableWidget::item:hover {
                background-color: #eaf6ff;
            }
        """)

        # Add components to main layout
        outer_layout.addLayout(header_layout)
        outer_layout.addWidget(self.excel_table_view)

        return page

    def handle_analysis(self):
        # Logic to analyze the uploaded DataFrame
        pass

    def handle_save_to_db(self):
        # Logic to save the table to database
        pass

    def get_card_style(self):
        """Returns the CSS style for card buttons"""
        return """
            QPushButton {
                background-color: #ffffff;
                border: 2px solid #d0d0d0;
                border-radius: 16px;
                font-size: 15pt;
                font-weight: 500;
                padding: 30px;
                color: #333;
            }

            QPushButton:hover {
                background-color: #f2f9ff;
                border: 2px solid #7cbfff;
                color: #005999;
            }

            QPushButton:pressed {
                background-color: #e6f0ff;
                border: 2px solid #5aa0ff;
            }
        """

    def open_resource_popup(self):
        """Opens a dialog for uploading resource mapping"""
        dialog = QDialog(self)
        dialog.setWindowTitle("Upload Resource Mapping")
        dialog.setFixedSize(400, 250)

        layout = QVBoxLayout(dialog)
        label = QLabel("Upload your resource mapping file (.csv or .xlsx)")
        layout.addWidget(label)

        file_btn = QPushButton("Choose File")
        file_btn.clicked.connect(lambda: self.choose_file(dialog))
        layout.addWidget(file_btn)

        save_btn = QPushButton("Close")
        save_btn.clicked.connect(dialog.accept)
        layout.addWidget(save_btn)

        dialog.exec()

    def handle_custom_file_upload(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "Select Excel or CSV File", "",
                                                   "Excel Files (*.xlsx *.xls);;CSV Files (*.csv)")
        if file_path:
            try:
                # Load the file
                if file_path.endswith(".csv"):
                    df = pd.read_csv(file_path)
                else:
                    df = pd.read_excel(file_path)

                # Validate the schema
                validation_result = self.validate_excel_schema(df, file_path)

                # Show appropriate message based on validation result
                if validation_result["status"] == "success":
                    # Show success message and proceed on OK
                    msg_box = QMessageBox(self)
                    msg_box.setIcon(QMessageBox.Icon.Information)
                    msg_box.setWindowTitle("File Validation Success")
                    msg_box.setText(validation_result["message"])
                    msg_box.setStandardButtons(QMessageBox.StandardButton.Ok)

                    if msg_box.exec() == QMessageBox.StandardButton.Ok:
                        # Proceed to table creation form
                        headers = df.columns.tolist()
                        self.show_table_creation_form(headers, df)

                elif validation_result["status"] == "warning":
                    # Show warning message and let user decide
                    msg_box = QMessageBox(self)
                    msg_box.setIcon(QMessageBox.Icon.Warning)
                    msg_box.setWindowTitle("File Validation Warning")
                    msg_box.setText(validation_result["message"])
                    msg_box.setStandardButtons(QMessageBox.StandardButton.Ok | QMessageBox.StandardButton.Cancel)
                    msg_box.setDefaultButton(QMessageBox.StandardButton.Ok)

                    if msg_box.exec() == QMessageBox.StandardButton.Ok:
                        # User chose to proceed despite warnings
                        headers = df.columns.tolist()
                        self.show_table_creation_form(headers, df)  # If Cancel clicked, do nothing (don't proceed)

                else:  # error status
                    # Show error message - only OK button, don't proceed
                    msg_box = QMessageBox(self)
                    msg_box.setIcon(QMessageBox.Icon.Critical)
                    msg_box.setWindowTitle("File Validation Error")
                    msg_box.setText(validation_result["message"])
                    msg_box.setStandardButtons(QMessageBox.StandardButton.Ok)
                    msg_box.exec()  # Do not proceed regardless of OK click

            except Exception as e:
                # Show critical error for file loading issues
                msg_box = QMessageBox(self)
                msg_box.setIcon(QMessageBox.Icon.Critical)
                msg_box.setWindowTitle("File Loading Error")
                msg_box.setText(f"❌ Error loading file:\n\n{str(e)}")
                msg_box.setStandardButtons(QMessageBox.StandardButton.Ok)
                msg_box.exec()

    def validate_excel_schema(self, df, file_path):
        """
        Validate Excel/CSV schema for SQL compatibility
        Returns: dict with 'status' ('success'/'warning'/'error'), 'message' string
        """
        errors = []

        # Check 1: File must not be empty
        if df.empty:
            return {"status": "error",
                "message": "❌ File Validation Failed:\n\n• The uploaded file is empty. Please upload a file with data."}

        # Check 2: Must have headers (column names)
        if df.columns.empty:
            return {"status": "error",
                "message": "❌ File Validation Failed:\n\n• No column headers found. Excel/CSV files must have column headers in the first row."}

        # Check 3: Validate column headers
        header_errors = []
        for i, col in enumerate(df.columns):
            col_name = str(col).strip()

            # Check for empty/null headers
            if pd.isna(col) or col_name == '' or col_name.lower() in ['nan', 'unnamed']:
                header_errors.append(f"Column {i + 1}: Empty or invalid header name")
                continue

            # Check for duplicate headers
            if list(df.columns).count(col) > 1:
                header_errors.append(f"Column '{col_name}': Duplicate header name found")

            # Check for SQL reserved words (basic check)
            sql_reserved = ['select', 'insert', 'update', 'delete', 'create', 'drop', 'alter', 'table', 'database', 'index',
                            'view', 'trigger', 'procedure', 'function', 'where', 'order', 'group', 'having', 'join', 'union',
                            'distinct']
            if col_name.lower() in sql_reserved:
                header_errors.append(f"Column '{col_name}': Uses SQL reserved word (not recommended)")

        if header_errors:
            errors.extend(header_errors)

        # Check 4: Validate data structure
        data_errors = []

        # :: TODO - Enable this section if you need validation for all empty rows
        # Check for completely empty columns
        # for col in df.columns:
        #     if df[col].isna().all():
        #         data_errors.append(f"Column '{col}': Contains no data (all empty cells)")

        # Check 5: Row structure validation
        row_errors = []

        # Check if all rows are empty
        if df.isna().all().all():
            row_errors.append("All rows contain only empty cells")

        # Check for rows with all empty values (except header)
        empty_rows = df.index[df.isna().all(axis=1)].tolist()
        if len(empty_rows) > len(df) * 0.5:  # More than 50% empty rows
            row_errors.append(f"Too many empty rows found ({len(empty_rows)} out of {len(df)} rows)")

        # Check 6: Data type consistency warnings
        consistency_warnings = []
        for col in df.columns:
            non_null_data = df[col].dropna()
            if len(non_null_data) > 0:
                # Check for mixed data types that might cause issues
                data_types = set(type(val).__name__ for val in non_null_data)
                if len(data_types) > 2:  # More than 2 different data types
                    consistency_warnings.append(f"Column '{col}': Contains mixed data types - {', '.join(data_types)}")

        # Compile all errors
        all_errors = []
        if header_errors:
            all_errors.append("📋 Header Issues:")
            all_errors.extend([f"  • {error}" for error in header_errors])
            all_errors.append("")

        if data_errors:
            all_errors.append("📊 Data Issues:")
            all_errors.extend([f"  • {error}" for error in data_errors])
            all_errors.append("")

        if row_errors:
            all_errors.append("📝 Row Structure Issues:")
            all_errors.extend([f"  • {error}" for error in row_errors])
            all_errors.append("")

        if consistency_warnings:
            all_errors.append("⚠️ Data Type Warnings:")
            all_errors.extend([f"  • {warning}" for warning in consistency_warnings])
            all_errors.append("")

        # Determine validation result
        critical_errors = header_errors + data_errors + row_errors

        if critical_errors:
            error_message = "❌ File Validation Failed:\n\n" + "\n".join(all_errors)
            error_message += "\n💡 Suggestions:\n"
            error_message += "  • Ensure first row contains unique column headers\n"
            error_message += "  • Remove completely empty rows/columns\n"
            error_message += "  • Fix duplicate or missing headers\n"
            error_message += "  • Ensure data is properly structured in rows and columns"

            return {"status": "error", "message": error_message}
        elif consistency_warnings:
            # Show warnings but allow to proceed
            warning_message = "⚠️ File Validation Warnings:\n\n" + "\n".join(all_errors)
            warning_message += "\n✅ File structure is valid but contains data type inconsistencies.\n"
            warning_message += "You can proceed, but please review the data types during table creation."

            return {"status": "warning", "message": warning_message}
        else:
            # All good
            success_message = f"✅ File validation successful!\n\n📄 File: {file_path.split('/')[-1]}\n📊 Rows: {len(df)}\n📋 Columns: {len(df.columns)}"
            return {"status": "success", "message": success_message}

    def show_table_creation_form(self, headers, df):
        dialog = QDialog(self)
        dialog.setWindowTitle("Define Table Structure")
        dialog.resize(700, 500)  # Wider dialog

        main_layout = QVBoxLayout(dialog)

        # Table Name input
        table_name_input = QLineEdit()
        table_name_input.setPlaceholderText("Enter Table Name")
        table_name_input.setObjectName("tableNameInput")  # Set object name for styling

        main_layout.addWidget(QLabel("Table Name:"))
        main_layout.addWidget(table_name_input)

        # Scrollable area
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)

        scroll_widget = QWidget()
        grid_layout = QGridLayout(scroll_widget)
        grid_layout.setSpacing(15)

        dropdowns = {}
        row = 0

        for header in headers:
            label = QLabel(str(header))
            combo = QComboBox()
            combo.addItems(["TEXT", "INTEGER", "REAL", "DATE"])
            dropdowns[header] = combo

            col = row // ((len(headers) + 1) // 2)  # auto split into 2 columns
            grid_layout.addWidget(label, row % ((len(headers) + 1) // 2), col * 2)
            grid_layout.addWidget(combo, row % ((len(headers) + 1) // 2), col * 2 + 1)

            row += 1

        scroll_area.setWidget(scroll_widget)
        main_layout.addWidget(scroll_area)

        # Submit button
        submit_btn = QPushButton("Load Table")
        submit_btn.setEnabled(False)
        main_layout.addWidget(submit_btn)

        # --- VALIDATION LOGIC ---
        # A simple validator for common SQL table naming rules
        # Starts with a letter or underscore, followed by letters, numbers, or underscores.
        table_name_regex = re.compile(r'^[a-zA-Z_][a-zA-Z0-9_]*$')

        # A list of common SQL reserved keywords to avoid
        reserved_keywords = {
            'SELECT', 'CREATE', 'TABLE', 'FROM', 'WHERE', 'AND', 'OR', 'NOT', 'INSERT',
            'INTO', 'UPDATE', 'SET', 'DELETE', 'GROUP', 'BY', 'ORDER', 'ASC', 'DESC',
            'JOIN', 'LEFT', 'RIGHT', 'INNER', 'OUTER', 'PRIMARY', 'KEY', 'FOREIGN',
            'REFERENCES', 'NULL', 'UNIQUE', 'CHECK', 'DEFAULT'
        }

        def is_valid_table_name(name):
            if not name:
                return False
            # Check against regex and reserved keywords
            if not table_name_regex.match(name) or name.upper() in reserved_keywords:
                return False
            return True

        def set_name_input_style(is_valid):
            if is_valid:
                # Revert to default style
                table_name_input.setStyleSheet("")
                table_name_input.setToolTip("")
            else:
                # Set background to light red and add a tooltip
                table_name_input.setStyleSheet("background-color: #ffe6e6;")
                table_name_input.setToolTip("Invalid table name. "
                                            "It must start with a letter or underscore and "
                                            "contain only letters, numbers, and underscores. "
                                            "Spaces and special characters are not allowed.")

        # --- CONNECTORS AND HANDLERS ---
        def check_form_complete():
            is_name_valid = is_valid_table_name(table_name_input.text().strip())
            set_name_input_style(is_name_valid)

            if not is_name_valid:
                submit_btn.setEnabled(False)
                return

            for cb in dropdowns.values():
                if cb.currentText() == "":
                    submit_btn.setEnabled(False)
                    return

            submit_btn.setEnabled(True)

        table_name_input.textChanged.connect(check_form_complete)
        for cb in dropdowns.values():
            cb.currentIndexChanged.connect(check_form_complete)

        def on_submit():
            table_name = table_name_input.text().strip()

            if not is_valid_table_name(table_name):
                # Show error message box
                msg_box = QMessageBox(dialog)
                msg_box.setWindowTitle("Invalid Table Name")
                msg_box.setText("The table name you entered is not valid.")
                msg_box.setInformativeText(
                    "Table names must:\n"
                    "- Start with a letter (a-z, A-Z) or an underscore (_)\n"
                    "- Contain only letters, numbers, and underscores\n"
                    "- Not be a SQL reserved keyword\n"
                    "Please correct the name and try again."
                )
                msg_box.setIcon(QMessageBox.Warning)
                msg_box.exec_()
                return  # Do not proceed

            # If validation passes, proceed with table creation
            column_defs = {col: dropdowns[col].currentText() for col in headers}

            #:: TODO - Enable this below section if needed
            # self.create_dynamic_table(table_name, column_defs, df)
            self.create_and_show_table(table_name, column_defs, df)
            dialog.accept()

        submit_btn.clicked.connect(on_submit)

        # Initial check to set the button state correctly on form load
        check_form_complete()
        dialog.exec()

    def create_and_show_table(self, table_name, column_defs, df):
        """Create and display a stylized table from a DataFrame with maximum efficiency."""
        try:
            # Early exit for empty data
            if df.empty:
                self.excel_table_view.clear()
                return

            # Get data once and cache
            columns = df.columns.tolist()
            num_rows, num_cols = df.shape

            # Batch all UI updates
            self.excel_table_view.setUpdatesEnabled(False)
            self.excel_table_view.setSortingEnabled(False)

            try:
                # Clear efficiently
                self.excel_table_view.clearContents()

                # Set dimensions once
                self.excel_table_view.setRowCount(num_rows)
                self.excel_table_view.setColumnCount(num_cols)
                self.excel_table_view.setHorizontalHeaderLabels(columns)

                # Configure appearance (do this once, not per item)
                self._configure_table_appearance()

                # Bulk populate using vectorized operations
                self._populate_table_fast(df, num_rows, num_cols)

            finally:
                # Re-enable updates once at the end
                self.excel_table_view.setUpdatesEnabled(True)
                self.excel_table_view.setSortingEnabled(True)

        except Exception as e:
            self._handle_table_error(e)

    def _configure_table_appearance(self):
        """Configure table appearance once - separated for clarity."""
        table = self.excel_table_view

        # Basic settings
        table.setAlternatingRowColors(True)
        table.setWordWrap(False)  # Disable word wrap for performance
        table.setShowGrid(True)

        # Headers
        v_header = table.verticalHeader()
        h_header = table.horizontalHeader()

        # Use faster resize modes
        v_header.setSectionResizeMode(QHeaderView.ResizeMode.Fixed)
        v_header.setDefaultSectionSize(25)  # Fixed row height

        # Set scroll policies
        table.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        table.setVerticalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)

        # Lightweight stylesheet - removed gradients and complex selectors
        table.setStyleSheet("""
            QTableWidget {
                background-color: white;
                border: 1px solid #ddd;
                gridline-color: #f0f0f0;
                selection-background-color: #e3f2fd;
                font-size: 11px;
                font-family: 'Segoe UI', Arial;
            }
            QTableWidget::item {
                padding: 8px;
                border: none;
            }
            QTableWidget::item:selected {
                background-color: #e3f2fd;
            }
            QHeaderView::section {
                background-color: #f8f9fa;
                color: #495057;
                font-weight: bold;
                padding: 8px;
                border: 1px solid #dee2e6;
            }
        """)

        # Set column widths efficiently
        total_cols = table.columnCount()
        if total_cols <= 3:
            default_width = 250
        else:
            default_width = 150

        for col in range(total_cols):
            table.setColumnWidth(col, default_width)

    def _populate_table_fast(self, df, num_rows, num_cols):
        """Ultra-fast table population using batch operations."""
        # Pre-compute color mappings for reuse
        colors = {'positive': QColor("#2e7d32"), 'negative': QColor("#c62828"), 'zero': QColor("#1565c0"),
            'true': QColor("#2e7d32"), 'false': QColor("#c62828"), 'warning': QColor("#f57c00"),
            'default': QColor("#2c3e50")}

        # Pre-define alignment flags
        align_right = Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignRight
        align_left = Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft
        non_editable = Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsEnabled

        # Process data by chunks for better memory usage
        chunk_size = min(1000, num_rows)

        for chunk_start in range(0, num_rows, chunk_size):
            chunk_end = min(chunk_start + chunk_size, num_rows)

            # Get chunk of data
            chunk_data = df.iloc[chunk_start:chunk_end]

            for local_row_idx, (_, row) in enumerate(chunk_data.iterrows()):
                actual_row_idx = chunk_start + local_row_idx

                for col_idx, value in enumerate(row):
                    # Fast type checking and formatting
                    item_text, color_key, alignment = self._format_cell_value(value)

                    # Create item with pre-computed properties
                    item = QTableWidgetItem(item_text)
                    item.setTextAlignment(alignment)
                    item.setFlags(non_editable)

                    # Set color efficiently
                    if color_key in colors:
                        item.setForeground(colors[color_key])

                    # Set item once
                    self.excel_table_view.setItem(actual_row_idx, col_idx, item)

    def _format_cell_value(self, value):
        """Fast cell value formatting with minimal type checking."""

        # Handle None/NaN first (most common case)
        if pd.isna(value) or value is None:
            return "", "default", Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft

        # Handle numeric types (use isinstance once)
        if isinstance(value, (int, float, complex)):
            if isinstance(value, bool):  # bool is subclass of int, check first
                return "Yes" if value else "No", "true" if value else "false", Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft
            elif isinstance(value, float):
                if abs(value) < 1e-10:
                    return "0.0000", "zero", Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignRight
                else:
                    return f"{value:.4f}", "positive" if value > 0 else "negative", Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignRight
            else:  # int or complex
                text = str(value)
                if hasattr(value, 'real') and value != 0:  # complex or non-zero
                    color = "positive" if (getattr(value, 'real', value) > 0) else "negative"
                else:
                    color = "zero"
                return text, color, Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignRight

        # Handle strings efficiently
        elif isinstance(value, str):
            if not value.strip():  # empty string
                return "", "default", Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft

            # Fast string matching - check first character for speed
            lower_val = value.lower().strip()
            first_char = lower_val[0] if lower_val else ''

            if first_char in 'ty' and lower_val in {'true', 'yes', 'y'}:
                return value, "true", Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft
            elif first_char in 'fn' and lower_val in {'false', 'no', 'n'}:
                return value, "false", Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft
            elif first_char == 'w' and 'warn' in lower_val:
                return value, "warning", Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft
            else:
                return value, "default", Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft

        # Fallback for other types
        else:
            return str(value), "default", Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft

    def _handle_table_error(self, error):
        """Lightweight error handling."""
        print(f"Table Error: {error}")

        # Simple fallback - just clear the table
        try:
            self.excel_table_view.clear()
            # Add a single error message item
            self.excel_table_view.setRowCount(1)
            self.excel_table_view.setColumnCount(1)
            self.excel_table_view.setHorizontalHeaderLabels(['Error'])
            error_item = QTableWidgetItem(f"Error loading data: {str(error)}")
            self.excel_table_view.setItem(0, 0, error_item)
        except:
            pass  # If even this fails, just leave it empty

    # def setup_filter_row(self, columns):
    #     filter_container = QWidget()
    #     filter_layout = QGridLayout(filter_container)
    #     filter_layout.setContentsMargins(12, 8, 12, 8)
    #     filter_layout.setSpacing(6)
    #
    #     filter_label = QLabel("🔍 Filters:")
    #     filter_label.setStyleSheet("font-weight: bold; font-size: 12px; color: #424242;")
    #     filter_layout.addWidget(filter_label, 0, 0, 1, len(columns))
    #
    #     self.filter_excel_inputs = []
    #     for col_idx, col_name in enumerate(columns):
    #         input_box = QLineEdit()
    #         input_box.setPlaceholderText(col_name)
    #         input_box.setStyleSheet("""
    #             QLineEdit {
    #                 border: 2px solid #e0e0e0;
    #                 border-radius: 6px;
    #                 padding: 6px;
    #                 font-size: 11px;
    #             }
    #             QLineEdit:focus {
    #                 border-color: #1976d2;
    #                 background-color: #fafafa;
    #             }
    #         """)
    #         input_box.textChanged.connect(self.apply_spreadsheet_filters)
    #         self.filter_excel_inputs.append(input_box)
    #         filter_layout.addWidget(input_box, 1, col_idx)
    #
    #     if hasattr(self, 'filter_layout'):
    #         self.filter_layout.addWidget(filter_container)
    #
    # def apply_spreadsheet_filters(self):
    #     """Filter the table based on values typed into the filter input fields."""
    #     try:
    #         for row in range(self.excel_table_view.rowCount()):
    #             self.excel_table_view.setRowHidden(row, False)
    #
    #         for col_idx, filter_input in enumerate(self.filter_excel_inputs):
    #             filter_text = filter_input.text().strip().lower()
    #             if not filter_text:
    #                 continue
    #
    #             for row in range(self.excel_table_view.rowCount()):
    #                 item = self.excel_table_view.item(row, col_idx)
    #                 if item is None:
    #                     continue
    #
    #                 cell_text = item.text().strip().lower()
    #                 if filter_text not in cell_text:
    #                     self.excel_table_view.setRowHidden(row, True)
    #
    #     except Exception as e:
    #         QMessageBox.warning(self, "Filter Error", f"An error occurred while filtering:\n{str(e)}")
    def sanitize_column_name(self, col):
        if not isinstance(col, str):
            col = str(col)
        col = re.sub(r'\W+', '_', col)
        if col and col[0].isdigit():
            col = f'col_{col}'
        return col.strip('_').lower()

    def create_dynamic_table(self, table_name, column_defs, df):
        cursor = None
        try:
            cursor = self.db_connection.cursor()

            # Step 1: Sanitize column names and build mapping
            sanitized_map = {col: self.sanitize_column_name(col) for col in column_defs.keys()}

            # Step 2: Create table SQL
            column_sql_parts = [f'"{sanitized_map[col]}" {column_defs[col]}' for col in column_defs]
            columns_sql = ", ".join(column_sql_parts)
            create_table_sql = f'CREATE TABLE IF NOT EXISTS "{table_name}" ({columns_sql})'
            cursor.execute(create_table_sql)

            # Step 3: Insert data
            placeholders = ", ".join(["?"] * len(column_defs))
            sanitized_cols = [sanitized_map[col] for col in column_defs]
            insert_sql = f'INSERT INTO "{table_name}" ({", ".join(sanitized_cols)}) VALUES ({placeholders})'

            for _, row in df.iterrows():
                values = [row[col].isoformat() if column_defs[col] == "DATE" and pd.notnull(row[col]) else row[col] for col
                          in column_defs]
                cursor.execute(insert_sql, values)

            # Step 4: Optional metadata logging
            metadata_sql = """
                CREATE TABLE IF NOT EXISTS imported_files_metadata (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    table_name TEXT,
                    columns TEXT,
                    imported_at TEXT
                )
            """
            cursor.execute(metadata_sql)
            cursor.execute("INSERT INTO imported_files_metadata (table_name, columns, imported_at) VALUES (?, ?, ?)",
                           (table_name, json.dumps(sanitized_map), datetime.now().isoformat()))

            self.db_connection.commit()

            # Show success message with QMessageBox
            msg_box = QMessageBox(self)
            msg_box.setIcon(QMessageBox.Icon.Information)
            msg_box.setWindowTitle("Table Creation Success")
            msg_box.setText(f"✅ Table '{table_name}' created successfully!\n\n📊 {len(df)} records imported.")
            msg_box.setStandardButtons(QMessageBox.StandardButton.Ok)
            msg_box.exec()

            self.initialize_database()#again intialize the db page

        except sqlite3.Error as e:
            # Show database error with QMessageBox
            msg_box = QMessageBox(self)
            msg_box.setIcon(QMessageBox.Icon.Critical)
            msg_box.setWindowTitle("Database Error")
            msg_box.setText(f"❌ Database Error:\n\n{str(e)}")
            msg_box.setStandardButtons(QMessageBox.StandardButton.Ok)
            msg_box.exec()

            if self.db_connection:
                self.db_connection.rollback()

        finally:
            if cursor:
                cursor.close()

    def add_data_resource_tab(self, df):
        required_columns = {"Full Name", "521 ID", "Point of Contact", "Team", "Start Date", "End Date"}
        if not required_columns.issubset(df.columns):
            QMessageBox.critical(self, "Error", "Invalid file format. Ensure correct column names.")
            return

        cursor = self.db_connection.cursor()
        for _, row in df.iterrows():
            cursor.execute("SELECT COUNT(*) FROM resource_mapping WHERE id_521 = ?", (row["521 ID"],))
            exists = cursor.fetchone()[0] > 0

            if exists:
                # ✅ Update the existing record
                cursor.execute("""
                                UPDATE resource_mapping 
                                SET full_name = ?, point_of_contact = ?, team = ?, start_date = ?, end_date = ?
                                WHERE id_521 = ?
                            """, (
                    clean_string(row["Full Name"]), row["Point of Contact"], row["Team"], clean_date(row["Start Date"]),
                    clean_date(row["End Date"]), row["521 ID"]))
            else:
                # ✅ Insert new record
                cursor.execute("""
                                INSERT INTO resource_mapping (full_name, id_521, point_of_contact, team, start_date, end_date)
                                VALUES (?, ?, ?, ?, ?, ?)
                            """, (clean_string(row["Full Name"]), row["521 ID"], row["Point of Contact"], row["Team"],
                                  clean_date(row["Start Date"]), clean_date(row["End Date"])))

        self.db_connection.commit()
        cursor.close()

    def choose_file(self, parent_dialog):
        """Opens a file dialog to select a file and imports its data"""
        file_path, _ = QFileDialog.getOpenFileName(self, "Select File", "", "CSV/Excel Files (*.csv *.xlsx *.xls)")
        if not file_path:
            return

        try:
            if file_path.endswith(".csv"):
                df = pd.read_csv(file_path)
            else:  # For .xlsx files, read only "PublicCloudResourceList" sheet
                df = pd.read_excel(file_path, sheet_name="PublicCloudResourceList")

            self.add_data_resource_tab(df)

            QMessageBox.information(self, "Success", "Resource mapping data uploaded successfully!")

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to import file: {str(e)}")

        parent_dialog.close()

    def show_holiday_import_dialog(self):
        """Show holiday import prompt and handle file selection"""
        reply = QMessageBox.question(self, "Holiday Data Required",
                                     f"No holidays found for {self.current_year}. Would you like to import from Excel?",
                                     QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)

        if reply == QMessageBox.StandardButton.Yes:
            self.import_holidays_from_excel()

    def import_holidays_from_excel(self):
        """Handle Excel and Numbers import with datetime-formatted cells"""
        file_path, _ = QFileDialog.getOpenFileName(self, "Select Holiday File", "",
                                                   "Spreadsheet Files (*.xlsx *.xls *.numbers);;Excel Files (*.xlsx *.xls);;Numbers Files (*.numbers);;All Files (*)")

        if not file_path:
            return

        try:
            file_extension = os.path.splitext(file_path)[1].lower()
            holidays = []
            excel_year = None

            if file_extension in ['.xlsx', '.xls']:
                # Handle Excel files with openpyxl
                wb = load_workbook(filename=file_path)
                sheet = wb.active

                # Get year from first cell
                year_cell = sheet['A1'].value
                if not isinstance(year_cell, int) or len(str(year_cell)) != 4:
                    raise ValueError("First cell must contain a 4-digit year (e.g., 2025)")
                excel_year = str(year_cell)

                # Process date cells
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    cell_value = row[0]
                    if not cell_value:
                        break

                    # Handle different cell types
                    if isinstance(cell_value, datetime):
                        date_obj = cell_value
                    else:
                        try:
                            # Try parsing string format
                            if isinstance(cell_value, str):
                                # Try multiple date formats
                                date_formats = ["%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y", "%d/%m/%Y"]
                                date_obj = None
                                for fmt in date_formats:
                                    try:
                                        date_obj = datetime.strptime(cell_value, fmt)
                                        break
                                    except ValueError:
                                        continue

                                if date_obj is None:
                                    raise ValueError(f"Unable to parse date format: {cell_value}")
                            else:
                                # Convert other types to string and try parsing
                                date_obj = datetime.strptime(str(cell_value), "%Y-%m-%d %H:%M:%S")
                        except ValueError as ve:
                            raise ValueError(f"Invalid date format: {cell_value} - {str(ve)}")

                    # Validate year match
                    if str(date_obj.year) != excel_year:
                        raise ValueError(f"Date {date_obj.date()} doesn't match file year {excel_year}")

                    formatted_date = date_obj.strftime("%d-%m-%Y")
                    if formatted_date not in holidays:  # Avoid duplicates
                        holidays.append(formatted_date)

            elif file_extension == '.numbers':
                # Handle Numbers files
                try:
                    # First try: attempt direct reading with pandas
                    try:
                        df = pd.read_excel(file_path, header=None, engine='openpyxl')
                    except:
                        # Second try: read as CSV if Numbers exported as such
                        df = pd.read_csv(file_path, header=None)

                    if df.empty:
                        raise ValueError("File appears to be empty")

                    # Get year from first cell
                    year_cell = df.iloc[0, 0]
                    if not str(year_cell).isdigit() or len(str(year_cell)) != 4:
                        raise ValueError("First cell must contain a 4-digit year (e.g., 2025)")
                    excel_year = str(year_cell)

                    # Process dates
                    for idx in range(1, len(df)):
                        cell_value = df.iloc[idx, 0]

                        if pd.isna(cell_value):
                            continue

                        try:
                            # Handle different data types
                            if isinstance(cell_value, str):
                                # Try multiple date formats
                                date_formats = ["%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y", "%d/%m/%Y",
                                                "%Y/%m/%d"]
                                date_obj = None
                                for fmt in date_formats:
                                    try:
                                        date_obj = datetime.strptime(cell_value, fmt)
                                        break
                                    except ValueError:
                                        continue

                                if date_obj is None:
                                    raise ValueError(f"Unable to parse date format: {cell_value}")

                            elif isinstance(cell_value, pd.Timestamp):
                                date_obj = cell_value.to_pydatetime()
                            elif hasattr(cell_value, 'date'):
                                date_obj = cell_value
                            else:
                                # Try to convert to datetime
                                date_obj = pd.to_datetime(cell_value).to_pydatetime()

                            # Validate year match
                            if str(date_obj.year) != excel_year:
                                raise ValueError(f"Date {date_obj.date()} doesn't match file year {excel_year}")

                            formatted_date = date_obj.strftime("%d-%m-%Y")
                            if formatted_date not in holidays:  # Avoid duplicates
                                holidays.append(formatted_date)

                        except Exception as e:
                            raise ValueError(f"Error processing row {idx + 1}: {str(e)}")

                except Exception as numbers_error:
                    # If all reading attempts fail, show conversion instructions
                    conversion_msg = ("Unable to read Numbers file directly.\n\n"
                                      "Please convert to Excel format:\n"
                                      "1. Open your Numbers file\n"
                                      "2. Go to File → Export To → Excel...\n"
                                      "3. Save as .xlsx format\n"
                                      "4. Use the exported .xlsx file\n\n"
                                      f"Technical error: {str(numbers_error)}")

                    reply = QMessageBox.question(self, "Numbers File Conversion Required",
                                                 "Numbers file could not be read directly.\n\n"
                                                 "Would you like to see conversion instructions?",
                                                 QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)

                    if reply == QMessageBox.StandardButton.Yes:
                        QMessageBox.information(self, "Conversion Instructions", conversion_msg)

                    return

            else:
                raise ValueError(f"Unsupported file format: {file_extension}")

            if not holidays:
                QMessageBox.warning(self, "No Data", "No valid holiday dates found in the file")
                return

            # Check for existing year in database
            cursor = self.db_connection.cursor()
            cursor.execute("SELECT year FROM holiday WHERE year = ?", (excel_year,))
            exists = cursor.fetchone()

            if exists:
                confirm = QMessageBox.question(self, "Override Confirmation",
                                               f"Holidays for {excel_year} already exist. Override?",
                                               QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
                if confirm != QMessageBox.StandardButton.Yes:
                    return

            # Insert/Update data
            cursor.execute("INSERT OR REPLACE INTO holiday (year, holidays) VALUES (?, ?)",
                           (excel_year, json.dumps(holidays)))
            self.db_connection.commit()

            file_type = "Excel" if file_extension in ['.xlsx', '.xls'] else "Numbers"
            QMessageBox.information(self, "Import Successful",
                                    f"Successfully imported {len(holidays)} holidays for {excel_year} from {file_type} file")

        except Exception as e:
            QMessageBox.critical(self, "Import Error", f"Failed to import holidays:\n{str(e)}")
        finally:
            if 'cursor' in locals() and cursor:
                cursor.close()

    def update_database_page(self, all_tables, new_tables=None):
        """Update the database page with current table information"""
        self.db_table_list.clear()
        self.db_table_list.addItems(all_tables)

        status_text = f"Loaded {len(all_tables)} tables"
        if new_tables:
            status_text += f"\nCreated new tables: {', '.join(new_tables)}"

        # Show status message in home page for demo
        if hasattr(self, 'home_status_label'):
            self.home_status_label.setText(status_text)

    def create_home_page(self):
        page = QWidget()
        main_layout = QVBoxLayout(page)
        main_layout.setContentsMargins(20, 15, 20, 15)
        main_layout.setSpacing(15)

        # Year/Month Selection Group
        year_month_group = QGroupBox("Select Month & Year")
        year_month_group.setStyleSheet("""
                QGroupBox {
                    border: 1px solid #e0e0e0;
                    border-radius: 6px;
                    margin-top: 8px;
                    padding-top: 12px;
                }
                QGroupBox::title {
                    subcontrol-origin: margin;
                    left: 10px;
                    color: #118370;
                    font-weight: 500;
                }
            """)

        # Layout for month and year
        date_layout = QHBoxLayout(year_month_group)
        date_layout.setContentsMargins(10, 15, 10, 15)
        date_layout.setSpacing(15)

        # Month ComboBox
        self.month_combo = QComboBox()
        self.month_combo.addItems(
            ["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November",
             "December"])
        self.month_combo.setFixedHeight(35)

        # Year ComboBox
        self.year_combo = QComboBox()
        current_year = QDate.currentDate().year()
        self.year_combo.addItems([str(y) for y in range(2022, 2051)])
        self.year_combo.setFixedHeight(35)

        # Set current date
        current_month = QDate.currentDate().month()
        self.month_combo.setCurrentIndex(current_month - 1)
        self.year_combo.setCurrentText(str(current_year))

        # Function to update month combo box based on selected year
        def update_month_combo():
            selected_year = int(self.year_combo.currentText())
            is_current_year = selected_year == current_year

            # If switching back to the current year, reset month to the current month
            if is_current_year:
                self.month_combo.setCurrentIndex(current_month - 1)

            for idx in range(self.month_combo.count()):
                month_enabled = (idx + 1) <= current_month if is_current_year else True
                self.month_combo.model().item(idx).setEnabled(month_enabled)

        # Disable future years
        for idx in range(self.year_combo.count()):
            year = int(self.year_combo.itemText(idx))
            self.year_combo.model().item(idx).setEnabled(year <= current_year)

        # Connect the year selection change to update months dynamically
        self.year_combo.currentTextChanged.connect(update_month_combo)

        # Initial check on startup
        update_month_combo()

        # Add widgets to layout
        date_layout.addWidget(self.month_combo, 70)  # 70% width
        date_layout.addWidget(self.year_combo, 30)  # 30% width

        # Styling
        combo_style = """
            QComboBox {
                padding: 8px;
                border: 1px solid #cccccc;
                border-radius: 4px;
                background-color: white;
            }
            QComboBox::drop-down {
                subcontrol-origin: padding;
                subcontrol-position: top right;
                width: 25px;
            }
        """
        self.month_combo.setStyleSheet(combo_style)
        self.year_combo.setStyleSheet(combo_style)

        # Holiday Information Section
        holiday_group = QGroupBox("Holiday Management")
        holiday_group.setStyleSheet("""
                QGroupBox {
                    border: 1px solid #e0e0e0;
                    border-radius: 6px;
                    margin-top: 8px;
                    padding-top: 12px;
                }
                QGroupBox::title {
                    color: #118370;
                    font-weight: 500;
                }
            """)
        holiday_layout = QVBoxLayout(holiday_group)
        holiday_layout.setContentsMargins(10, 15, 10, 10)
        holiday_layout.setSpacing(8)

        # Input row with buttons
        input_row = QHBoxLayout()
        input_row.setSpacing(8)

        # Input field
        self.holiday_input = QPlainTextEdit()
        self.holiday_input.setFixedHeight(35)
        self.holiday_input.setReadOnly(True)
        self.holiday_input.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)

        # Buttons
        self.load_holiday_btn = QPushButton("Load Holiday")
        self.load_holiday_btn.setToolTip("Please Upload Excel file as per the format")
        self.load_holiday_btn.setFixedSize(30, 30)

        self.view_holiday_btn = QPushButton("View Holidays")
        self.view_holiday_btn.setFixedSize(120, 35)

        # New action button
        self.format_info_btn = QPushButton("Format")
        self.format_info_btn.setToolTip("Holiday Excel Format")
        self.format_info_btn.setFixedSize(100, 35)

        input_row.addWidget(self.holiday_input)
        input_row.addWidget(self.load_holiday_btn)
        input_row.addWidget(self.view_holiday_btn)
        input_row.addWidget(self.format_info_btn)  # Added new button

        # Connect buttons
        self.format_info_btn.clicked.connect(self.show_format_guide)
        self.load_holiday_btn.clicked.connect(self.load_holidays_to_db)
        self.view_holiday_btn.clicked.connect(self.show_holiday_viewer)

        # When creating buttons
        self.format_info_btn.setObjectName("format_info_btn")
        self.load_holiday_btn.setObjectName("load_holiday_btn")
        self.view_holiday_btn.setObjectName("view_holiday_btn")

        holiday_layout.addLayout(input_row)

        # Error message label
        self.holiday_error_label = QLabel()
        self.holiday_error_label.setStyleSheet("color: red; font-size: 11px;")
        self.holiday_error_label.setWordWrap(True)
        holiday_layout.addWidget(self.holiday_error_label)

        # Styling for new button
        button_style = """
                QPushButton {
                    border-radius: 4px;
                    padding: 6px;
                    font-weight: 500;
                }
            """
        self.format_info_btn.setStyleSheet(button_style + "background-color: #aae0a4; color: black;")

        # Category Information Section
        category_group = QGroupBox("User Categories")
        category_group.setStyleSheet("""
                QGroupBox {
                    border: 1px solid #e0e0e0;
                    border-radius: 6px;
                    margin-top: 8px;
                    padding-top: 12px;
                }
                QGroupBox::title {
                    color: #118370;
                    font-weight: 500;
                }
            """)
        category_layout = QVBoxLayout(category_group)
        category_layout.setContentsMargins(10, 15, 10, 10)
        category_layout.setSpacing(8)

        # Input row with buttons
        catergory_input_row = QHBoxLayout()
        catergory_input_row.setSpacing(8)

        # Input field
        self.category_input = QPlainTextEdit()
        self.category_input.setFixedHeight(35)
        self.category_input.setReadOnly(True)
        self.category_input.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)

        # Buttons
        self.category_btn = QPushButton("Load Category")
        self.category_btn.setToolTip("Please Upload Excel file as per the format")
        self.category_btn.setFixedSize(30, 30)

        catergory_input_row.addWidget(self.category_input)
        catergory_input_row.addWidget(self.category_btn)

        # Connect buttons
        self.category_btn.clicked.connect(self.select_category)

        # When creating buttons
        self.category_btn.setObjectName("category_btn")
        category_layout.addLayout(catergory_input_row)

        # Main group box
        groupBox = QGroupBox("Please enter raw attendance file to generate Attendance Billing Report")
        groupBox.setStyleSheet("""
                QGroupBox {
                border: 1px solid #e0e0e0;
                border-radius: 6px;
                margin-top: 8px;
                padding-top: 12px;
            }
            QGroupBox::title {
                color: #118370;
                font-weight: 500;
            }
            # QGroupBox::title {
            #     subcontrol-origin: margin;
            #     left: 10px;
            #     padding: 0 3px;
            # }
        """)
        group_layout = QVBoxLayout(groupBox)
        group_layout.setContentsMargins(10, 15, 10, 10)  # Reduced top/bottom margins (was 10,15,10,10)
        group_layout.setSpacing(8)  # Reduced spacing between widgets

        # File input box
        main_input_row = QHBoxLayout()
        main_input_row.setSpacing(8)

        # Input field
        self.main_input = QPlainTextEdit()
        self.main_input.setFixedHeight(35)
        self.main_input.setReadOnly(True)
        self.main_input.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)

        # Upload button
        self.upload_button = QPushButton("Upload")
        self.upload_button.setIcon(QIcon.fromTheme("folder"))
        self.upload_button.setToolTip("Select file")

        main_input_row.addWidget(self.main_input)
        main_input_row.addWidget(self.upload_button)

        # Connect buttons
        self.upload_button.clicked.connect(self.upload_file)

        # Add to main group
        # When creating buttons
        self.upload_button.setObjectName("upload_button")
        group_layout.addLayout(main_input_row)

        # Status message system
        # Create the GroupBox
        statusBox = QGroupBox("Status Message")
        statusBox.setStyleSheet("""
                        QGroupBox {
                        border: 1px solid #e0e0e0;
                        border-radius: 6px;
                        margin-top: 8px;
                        padding-top: 12px;
                    }
                    QGroupBox::title {
                        color: #118370;
                        font-weight: 500;
                    }
                    # QGroupBox::title {
                    #     subcontrol-origin: margin;
                    #     left: 10px;
                    #     padding: 0 3px;
                    # }
                """)
        status_group_box = QVBoxLayout(statusBox)
        status_group_box.setContentsMargins(10, 15, 10, 10)  # Reduced top/bottom margins (was 10,15,10,10)
        status_group_box.setSpacing(8)  # Reduced spacing between widgets

        # File input box
        status_input_row = QHBoxLayout()
        status_input_row.setSpacing(8)

        self.msg_label = QLabel()
        self.msg_label.setObjectName("statusMsg")
        self.msg_label.setStyleSheet("""
                QLabel {
                    padding: 10px;
                    border-radius: 4px;
                    margin: 8px 0;
                    font-size: 12px;
                    min-height: 40px;
                }
                QLabel[messageType="error"] {
                    color: #dc3545;
                    background-color: #f8d7da;
                    border: 1px solid #f5c6cb;
                }
                QLabel[messageType="success"] {
                    color: #28a745;
                    background-color: #d4edda;
                    border: 1px solid #c3e6cb;
                }
                QLabel[messageType="info"] {
                    color: #004085;
                    background-color: #cce5ff;
                    border: 1px solid #b8daff;
                }
            """)
        self.msg_label.setWordWrap(True)
        self.msg_label.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        self.msg_label.hide()

        status_input_row.addWidget(self.msg_label)
        status_group_box.addLayout(status_input_row)

        # # Show error message (red)
        # self.show_message("Error: File not found!", "error", 5000)
        #
        # # Show success message (green)
        # self.show_message("Data loaded successfully!", "success", 3000)
        #
        # # Show info message (blue)
        # self.show_message("Processing completed", "info", 2000)

        # Progress Bar
        # progress_input_row = QHBoxLayout()
        # progress_input_row.setSpacing(8)
        progress_group = QWidget()
        progress_layout = QVBoxLayout(progress_group)
        progress_layout.setContentsMargins(0, 10, 0, 0)

        self.progress_bar = QProgressBar()
        self.progress_bar.setObjectName("progress_bar")
        self.progress_bar.setValue(0)
        self.progress_bar.setTextVisible(True)
        self.progress_bar.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.progress_bar.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        self.progress_bar.setStyleSheet("""
                QProgressBar {
                    height: 20px;
                    border: 1px solid #cccccc;
                    border-radius: 4px;
                    text-align: center;
                }
                QProgressBar::chunk {
                    background-color: #118370;
                    border-radius: 3px;
                }
            """)

        # progress_input_row.addWidget(self.progress_bar)
        progress_layout.addWidget(self.progress_bar)

        # Generate Button
        generate_container = QHBoxLayout()
        generate_container.addStretch()  # Left stretch

        self.generate_button = QPushButton("Generate Report")
        self.generate_button.setObjectName("generate_button")
        self.generate_button.setFixedHeight(35)
        self.generate_button.setFixedSize(280, 40)  # Slightly larger for emphasis
        self.generate_button.setStyleSheet("""
                QPushButton {
                    background-color: #118370;
                    color: white;
                    border-radius: 4px;
                    padding: 8px;
                    font-weight: 500;
                }
                QPushButton:hover {
                    background-color: #0f7460;
                }
                QPushButton:disabled {
                    background-color: #cccccc;
                    color: #666666;
                }
            """)
        generate_container.addWidget(self.generate_button)
        generate_container.addStretch()  # Right stretch

        progress_layout.addLayout(generate_container)

        # Create button container for alignment
        button_container = QHBoxLayout()
        button_container.addStretch()
        button_container.addWidget(self.generate_button)
        self.generate_button.clicked.connect(self.generate_report)

        # Add to main layout without stretching
        main_layout.addWidget(year_month_group)
        main_layout.addWidget(holiday_group)
        main_layout.addWidget(category_group)
        # Add main group to page
        main_layout.addWidget(groupBox)

        main_layout.addWidget(self.progress_bar)
        main_layout.addLayout(button_container)
        # main_layout.addWidget(self.msg_label)
        main_layout.addWidget(statusBox)

        main_layout.addStretch()
        # Update progress
        self.progress_bar.setValue(0)  # 0-100%
        # Set fixed height for message label to prevent layout shift
        self.msg_label.setFixedHeight(0)

        def adjust_message_height():
            if self.msg_label.isVisible():
                self.msg_label.setFixedHeight(self.msg_label.sizeHint().height())
            else:
                self.msg_label.setFixedHeight(0)

        self.msg_label.showEvent = lambda e: adjust_message_height()
        self.msg_label.hideEvent = lambda e: adjust_message_height()

        return page

    def show_message(self, text, msg_type="info", timeout=5000):
        """Show dynamic message with auto-hide"""
        self.msg_label.setProperty("messageType", msg_type)
        self.msg_label.setText(text)
        self.msg_label.style().polish(self.msg_label)

        # Adjust height dynamically
        self.msg_label.setFixedHeight(self.msg_label.sizeHint().height())

        self.msg_label.show()

        # Auto-hide after timeout
        if timeout > 0:
            QTimer.singleShot(timeout, self.clear_message)

    def clear_message(self):
        """Clear the status message"""
        self.msg_label.hide()
        self.msg_label.setText("")

    def get_holidays_for_year(self, year):
        """
        Fetch holidays for a specific year from the database
        Returns:
            list: List of holidays (YYYY-MM-DD format) or None if error occurs
            None: Returns None if database error occurs
        """
        try:
            cursor = self.db_connection.cursor()
            cursor.execute("SELECT holidays FROM holiday WHERE year = ?", (str(year),))
            result = cursor.fetchone()

            if result:
                return json.loads(result[0])
            return []  # Return empty list if no entries found

        except sqlite3.Error as e:
            print(f"Database error fetching holidays: {str(e)}")
            return None
        except json.JSONDecodeError as e:
            print(f"Error parsing holiday data: {str(e)}")
            return None
        finally:
            if cursor:
                cursor.close()

    def clean_keys(self, dict_list):
        cleaned_list = []
        for data_dict in dict_list:
            cleaned_dict = {}
            for key, value in data_dict.items():
                # Clean the key by removing '\n' and extra spaces
                clean_key = key.replace("\n", " ").strip()
                cleaned_dict[clean_key] = value
            cleaned_list.append(cleaned_dict)
        return cleaned_list

    def fetch_all_resource_mappings(self):
        """Fetch all records from the resource_mapping table"""
        try:
            cursor = self.db_connection.cursor()
            cursor.execute("SELECT * FROM resource_mapping")
            records = cursor.fetchall()

            # Get column names
            columns = [desc[0] for desc in cursor.description]

            # Convert to a list of dictionaries
            result = [dict(zip(columns, row)) for row in records]

            # Convert the DataFrame to a list of dictionaries
            self.raw_category_list = result

            # Standardize the "Full Name" field by removing commas and spaces
            for item in self.raw_category_list:
                name = item['full_name']
                team = item["team"]

                # Check if the team is already in the dictionary
                if team in self.categories:
                    self.categories[team].append(name)
                else:
                    self.categories[team] = [name]

                self.name_mapping.update(
                    {name: [item['id_521'], item['point_of_contact'], item['start_date'], item['end_date']]})

            for k, v in self.categories.items():
                temp_list = sorted(v)
                self.categories[k] = temp_list
                self.name_order_list.extend(temp_list)

            return True
        except sqlite3.Error as e:
            print(f"Database error: {e}")
            return False
        finally:
            if cursor:
                cursor.close()

    def non_compliance_resources(self, data, filename="non_complaint_user.xlsx"):

        # Create a new workbook and select the active worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance Data"

        # Define the header names and order
        headers = ["Name", "Month", "Listed Month Holiday", "Attendance Marked on Holiday"]

        # Set header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

        # Write the headers to the worksheet
        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

            # Set column width
            # Set column width; assign a larger width for the "Name" column
            if header == "Name":
                column_width = 40  # Increase this value as needed
            else:
                column_width = max(len(header), 20)
            ws.column_dimensions[get_column_letter(col_num)].width = column_width

        # Write the data to the worksheet
        for row_num, entry in enumerate(data, start=2):
            ws.cell(row=row_num, column=1, value=entry["Name"])
            ws.cell(row=row_num, column=2, value=entry["Month"])
            ws.cell(row=row_num, column=3, value=", ".join(entry["Listed Month Holiday"]))
            ws.cell(row=row_num, column=4, value=", ".join(entry["Attendance Marked on Holiday"]))

            # Set alignment for data cells
            for col_num in range(1, 5):
                cell = ws.cell(row=row_num, column=col_num)
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Save the workbook to the specified filename
        wb.save(filename)
        print(f"Data written to {filename} with formatting.")

    def add_summary_page(self, data, filename="my_workbook.xlsx"):
        """
        Creates a new sheet called 'Summary' at index 0 in the workbook.
        Writes headers from B4 and then data from row 5 onward.
        Adjusts column width automatically, adds a 'Total' row at the end,
        aligns only the Name column to left (others center) and applies borders
        to the entire table.
        """
        # Load existing workbook
        wb = load_workbook(filename)
        # Create a new sheet at index 0
        sheet = wb.create_sheet("Summary", 0)
        # Set tab color
        sheet.sheet_properties.tabColor = "34b1eb"

        # Define a thin border style for the entire table
        thin_side = Side(border_style="thin", color="000000")
        cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        # ------------------------------------------------------
        # 1) Write Headers at Row 4 (B4, C4, D4)
        # ------------------------------------------------------
        headers = {2: "Name", 3: "Total Number of Billable Days", 4: "Leave Days"}

        for col_idx, header_text in headers.items():
            cell = sheet.cell(row=4, column=col_idx, value=header_text)
            cell.font = Font(bold=True, color="111212")
            cell.fill = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
            # Only "Name" header is left-aligned; others are center-aligned
            if col_idx == 2:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = cell_border

        # ------------------------------------------------------
        # 2) Write Data Starting from Row 5
        # ------------------------------------------------------
        start_row = 5
        for idx, entry in enumerate(data, start=start_row):
            # Column B: Name (left-aligned)
            sheet_name = entry.get("Name", "")
            cell_B = sheet.cell(row=idx, column=2, value=sheet_name)
            if sheet_name:
                cell_B.hyperlink = f"#'{sheet_name}'!A1"
                cell_B.style = "Hyperlink"
                cell_B.font = Font(color="000000")
            cell_B.alignment = Alignment(horizontal="left", vertical="center")
            cell_B.border = cell_border

            # Column C: Total Number of Billable Days (center-aligned)
            cell_C = sheet.cell(row=idx, column=3, value=entry.get("Total Number of Billable Days", 0))
            cell_C.alignment = Alignment(horizontal="center", vertical="center")
            cell_C.border = cell_border

            # Column D: Service Credit Pool Days (center-aligned)
            cell_D = sheet.cell(row=idx, column=4, value=entry.get("Service Credit Pool Days", 0))
            cell_D.alignment = Alignment(horizontal="center", vertical="center")
            cell_D.border = cell_border

        # ------------------------------------------------------
        # 3) Add a "Total" Row
        # ------------------------------------------------------
        last_data_row = start_row + len(data) - 1
        total_row = last_data_row + 1  # One row below the last data row

        # Write "Total" in column B (center-aligned for total row)
        total_cell_B = sheet.cell(row=total_row, column=2, value="Total")
        total_cell_B.alignment = Alignment(horizontal="center", vertical="center")
        total_cell_B.border = cell_border

        # Sum formula for column C
        total_cell_C = sheet.cell(row=total_row, column=3,
            value=f"=SUM({get_column_letter(3)}{start_row}:{get_column_letter(3)}{last_data_row})")
        total_cell_C.alignment = Alignment(horizontal="center", vertical="center")
        total_cell_C.border = cell_border

        # Sum formula for column D
        total_cell_D = sheet.cell(row=total_row, column=4,
            value=f"=SUM({get_column_letter(4)}{start_row}:{get_column_letter(4)}{last_data_row})")
        total_cell_D.alignment = Alignment(horizontal="center", vertical="center")
        total_cell_D.border = cell_border

        # Apply fill for total row
        total_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        for col_num in range(2, 5):
            cell = sheet.cell(row=total_row, column=col_num)
            cell.fill = total_fill

        # ------------------------------------------------------
        # 4) Auto-Adjust Column Widths Based on Longest Value
        # ------------------------------------------------------
        max_lengths = {col_idx: 0 for col_idx in headers}

        # Check headers
        for col_idx, header_text in headers.items():
            max_lengths[col_idx] = max(max_lengths[col_idx], len(str(header_text)))

        # Check data rows
        for row_idx in range(start_row, last_data_row + 1):
            for col_idx in range(2, 5):
                cell_value = sheet.cell(row=row_idx, column=col_idx).value
                if cell_value is not None:
                    max_lengths[col_idx] = max(max_lengths[col_idx], len(str(cell_value)))

        # Check the "Total" row
        max_lengths[2] = max(max_lengths[2], len("Total"))
        for col_idx in [3, 4]:
            formula_text = sheet.cell(row=total_row, column=col_idx).value
            max_lengths[col_idx] = max(max_lengths[col_idx], len(str(formula_text)))

        # Set column widths
        for col_idx in range(2, 5):
            col_letter = get_column_letter(col_idx)
            sheet.column_dimensions[col_letter].width = max_lengths[col_idx] + 2

        # Save the workbook
        wb.save(filename)
        print(f"Summary sheet created and data written to '{filename}'")

    def generate_report(self):

        if not all([self.raw_category_list, self.categories, self.name_mapping, self.name_order_list]):
            if not (result := self.fetch_all_resource_mappings()):
                self.show_message("Error: Please select proper category file!", "error", 5000)
                return None

        # Implement report generation logic here
        # self.month_combo.currentText()
        self.selected_month = self.month_combo.currentText()
        self.selected_year = self.year_combo.currentText()

        self.HOLIDAY_LIST = self.get_holidays_for_year(self.selected_year)

        if not self.HOLIDAY_LIST:
            # Show error message (red)
            self.show_message(f"Error: Please load the holiday corresponding to the year {self.selected_year}.", "error",
                              5000)
            return None

        if self.df:
            self.df = self.clean_keys(self.df)

            for category, valid_rsnames in self.categories.items():
                self.file_name = f"{category}_Timesheet_{self.selected_month} {self.selected_year}.xlsx"
                # Filter df based on Rsname matching any valid_rsnames with 100% coverage
                filtered_df = [record for record in self.df if any(
                    coverage_percentage(clean_string(record["Rsname"]), clean_string(valid_rs)) >= 60 for valid_rs in
                    valid_rsnames)]
                if filtered_df:
                    status, response, user_data, non_complaince_resources, user_leave_record = (
                    generate_excel(self.selected_month, self.selected_year, self.file_name, filtered_df, self.HOLIDAY_LIST,
                        self.name_mapping, self.name_order_list, self.progress_bar))
                    if status == 200:
                        remaining_val = 100 - self.progress_bar.value()

                        step = remaining_val / 3
                        if self.categories:
                            # self.add_category_data(user_data)
                            self.progress_bar.setValue(self.progress_bar.value() + int(step))
                            self.add_summary_page(user_data, self.file_name)
                            self.progress_bar.setValue(self.progress_bar.value() + int(step))
                    else:
                        self.show_message(f"Error: {response}", "error", 5000)

                    if non_complaince_resources:
                        self.non_compliance_resources(non_complaince_resources,
                                                      f"{category}_non_complaint_{self.selected_month} {self.selected_year}.xlsx")
                        self.update_non_complaint_user(non_complaince_resources)

                    if user_leave_record:
                        self.update_user_leave(user_leave_record)
                else:
                    print(f"No Data for the category - {category}")

            # sys.exit(1)
            self.show_message("Report Generated Successfully!", "success", 10000)
            self.progress_bar.setValue(100)
        else:
            self.show_message(f"Error: Please provide raw excel file as an input.", "error", 5000)

    def update_user_leave(self, data):
        """
        Updates the user_leave table with the provided data.
        Data is expected as a list of dictionaries in the following format:
          [{"name": sheet_name, "id_521": details[0], "year": year, "month": month, "leave_days": leave_dates}, ...]

        The function checks for an existing record using either (name, year, month) or (id_521, year, month).
        If a record exists, it updates it; otherwise, it inserts a new record.
        """
        try:
            cursor = self.db_connection.cursor()

            for d in data:
                name = d.get("name", "")
                id_521 = d.get("id_521", "")
                year = str(d.get("year", ""))
                month = d.get("month", "")
                leave_days = ",".join(str(x) for x in d.get("leave_days", [])) if d.get("leave_days") else ""

                try:
                    # Check if a record exists using either (name, year, month) or (id_521, year, month)
                    cursor.execute("""
                        SELECT COUNT(*) FROM user_leave 
                        WHERE (name=? AND year=? AND month=?) OR (id_521=? AND year=? AND month=?)
                        """, (name, year, month, id_521, year, month))
                    exists = cursor.fetchone()[0]
                except sqlite3.Error as e:
                    print(f"Error checking existence for {name}: {e}")
                    continue  # Skip this record if an error occurs

                if exists:
                    try:
                        # Update the existing record
                        cursor.execute("""
                            UPDATE user_leave
                            SET id_521 = ?, leave_days = ?
                            WHERE (name=? AND year=? AND month=?) OR (id_521=? AND year=? AND month=?)
                            """, (id_521, leave_days, name, year, month, id_521, year, month))
                    except sqlite3.Error as e:
                        print(f"Error updating record for {name}: {e}")
                        continue
                else:
                    try:
                        # Insert a new record
                        cursor.execute("""
                            INSERT INTO user_leave (name, id_521, year, month, leave_days)
                            VALUES (?, ?, ?, ?, ?)
                            """, (name, id_521, year, month, leave_days))
                    except sqlite3.Error as e:
                        print(f"Error inserting record for {name}: {e}")
                        continue

            self.db_connection.commit()
        except sqlite3.Error as e:
            print(f"Database error: {e}")
            self.db_connection.rollback()
        finally:
            cursor.close()

    def update_non_complaint_user(self, data):
        """
        Updates the non_complaint_user table with the new columns:
        observed_leave_count, observed_leave_dates, month_holiday_count, and month_holiday_dates.

        For each record in the data list:
          - observed_leave_count is the number of entries in "Attendance Marked on Holiday"
          - observed_leave_dates is a comma-separated list from "Attendance Marked on Holiday"
          - month_holiday_count is the number of entries in "Listed Month Holiday"
          - month_holiday_dates is a comma-separated list from "Listed Month Holiday"

        The function checks if a record exists using either (name, year, month) or (id_521, year, month).
        If a record exists, it updates it; otherwise, it inserts a new record.
        """
        try:
            cursor = self.db_connection.cursor()

            for d in data:
                name = d.get("Name", "")
                id_521 = d.get("521_ID", "")
                year = str(d.get("Year", ""))
                month = d.get("Month", "")

                # Get lists of holidays and attendance marked dates
                listed_holidays = d.get("Listed Month Holiday", [])
                attendance_marked = d.get("Attendance Marked on Holiday", [])

                # Prepare new column values
                observed_leave_count = str(len(attendance_marked)) if attendance_marked else "0"
                observed_leave_dates = ",".join(attendance_marked) if attendance_marked else ""
                month_holiday_count = str(len(listed_holidays)) if listed_holidays else "0"
                month_holiday_dates = ",".join(listed_holidays) if listed_holidays else ""

                try:
                    # Check if a record already exists (by either key combination)
                    cursor.execute("""
                        SELECT COUNT(*) FROM non_complaint_user 
                        WHERE (name=? AND year=? AND month=?) OR (id_521=? AND year=? AND month=?)
                        """, (name, year, month, id_521, year, month))
                    exists = cursor.fetchone()[0]
                except sqlite3.Error as e:
                    print(f"Error checking existence for {name}: {e}")
                    continue  # Skip this record if an error occurs

                if exists:
                    try:
                        # Update existing record
                        cursor.execute("""
                            UPDATE non_complaint_user
                            SET id_521 = ?,
                                observed_leave_count = ?,
                                observed_leave_dates = ?,
                                month_holiday_count = ?,
                                month_holiday_dates = ?
                            WHERE (name=? AND year=? AND month=?) OR (id_521=? AND year=? AND month=?)
                            """, (
                        id_521, observed_leave_count, observed_leave_dates, month_holiday_count, month_holiday_dates, name,
                        year, month, id_521, year, month))
                    except sqlite3.Error as e:
                        print(f"Error updating record for {name}: {e}")
                        continue
                else:
                    try:
                        # Insert new record
                        cursor.execute("""
                            INSERT INTO non_complaint_user 
                            (name, id_521, year, month, observed_leave_count, observed_leave_dates, month_holiday_count, month_holiday_dates)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                            """, (name, id_521, year, month, observed_leave_count, observed_leave_dates, month_holiday_count,
                                  month_holiday_dates))
                    except sqlite3.Error as e:
                        print(f"Error inserting record for {name}: {e}")
                        continue

            self.db_connection.commit()
        except sqlite3.Error as e:
            print(f"Database error: {e}")
            self.db_connection.rollback()
        finally:
            cursor.close()

    def upload_file(self):
        file_dialog = QFileDialog(self)
        filepath, _ = file_dialog.getOpenFileName(self, "Open Excel File", "", "Excel Files (*.xlsx *.xls, *.csv)")

        if filepath:
            fileInfo = QtCore.QFileInfo(filepath)
            file_name = fileInfo.fileName()
            file_size = fileInfo.size()  # in bytes
            # Convert file size to kilobytes
            file_size_kb = file_size / 1024.0
            print(f"File Name: {file_name}, File Size: {file_size_kb:.2f} KB")
            self.main_input.setPlainText(f"{file_name} ({file_size_kb:.2f} KB)")

            # Create DataFrame with Pandas
            self.df = read_file(filepath)

            # Adding Validation to file upload
            try:
                dict_value = dict(Counter(list(itertools.chain.from_iterable(
                    [[item.split("-")[-1] for item in j] for j in [list(i.keys())[4:-2] for i in self.df]]))))
                value = max(dict_value, key=dict_value.get)

                self.selected_month = self.month_combo.currentText()

                if value == self.selected_month[:3]:
                    self.show_message("Valid Raw Excel Loaded", "info", 4000)
                else:
                    self.show_message("Invalid Raw excel, please check the file or selected month.", "error", 5000)
                    self.main_input.setPlainText("")
                    self.df = None  # # Update progress  # self.progress_bar.setValue(0)  # 0-100%  #  # # Show/hide when needed  # self.progress_bar.setVisible(True)

                # Reset on completion  # self.progress_bar.reset()
            except Exception as e:
                print(e)
                self.show_message("Invalid Raw excel, please check the file.", "error", 5000)
                self.df = None

    def select_category(self):
        file_dialog = QFileDialog(self)
        filepath, _ = file_dialog.getOpenFileName(self, "Open Excel File", "", "Excel Files (*.xlsx *.xls, *.csv)")

        self.raw_category_list, self.name_order_list = ([],) * 2
        self.categories = dict()
        self.name_mapping = dict()

        try:
            if filepath:
                fileInfo = QtCore.QFileInfo(filepath)
                file_name = fileInfo.fileName()
                file_size = fileInfo.size()  # in bytes
                # Convert file size to kilobytes
                file_size_kb = file_size / 1024.0
                print(f"File Name: {file_name}, File Size: {file_size_kb:.2f} KB")

                self.category_input.setPlainText(f"{file_name} ({file_size_kb:.2f} KB)")

                sheet_name = "PublicCloudResourceList"
                # Read the Excel file into a DataFrame
                df = pd.read_excel(filepath, sheet_name=sheet_name)
                df.replace({np.nan: None}, inplace=True)
                self.add_data_resource_tab(df)
                # Convert the DataFrame to a list of dictionaries
                self.raw_category_list = df.to_dict(orient="records")

                # Standardize the "Full Name" field by removing commas and spaces
                for item in self.raw_category_list:
                    name = clean_string(item['Full Name'])
                    team = item["Team"]

                    # Check if the team is already in the dictionary
                    if team in self.categories:
                        self.categories[team].append(name)
                    else:
                        self.categories[team] = [name]

                    self.name_mapping.update(
                        {name: [item["521 ID"], item["Point of Contact"], item["Start Date"], item["End Date"]]})

                # Create a mapping from Full Name to 521 ID for quick lookup
                # self.name_mapping = {
                #     preprocess_name(item["Full Name"]): [item["521 ID"], item["Point of Contact"], item["Start Date"],
                #         item["End Date"], ] for item in self.raw_category_list}

                for k, v in self.categories.items():
                    temp_list = sorted(v)
                    self.categories[k] = temp_list
                    self.name_order_list.extend(temp_list)
                # QMessageBox.information(self, "Success", "Category data successfully created!",
                #     QMessageBox.StandardButton.Ok)
                self.show_message("Category data successfully created!", "success", 3000)

        except Exception as e:
            self.raw_category_list, self.name_order_list = [], []  # ✅ Separate lists
            self.categories, self.name_mapping = {}, {}  # ✅ Separate dictionaries
            self.category_input.setPlainText(f"")
            self.show_message(f"Error: Not a valid category file", "error", 5000)

    def show_format_guide(self):
        """Show Excel format requirements"""
        guide_text = """
                <html>
                    <head>
                        <style>
                            body {
                                font-family: Arial, sans-serif;
                            }
                            table {
                                font-family: Arial, sans-serif;
                                border-collapse: collapse;
                                width: 100%;
                                margin-top: 10px;
                            }
                            td, th {
                                border: 1px solid #dddddd;
                                text-align: left;
                                padding: 8px;
                            }
                            tr:nth-child(even) {
                                background-color: #f2f2f2;
                            }
                            .info {
                                font-size: 14px;
                                margin-bottom: 10px;
                            }
                            .title {
                                font-size: 18px;
                                font-weight: bold;
                                margin-bottom: 10px;
                            }
                        </style>
                    </head>
                    <body>
                        <div class="title">Excel Format Requirements</div>
                        <div class="info">
                            <ol>
                                <li>First cell must contain the 4-digit year (e.g., <b>2025</b>).</li>
                                <li>Subsequent cells should contain dates in any valid Excel date format.</li>
                                <li>Dates must belong to the specified year.</li>
                            </ol>
                        </div>

                        <h2>Holiday List</h2>
                        <table>
                            <tr>
                                <th>Year</th>
                            </tr>
                            <tr>
                                <td>DD-MM-YYYY</td>
                            </tr>
                            <tr>
                                <td>DD-MM-YYYY</td>
                            </tr>
                            <tr>
                                <td>etc ...</td>
                            </tr>
                        </table>
                    </body>
                </html>
            """

        msg_box = QMessageBox(self)
        msg_box.setWindowTitle("Format Guide")
        msg_box.setTextFormat(Qt.TextFormat.RichText)  # Enable HTML rendering
        msg_box.setText(guide_text)
        msg_box.exec()

    def load_holidays_to_db(self):
        """Open file dialog and import holidays from Excel or Numbers"""
        file_path, _ = QFileDialog.getOpenFileName(self, "Select Holiday File", "",
                                                   "Spreadsheet Files (*.xlsx *.xls *.numbers);;Excel Files (*.xlsx *.xls);;Numbers Files (*.numbers);;All Files (*)")

        if not file_path:
            return

        try:
            # Get file info before processing
            file_name = os.path.basename(file_path)
            file_size = os.path.getsize(file_path)
            file_extension = os.path.splitext(file_path)[1].lower()

            # Convert bytes to human-readable format
            def sizeof_fmt(num, suffix='B'):
                for unit in ['', 'K', 'M', 'G', 'T', 'P', 'E', 'Z']:
                    if abs(num) < 1024.0:
                        return f"{num:3.1f} {unit}{suffix}"
                    num /= 1024.0
                return f"{num:.1f} Y{suffix}"

            size_str = sizeof_fmt(file_size)
            display_text = f"{file_name} ({size_str})"
            self.holiday_input.setPlainText(display_text)

            # Read file based on extension
            df = None

            if file_extension in ['.xlsx', '.xls']:
                # Read Excel file
                df = pd.read_excel(file_path, header=None)

            elif file_extension == '.numbers':
                # Handle Numbers files
                try:
                    # Try using pandas with xlrd engine (if Numbers file is converted)
                    df = pd.read_excel(file_path, header=None, engine='openpyxl')
                except:
                    try:
                        # Alternative approach - try to read as CSV if Numbers exported as such
                        df = pd.read_csv(file_path, header=None)
                    except:
                        # If direct reading fails, show instructions for manual conversion
                        msg = ("Numbers files need to be exported as Excel format first.\n\n"
                               "Steps:\n"
                               "1. Open your Numbers file\n"
                               "2. Go to File → Export To → Excel...\n"
                               "3. Save as .xlsx format\n"
                               "4. Use the exported .xlsx file with this application")
                        self.show_message(f"File Format Note: {msg}", "info", 8000)

                        # Ask user if they want to try alternative method
                        reply = QMessageBox.question(self, "Numbers File Detected",
                                                     "Numbers files require conversion to Excel format.\n\n"
                                                     "Would you like to:\n"
                                                     "• YES: Get instructions for manual conversion\n"
                                                     "• NO: Cancel and select a different file",
                                                     QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)

                        if reply == QMessageBox.StandardButton.Yes:
                            QMessageBox.information(self, "Conversion Instructions", "To convert your Numbers file:\n\n"
                                                                                     "1. Open the Numbers file\n"
                                                                                     "2. Click File → Export To → Excel...\n"
                                                                                     "3. Choose 'Advanced Options' if needed\n"
                                                                                     "4. Select .xlsx format\n"
                                                                                     "5. Save the file\n"
                                                                                     "6. Return here and select the .xlsx file\n\n"
                                                                                     "Your data structure should be:\n"
                                                                                     "• First row: Year (e.g., 2025)\n"
                                                                                     "• Following rows: Holiday dates")
                        return
            else:
                raise ValueError(f"Unsupported file format: {file_extension}")

            if df is None or df.empty:
                raise ValueError("Failed to read the file or file is empty")

            # Clear previous holiday list
            self.HOLIDAY_LIST = []

            # Get year from first cell
            excel_year = str(df.iloc[0, 0])
            if not excel_year.isdigit() or len(excel_year) != 4:
                raise ValueError("First cell must contain a 4-digit year")

            # Process dates
            processed_dates = 0
            for idx in range(1, len(df)):
                date_val = df.iloc[idx, 0]

                if pd.isna(date_val):
                    continue

                try:
                    # Handle different date formats
                    if isinstance(date_val, str):
                        # Try multiple date formats
                        date_formats = ["%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d"]
                        date_obj = None

                        for fmt in date_formats:
                            try:
                                date_obj = datetime.strptime(date_val, fmt)
                                break
                            except ValueError:
                                continue

                        if date_obj is None:
                            raise ValueError(f"Unable to parse date format: {date_val}")

                    elif isinstance(date_val, pd.Timestamp):
                        # Convert pandas Timestamp to datetime
                        date_obj = date_val.to_pydatetime()
                    elif hasattr(date_val, 'date'):
                        # Handle datetime objects
                        date_obj = date_val
                    else:
                        # Try to convert to datetime
                        date_obj = pd.to_datetime(date_val).to_pydatetime()

                    # Validate year match
                    if str(date_obj.year) != excel_year:
                        msg = f"Invalid Date: Error in row {idx + 1}:\nDate {date_obj.date()} doesn't match file year {excel_year}"
                        self.show_message(msg, "error", 5000)
                        return  # Stop the operation immediately

                    formatted_date = date_obj.strftime("%d-%m-%Y")
                    if formatted_date not in self.HOLIDAY_LIST:  # Avoid duplicates
                        self.HOLIDAY_LIST.append(formatted_date)
                        processed_dates += 1

                except Exception as e:
                    msg = f"Invalid Date: Error in row {idx + 1}:\n{str(e)}"
                    self.show_message(msg, "error", 5000)
                    return

            if processed_dates == 0:
                msg = "No valid holiday dates found in the file"
                self.show_message(msg, "warning", 5000)
                return

            # Check for existing entry
            cursor = self.db_connection.cursor()
            cursor.execute("SELECT year FROM holiday WHERE year = ?", (excel_year,))
            exists = cursor.fetchone()

            if exists:
                confirm = QMessageBox.question(self, "Override Confirmation",
                                               f"Holidays for {excel_year} already exist. Override?",
                                               QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
                if confirm != QMessageBox.StandardButton.Yes:
                    return

            # Insert/Update data
            cursor.execute("INSERT OR REPLACE INTO holiday (year, holidays) VALUES (?, ?)",
                           (excel_year, json.dumps(self.HOLIDAY_LIST)))
            self.db_connection.commit()

            # Update UI
            current_year = QDate.currentDate().year()
            if self.year_combo.findText(excel_year) == -1:  # If year not found in combo
                self.year_combo.addItem(excel_year)
            self.year_combo.setCurrentText(excel_year)

            msg = f"Successfully loaded {processed_dates} holidays for {excel_year} from {file_extension.upper()} file"
            self.show_message(msg, "success", 5000)

        except Exception as e:
            msg = f"Import Error: Failed to load holidays:\n{str(e)}"
            self.show_message(msg, "error", 5000)
        finally:
            if 'cursor' in locals() and cursor:
                cursor.close()

    def show_holiday_viewer(self):
        """Show holiday viewer dialog with table display"""
        dialog = QDialog(self)
        dialog.setWindowTitle("Holiday List")
        dialog.setMinimumSize(500, 400)  # Increased width and height for better visibility

        layout = QVBoxLayout(dialog)

        # Year selection
        year_layout = QHBoxLayout()
        year_label = QLabel("Select Year:")

        self.viewer_year_combo = QComboBox()
        self.viewer_year_combo.setFixedWidth(100)  # Increased width for better readability
        self.viewer_year_combo.addItems([str(y) for y in range(2024, 2051)])
        # Set the current year as the default selection
        current_year = str(QDate.currentDate().year())
        self.viewer_year_combo.setCurrentText(current_year)

        year_layout.addWidget(year_label)
        year_layout.addWidget(self.viewer_year_combo)
        year_layout.addStretch()

        # Table setup with scroll area
        self.holiday_table = QTableWidget()
        self.holiday_table.setColumnCount(2)
        self.holiday_table.setHorizontalHeaderLabels(["Date", "Day"])
        self.holiday_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)

        # Make table cells uneditable
        self.holiday_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)

        # Wrap the table in a scroll area
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)

        table_container = QWidget()
        table_layout = QVBoxLayout(table_container)
        table_layout.addWidget(self.holiday_table)
        scroll_area.setWidget(table_container)

        # Refresh button
        refresh_btn = QPushButton("Refresh")
        refresh_btn.setFixedHeight(35)  # Adjust button height for a better look
        refresh_btn.clicked.connect(lambda: self.populate_holiday_table(self.viewer_year_combo.currentText()))

        layout.addLayout(year_layout)
        layout.addWidget(scroll_area)  # Add scrollable table
        layout.addWidget(refresh_btn)

        # Initial population
        self.viewer_year_combo.currentTextChanged.connect(self.populate_holiday_table)
        self.populate_holiday_table(self.viewer_year_combo.currentText())

        dialog.exec()

    def populate_holiday_table(self, year):
        """Populate table with holidays for selected year"""
        try:
            cursor = self.db_connection.cursor()
            cursor.execute("SELECT holidays FROM holiday WHERE year = ?", (year,))
            result = cursor.fetchone()

            self.holiday_table.setRowCount(0)

            if result:
                holidays = json.loads(result[0])
                self.holiday_table.setRowCount(len(holidays))

                for row, date_str in enumerate(holidays):
                    date = QDateTime.fromString(date_str, "dd-MM-yyyy")

                    self.holiday_table.setItem(row, 0, QTableWidgetItem(date_str))
                    self.holiday_table.setItem(row, 1, QTableWidgetItem(date.toString("dddd")))

        except sqlite3.Error as e:
            QMessageBox.critical(self, "Database Error", f"Failed to load holidays:\n{str(e)}")

    def create_config_page(self):
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(40, 40, 40, 40)

        title = QLabel("Configuration Settings")
        title.setObjectName("title")

        form = QFormLayout()
        form.setVerticalSpacing(15)
        form.setHorizontalSpacing(20)

        self.username_input = QLineEdit()
        self.theme_combo = QComboBox()
        self.theme_combo.addItems(["Dark", "Light", "System"])
        self.theme_combo.currentTextChanged.connect(self.change_theme)

        form.addRow(QLabel("Username:"), self.username_input)
        form.addRow(QLabel("Interface Theme:"), self.theme_combo)

        layout.addWidget(title)
        layout.addSpacing(20)
        layout.addLayout(form)
        layout.addStretch()
        return page

    def create_database_page(self):
        """Create the database management page with dropdown"""
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)

        # Database selection section
        db_control_layout = QHBoxLayout()

        # Table selection dropdown
        self.db_table_combo = QComboBox()
        self.db_table_combo.setPlaceholderText("Select a table...")
        self.db_table_combo.currentTextChanged.connect(self.show_table_contents)
        db_control_layout.addWidget(QLabel("Select Table:"), 1)
        db_control_layout.addWidget(self.db_table_combo, 4)

        # Export button
        self.export_btn = QPushButton("📤 Export Records")
        self.export_btn.setFixedHeight(40)
        self.export_btn.clicked.connect(self.export_record)
        self.export_btn.setEnabled(True)
        db_control_layout.addWidget(self.export_btn)

        # Delete Table button
        self.delete_table_btn = QPushButton("🗑️ Delete Table")
        self.delete_table_btn.setFixedHeight(40)
        self.delete_table_btn.clicked.connect(self.delete_current_table)
        self.delete_table_btn.setEnabled(True)
        db_control_layout.addWidget(self.delete_table_btn)

        layout.addLayout(db_control_layout)

        # Add this before table view initialization
        self.filter_row = QWidget()
        self.filter_layout = QHBoxLayout()
        self.filter_layout.setContentsMargins(0, 0, 0, 0)
        self.filter_layout.setSpacing(5)
        self.filter_row.setLayout(self.filter_layout)
        layout.addWidget(self.filter_row)

        # Table view section
        self.table_view = QTableWidget()
        self.table_view.setSortingEnabled(True)
        self.table_view.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.table_view.horizontalHeader().setStretchLastSection(True)
        self.table_view.verticalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        self.table_view.setAlternatingRowColors(True)
        layout.addWidget(self.table_view)

        return page

    def show_table_contents(self, table_name):
        self.current_table = table_name
        """Display contents of selected table from dropdown with enhanced UI"""
        if not table_name:
            return

        try:
            cursor = self.db_connection.cursor()
            cursor.execute(f"PRAGMA table_info({table_name})")
            columns = [col[1] for col in cursor.fetchall()]

            cursor.execute(f"SELECT * FROM {table_name}")
            rows = cursor.fetchall()
            self.table_view.clear()  #reset the view in every table selected
            # Configure table view with modern styling
            self.table_view.setRowCount(len(rows))
            self.table_view.setColumnCount(len(columns) + 1)
            self.table_view.setHorizontalHeaderLabels(columns + ["Actions"])

            # Make table cells uneditable
            self.table_view.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)

            # Enhanced table styling
            self.table_view.setStyleSheet("""
                QTableWidget {
                    background-color: #ffffff;
                    border: 1px solid #e0e0e0;
                    border-radius: 8px;
                    gridline-color: #f0f0f0;
                    selection-background-color: #e3f2fd;
                    font-size: 12px;
                    font-family: 'Segoe UI', Arial, sans-serif;
                }
                QTableWidget::item {
                    padding: 12px 8px;
                    border-bottom: 1px solid #f5f5f5;
                }
                QTableWidget::item:selected {
                    background-color: #e3f2fd;
                    color: #1976d2;
                }
                QTableWidget::item:hover {
                    background-color: #f8f9fa;
                }
                QHeaderView::section {
                    background-color: #f8f9fa;
                    color: #424242;
                    font-weight: bold;
                    font-size: 11px;
                    text-transform: uppercase;
                    letter-spacing: 0.5px;
                    padding: 12px 8px;
                    border: none;
                    border-bottom: 2px solid #e0e0e0;
                    border-right: 1px solid #e0e0e0;
                }
                QHeaderView::section:first {
                    border-top-left-radius: 8px;
                }
                QHeaderView::section:last {
                    border-top-right-radius: 8px;
                    border-right: none;
                }
                QScrollBar:vertical {
                    border: none;
                    background: #f5f5f5;
                    width: 12px;
                    border-radius: 6px;
                }
                QScrollBar::handle:vertical {
                    background: #c0c0c0;
                    border-radius: 6px;
                    min-height: 20px;
                }
                QScrollBar::handle:vertical:hover {
                    background: #a0a0a0;
                }
                QScrollBar:horizontal {
                    border: none;
                    background: #f5f5f5;
                    height: 12px;
                    border-radius: 6px;
                }
                QScrollBar::handle:horizontal {
                    background: #c0c0c0;
                    border-radius: 6px;
                    min-width: 20px;
                }
                QScrollBar::handle:horizontal:hover {
                    background: #a0a0a0;
                }
                QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal {
                    width: 0;
                }
            """)

            # Enable word wrapping for all items
            self.table_view.setWordWrap(True)

            # Set row height to auto-adjust based on content
            self.table_view.verticalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
            self.table_view.verticalHeader().setDefaultSectionSize(80)  # Minimum row height
            # self.table_view.verticalHeader().hide()  # Hide row numbers for cleaner look

            # Set alternating row colors
            self.table_view.setAlternatingRowColors(True)

            # Clear existing filters with animation-like effect
            while self.filter_layout.count():
                if child := self.filter_layout.takeAt(0):
                    if widget := child.widget():
                        widget.deleteLater()

            # Enhanced filter section with proper alignment
            filter_container = QWidget()
            filter_container.setStyleSheet("""
                QWidget {
                    background-color: #f8f9fa;
                    border-radius: 8px;
                    margin: 4px;
                    padding: 8px;
                }
            """)

            # Create a grid layout for filters to align with table columns
            filter_grid_layout = QGridLayout(filter_container)
            filter_grid_layout.setSpacing(4)
            filter_grid_layout.setContentsMargins(12, 8, 12, 8)

            # Add filter label
            filter_label = QLabel("🔍 Filters:")
            filter_label.setStyleSheet("""
                QLabel {
                    color: #424242;
                    font-weight: bold;
                    font-size: 12px;
                    background: none;
                    padding: 0;
                }
            """)
            filter_grid_layout.addWidget(filter_label, 0, 0, 1, len(columns) + 1)

            # Add enhanced filter inputs aligned with columns
            self.filter_inputs = []
            for col_idx, col_name in enumerate(columns):
                filter_edit = QLineEdit()
                filter_edit.setPlaceholderText(col_name)
                filter_edit.setStyleSheet("""
                    QLineEdit {
                        background-color: white;
                        border: 2px solid #e0e0e0;
                        border-radius: 6px;
                        padding: 6px 8px;
                        font-size: 10px;
                        min-height: 20px;
                    }
                    QLineEdit:focus {
                        border-color: #2196f3;
                        background-color: #fafafa;
                    }
                    QLineEdit:hover {
                        border-color: #bdbdbd;
                    }
                """)
                filter_edit.textChanged.connect(self.apply_filters)
                filter_grid_layout.addWidget(filter_edit, 1, col_idx)
                self.filter_inputs.append(filter_edit)

            # Add clear filters button in the actions column
            clear_filters_btn = QPushButton("✖️")
            clear_filters_btn.setFixedSize(38, 38)
            clear_filters_btn.setStyleSheet("""
                QPushButton {
                    background-color: #ff9800;
                    color: white;
                    border: none;
                    border-radius: 14px;
                    font-weight: bold;
                    font-size: 12px;
                }
                QPushButton:hover {
                    background-color: #f57c00;
                }
                QPushButton:pressed {
                    background-color: #ef6c00;
                }
            """)
            clear_filters_btn.setToolTip("Clear All Filters")
            clear_filters_btn.clicked.connect(self.clear_all_filters)
            filter_grid_layout.addWidget(clear_filters_btn, 1, len(columns))

            # Add the filter container to the main layout
            self.filter_layout.addWidget(filter_container)

            # Configure columns with better spacing and horizontal scrolling
            header = self.table_view.horizontalHeader()

            # Enable horizontal scrolling when needed
            self.table_view.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
            self.table_view.setHorizontalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)

            # Action column with proper width calculation
            if len(columns) <= 3:
                # Set all columns to have max width and word wrap
                for col in range(len(columns)):
                    # Set max width to 250 (adjust as needed) with word wrap
                    header.setSectionResizeMode(col, QHeaderView.ResizeMode.Interactive)
                    self.table_view.setColumnWidth(col, 370)  # Reduced from 250 to make room for actions
                    header.setMaximumSectionSize(400)  # Max width for content columns
                action_column_width = 80  # Width for two buttons + spacing
            else:
                # Set all columns to have max width and word wrap
                for col in range(len(columns)):
                    # Set max width to 250 (adjust as needed) with word wrap
                    header.setSectionResizeMode(col, QHeaderView.ResizeMode.Interactive)
                    self.table_view.setColumnWidth(col, 200)  # Reduced from 250 to make room for actions
                    header.setMaximumSectionSize(300)  # Max width for content columns
                action_column_width = 210  # Width for two buttons + spacing

            header.setSectionResizeMode(len(columns), QHeaderView.ResizeMode.Fixed)
            self.table_view.setColumnWidth(len(columns), action_column_width)
            # Set minimum width for the action column header
            header.setMinimumSectionSize(action_column_width)

            # Populate data with enhanced styling
            for row_idx, row in enumerate(rows):
                for col_idx, value in enumerate(row):
                    item = QTableWidgetItem(str(value) if value is not None else "")
                    # Set alignment: left horizontal, center vertical
                    item.setTextAlignment(Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft)

                    # Add subtle styling based on data type
                    if isinstance(value, (int, float)) and value != 0:
                        item.setForeground(QColor("#1976d2"))  # Blue for numbers
                    elif str(value).lower() in ['true', 'false', 'yes', 'no']:
                        item.setForeground(QColor("#4caf50" if str(value).lower() in ['true', 'yes'] else "#f44336"))

                    self.table_view.setItem(row_idx, col_idx, item)

                # Enhanced action buttons - properly sized
                action_widget = QWidget()
                action_widget.setStyleSheet("background-color: transparent;")
                action_layout = QHBoxLayout(action_widget)
                action_layout.setContentsMargins(2, 2, 2, 2)  # Small margins
                action_layout.setSpacing(6)  # Space between buttons

                # Better sized edit button
                edit_btn = QPushButton("🖋️Edit")
                edit_btn.setFixedSize(28, 28)  # Increased from 24x24
                edit_btn.setStyleSheet("""
                    QPushButton {
                        background-color: #FFB300; /* Amber */
                        color: white;
                        border: none;
                        border-radius: 14px;
                        font-weight: bold;
                        font-size: 12px;
                        padding: 6px 12px;
                    }
                    QPushButton:hover {
                        background-color: #FFA000; /* Darker Amber */
                    }
                    QPushButton:pressed {
                        background-color: #FF8F00; /* Even deeper Amber */
                    }
                """)
                edit_btn.setToolTip("Edit Record")
                edit_btn.setCursor(Qt.CursorShape.PointingHandCursor)
                edit_btn.clicked.connect(lambda _, r=row, tn=table_name: self.open_edit_dialog(tn, r))

                # Better sized delete button
                delete_btn = QPushButton("🗑️Delete")
                delete_btn.setFixedSize(28, 28)  # Increased from 24x24
                delete_btn.setStyleSheet("""
                    QPushButton {
                        background-color: #D32F2F;
                        color: white;
                        border: none;
                        border-radius: 14px;
                        font-weight: bold;
                        font-size: 12px;
                        padding: 6px 12px;
                    }
                    QPushButton:hover {
                        background-color: #E53935;
                    }
                    QPushButton:pressed {
                        background-color: #C62828;
                    }
                """)
                delete_btn.setToolTip("Delete Record")
                delete_btn.setCursor(Qt.CursorShape.PointingHandCursor)
                delete_btn.clicked.connect(lambda _, r=row, tn=table_name: self.delete_row(tn, r))

                action_layout.addWidget(edit_btn)
                action_layout.addWidget(delete_btn)

                # action_layout.addWidget(edit_btn)
                # action_layout.addWidget(delete_btn)
                # action_layout.addStretch()

                self.table_view.setCellWidget(row_idx, len(columns), action_widget)

            # Apply initial filters
            self.apply_filters()

            # Add a subtle drop shadow effect to the table
            shadow_effect = QGraphicsDropShadowEffect()
            shadow_effect.setBlurRadius(15)
            shadow_effect.setColor(QColor(0, 0, 0, 30))
            shadow_effect.setOffset(0, 2)
            self.table_view.setGraphicsEffect(shadow_effect)

        except sqlite3.Error as e:
            QMessageBox.critical(self, "Database Error", f"Failed to load table:\n{str(e)}")
        finally:
            if cursor:
                cursor.close()

    def clear_all_filters(self):
        """Clear all filter inputs"""
        if hasattr(self, 'filter_inputs'):
            for filter_input in self.filter_inputs:
                filter_input.clear()

    def apply_filters(self):
        """Apply case-insensitive partial matching filters to table rows"""
        try:
            # Get lowercase filter texts
            filters = [edit.text().strip().lower() for edit in self.filter_inputs]

            # Check each row
            for row in range(self.table_view.rowCount()):
                should_show = True
                for col in range(len(filters)):
                    if filters[col]:  # Only check non-empty filters
                        item = self.table_view.item(row, col)
                        cell_text = item.text().lower() if item else ""

                        # Partial match check
                        if filters[col] not in cell_text:
                            should_show = False
                            break

                # Show/hide row based on filter match
                self.table_view.setRowHidden(row, not should_show)

        except Exception as e:
            print(f"Filter error: {str(e)}")

    def export_record(self):
        ...

    def open_edit_dialog(self, table_name, row_data):
        """Open spacious edit dialog with large input fields"""
        dialog = QDialog(self)
        dialog.setWindowTitle(f"Edit Record - {table_name}")
        dialog.setMinimumSize(800, 600)  # Larger dialog size

        # Main layout with spacing
        main_layout = QVBoxLayout(dialog)
        main_layout.setContentsMargins(30, 30, 30, 30)
        main_layout.setSpacing(25)

        # Header section
        header = QLabel(f"Editing Record in '{table_name}'")
        header.setStyleSheet("""
            QLabel {
                font-size: 20px;
                font-weight: bold;
                color: #2c3e50;
                padding-bottom: 15px;
                border-bottom: 2px solid #3498db;
            }
        """)
        main_layout.addWidget(header)

        # Get column information
        cursor = self.db_connection.cursor()
        cursor.execute(f"PRAGMA table_info({table_name})")
        columns = [col[1] for col in cursor.fetchall()]

        # Scroll area for long forms
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        content = QWidget()
        form_layout = QVBoxLayout(content)  # Changed to vertical layout
        form_layout.setContentsMargins(15, 15, 15, 15)
        form_layout.setSpacing(30)

        input_fields = {}
        for col_name, value in zip(columns, row_data):
            # Create field container
            field_container = QWidget()
            field_layout = QVBoxLayout(field_container)
            field_layout.setContentsMargins(0, 0, 0, 0)
            field_layout.setSpacing(8)

            # Label with larger font
            label = QLabel(col_name)
            label.setStyleSheet("""
                QLabel {
                    font-size: 14px;
                    font-weight: bold;
                    color: #34495e;
                }
            """)

            # For longer text fields
            input_field = QTextEdit() if isinstance(value, str) and len(str(value)) > 50 else QLineEdit()

            # Set initial value
            if isinstance(input_field, QLineEdit):
                input_field.setText(str(value))
                input_field.setClearButtonEnabled(True)
            else:
                input_field.setPlainText(str(value))

            # Common styling for both input types
            input_field.setStyleSheet("""
                QLineEdit, QTextEdit {
                    font-size: 16px;
                    padding: 12px;
                    border: 2px solid #bdc3c7;
                    border-radius: 6px;
                    min-height: 50px;
                }
                QLineEdit:focus, QTextEdit:focus {
                    border-color: #3498db;
                }
            """)

            field_layout.addWidget(label)
            field_layout.addWidget(input_field)
            form_layout.addWidget(field_container)
            input_fields[col_name] = input_field

        scroll.setWidget(content)
        main_layout.addWidget(scroll)

        # Button container
        button_container = QWidget()
        button_layout = QHBoxLayout(button_container)
        button_layout.setContentsMargins(0, 0, 0, 0)

        # Save button with icon
        save_btn = QPushButton(" Save Changes")
        save_btn.setIcon(QIcon.fromTheme("document-save"))
        save_btn.setStyleSheet("""
            QPushButton {
                background-color: #27ae60;
                color: white;
                border: none;
                padding: 15px 30px;
                border-radius: 6px;
                font-size: 16px;
                min-width: 150px;
            }
            QPushButton:hover {
                background-color: #219a52;
            }
        """)
        save_btn.clicked.connect(lambda: self.save_edited_row(dialog, table_name, columns, row_data, input_fields))

        # Cancel button
        cancel_btn = QPushButton("Cancel")
        cancel_btn.setStyleSheet("""
            QPushButton {
                background-color: #e74c3c;
                color: white;
                border: none;
                padding: 15px 30px;
                border-radius: 6px;
                font-size: 16px;
                min-width: 150px;
            }
            QPushButton:hover {
                background-color: #c0392b;
            }
        """)
        cancel_btn.clicked.connect(dialog.reject)

        button_layout.addStretch()
        button_layout.addWidget(cancel_btn)
        button_layout.addWidget(save_btn)

        main_layout.addWidget(button_container)

        dialog.setStyleSheet("""
            QDialog {
                background-color: #f8f9fa;
                font-family: 'Segoe UI', sans-serif;
            }
        """)

        dialog.exec()

    def save_edited_row(self, dialog, table_name, columns, old_row_data, input_fields):
        """Universal save function that works for all widget types and handles NULL values properly"""
        cursor = None
        try:
            cursor = self.db_connection.cursor()

            # Helper function to get text from any widget type
            def get_widget_value(widget):
                """Extract text from any Qt widget type"""
                if hasattr(widget, 'text'):  # QLineEdit, QLabel
                    return widget.text()
                elif hasattr(widget, 'toPlainText'):  # QTextEdit, QPlainTextEdit
                    return widget.toPlainText()
                elif hasattr(widget, 'currentText'):  # QComboBox
                    return widget.currentText()
                elif hasattr(widget, 'value'):  # QSpinBox, QDoubleSpinBox
                    return str(widget.value())
                elif hasattr(widget, 'isChecked'):  # QCheckBox
                    return str(widget.isChecked())
                elif hasattr(widget, 'date'):  # QDateEdit
                    return widget.date().toString()
                elif hasattr(widget, 'time'):  # QTimeEdit
                    return widget.time().toString()
                elif hasattr(widget, 'dateTime'):  # QDateTimeEdit
                    return widget.dateTime().toString()
                else:
                    return str(widget)

            # Helper function to build WHERE clause with proper NULL handling
            def build_where_clause(columns, values):
                """Build WHERE clause that properly handles NULL values"""
                conditions = []
                params = []

                for col, val in zip(columns, values):
                    if val is None:
                        conditions.append(f'"{col}" IS NULL')
                    else:
                        conditions.append(f'"{col}" = ?')
                        params.append(val)

                return " AND ".join(conditions), params

            # Helper function to normalize values for comparison
            def normalize_value(value):
                """Normalize values for consistent comparison"""
                if value is None or value == "":
                    return None
                return str(value).strip()

            # Get new values from input fields
            new_values = []
            for col in columns:
                widget = input_fields[col]
                raw_value = get_widget_value(widget)
                normalized_value = normalize_value(raw_value)
                new_values.append(normalized_value)

            # Normalize old values for comparison
            old_values = [normalize_value(val) for val in old_row_data]

            # Check if any values have actually changed
            if new_values == old_values:
                QMessageBox.information(self, "No Changes", "No changes were made to the record.")
                dialog.close()
                return

            # Build WHERE clause that handles NULL values properly
            where_clause, where_params = build_where_clause(columns, old_row_data)

            # Get the ROWID of the current row
            rowid_query = f'SELECT ROWID FROM "{table_name}" WHERE {where_clause} LIMIT 1'
            cursor.execute(rowid_query, where_params)
            rowid_result = cursor.fetchone()

            if not rowid_result:
                # If ROWID method fails, try to find the record by checking if it exists
                check_query = f'SELECT COUNT(*) FROM "{table_name}" WHERE {where_clause}'
                cursor.execute(check_query, where_params)
                count = cursor.fetchone()[0]

                if count == 0:
                    QMessageBox.warning(self, "Error",
                                        "Original record not found. It may have been deleted or modified by another process.")
                elif count > 1:
                    QMessageBox.warning(self, "Error", "Multiple matching records found. Cannot safely update the record.")
                else:
                    QMessageBox.warning(self, "Error", "Unable to locate the record for updating.")
                return

            rowid = rowid_result[0]

            # Prepare update query
            set_clause = ", ".join([f'"{col}" = ?' for col in columns])
            update_query = f'UPDATE "{table_name}" SET {set_clause} WHERE ROWID = ?'

            # Execute update
            cursor.execute(update_query, new_values + [rowid])

            if cursor.rowcount > 0:
                self.db_connection.commit()
                QMessageBox.information(self, "Success", "Record updated successfully!")
                dialog.close()

                # Refresh the table view
                if hasattr(self, 'show_table_contents'):
                    self.show_table_contents(table_name)
            else:
                QMessageBox.warning(self, "Warning",
                                    "No record was updated. The record may have been modified by another process.")

        except sqlite3.Error as e:
            # Handle database errors
            try:
                if hasattr(self, 'db_connection') and self.db_connection:
                    self.db_connection.rollback()
            except:
                pass

            error_msg = f"Database error occurred while updating the record:\n\n{str(e)}"
            QMessageBox.critical(self, "Database Error", error_msg)

        except Exception as e:
            # Handle any other unexpected errors
            try:
                if hasattr(self, 'db_connection') and self.db_connection:
                    self.db_connection.rollback()
            except:
                pass

            error_msg = f"An unexpected error occurred:\n\n{str(e)}"
            QMessageBox.critical(self, "Error", error_msg)

        finally:
            # Always close cursor
            if cursor:
                try:
                    self.show_table_contents(table_name)
                    cursor.close()
                except:
                    pass

    # Test the query manually to debug:
    def debug_rowid_query(self, table_name, columns, old_values):
        """Debug function to test the ROWID query"""
        cursor = self.db_connection.cursor()

        # Build WHERE clause that handles NULL values properly
        where_clause, where_params = self.build_where_clause_with_nulls(columns, old_values)

        rowid_query = f'SELECT ROWID, * FROM "{table_name}" WHERE {where_clause}'
        print(f"Query: {rowid_query}")
        print(f"Parameters: {where_params}")

        cursor.execute(rowid_query, where_params)
        results = cursor.fetchall()

        print(f"Found {len(results)} matching records:")
        for result in results:
            print(f"ROWID: {result[0]}, Data: {result[1:]}")

        cursor.close()
        return results

    def delete_row(self, table_name, row_data):
        """Delete a row from the database after confirmation"""
        confirm = QMessageBox.question(self, "Confirm Delete", "Are you sure you want to delete this record?",
                                       QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)

        if confirm == QMessageBox.StandardButton.Yes:
            try:
                cursor = self.db_connection.cursor()

                # Get column names for the table
                cursor.execute(f"PRAGMA table_info({table_name})")
                columns_info = cursor.fetchall()
                column_names = [col[1] for col in columns_info]  # col[1] contains column name

                # Ensure we have the same number of columns as data values
                if len(column_names) != len(row_data):
                    QMessageBox.critical(self, "Error", "Column count mismatch with row data.")
                    return

                # Prepare delete query using column names
                where_conditions = []
                where_values = []

                for col_name, value in zip(column_names, row_data):
                    if value is not None and value != '':  # Skip empty values
                        where_conditions.append(f"{col_name} = ?")
                        where_values.append(value)
                    else:
                        where_conditions.append(f"{col_name} IS NULL")

                where_clause = " AND ".join(where_conditions)
                query = f"DELETE FROM {table_name} WHERE {where_clause}"

                cursor.execute(query, where_values)
                self.db_connection.commit()  # Uncommented this - you need to commit the transaction

                if cursor.rowcount > 0:
                    QMessageBox.information(self, "Deleted", "Record deleted successfully.")
                    # Refresh table view
                    self.show_table_contents(table_name)
                else:
                    QMessageBox.warning(self, "Warning", "No matching record found to delete.")

            except sqlite3.Error as e:
                QMessageBox.critical(self, "Error", f"Failed to delete record:\n{str(e)}")
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Unexpected error:\n{str(e)}")

    def delete_current_table(self):
        """Delete currently selected table with user confirmation"""
        if not self.current_table:
            QMessageBox.warning(self, "No Table Selected", "Please select a table to delete.")
            return

        confirm = QMessageBox.question(self, "Confirm Delete",
                                       f"Are you sure you want to permanently delete the table '{self.current_table}'?",
                                       QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)

        if confirm == QMessageBox.StandardButton.Yes:
            try:
                cursor = self.db_connection.cursor()
                cursor.execute(f"DROP TABLE IF EXISTS '{self.current_table}'")
                self.db_connection.commit()

                QMessageBox.information(self, "Table Deleted", f"Table '{self.current_table}' was deleted successfully.")

                # Refresh UI
                self.initialize_database()  # Refresh dropdown/list
                self.table_view.clear()  # Clear the current view
                self.current_table = None  # Reset current table
                self.delete_table_btn.setEnabled(True)
                self.export_btn.setEnabled(True)
                self.initialize_database()

                self.show_table_contents(self.all_tables_name[0])
                self.db_table_combo.setPlaceholderText(self.all_tables_name[0])
            except sqlite3.Error as e:
                QMessageBox.critical(self, "Database Error", f"Failed to delete table:\n{str(e)}")

            finally:
                if cursor:
                    cursor.close()

    def closeEvent(self, event):
        """Close database connection when window closes"""
        if self.db_connection:
            self.db_connection.close()
        event.accept()

    def create_about_page(self):
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setAlignment(Qt.AlignmentFlag.AlignCenter)

        title = QLabel("About This Application")
        title.setObjectName("title")

        content = QLabel("Proprietary Software - Hitachi Digital Service\n\n"
                         "Version 1.3.0\n"
                         "Build Date: 2025-08-08\n\n"
                         "Developed using:\n"
                         "• Python 3.11\n"
                         "• PyQt6 Framework\n"
                         "• SQLite Database\n"
                         "** Developed by : Vimit **")
        content.setAlignment(Qt.AlignmentFlag.AlignCenter)

        layout.addWidget(title)
        layout.addSpacing(20)
        layout.addWidget(content)
        layout.addStretch()
        return page

    # ======================
    # CORE FUNCTIONALITY
    # ======================
    def switch_page(self, index):
        self.stacked_widget.setCurrentIndex(index)
        buttons = self.sidebar.findChildren(QPushButton)
        for i, btn in enumerate(buttons):
            btn.setChecked(i == index)

    def cycle_theme(self):
        current_index = self.themes.index(self.current_theme)
        new_index = (current_index + 1) % len(self.themes)
        self.current_theme = self.themes[new_index]
        self.update_theme_button()
        self.change_theme(self.current_theme)

    def update_theme_button(self):
        theme_data = {"Dark": {"icon": "🌙", "text": "Dark"}, "Light": {"icon": "☀️", "text": "Light"},
                      "System": {"icon": "⚙️", "text": "System"}}

        current_data = theme_data[self.current_theme]
        btn_text = f"{current_data['icon']} {current_data['text']}"

        # Calculate required width based on text content
        fm = self.theme_btn.fontMetrics()
        text_width = fm.horizontalAdvance(btn_text) + 20  # Add padding

        # Set dynamic width constraints
        self.theme_btn.setMinimumWidth(text_width)
        self.theme_btn.setMaximumWidth(text_width + 20)

        self.theme_btn.setText(btn_text)

        # Update button colors based on theme
        if self.current_theme == "Dark":
            self.theme_btn.setStyleSheet("""
                QPushButton {
                    background-color: #353535;
                    color: white;
                    border-color: #454545;
                }
            """)
        elif self.current_theme == "Light":
            self.theme_btn.setStyleSheet("""
                QPushButton {
                    background-color: #f0f0f0;
                    color: #333333;
                    border-color: #cccccc;
                }
            """)
        else:  # System
            self.theme_btn.setStyleSheet("""
                QPushButton {
                    background-color: #e0e0e0;
                    color: #000000;
                    border-color: #a0a0a0;
                }
            """)

    def change_theme(self, theme_name):
        app = QApplication.instance()
        if theme_name == "Dark":
            app.setStyleSheet(DARK_THEME)
        elif theme_name == "Light":
            app.setStyleSheet(LIGHT_THEME)
        else:
            app.setStyleSheet("")
        self.current_theme = theme_name


if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setFont(QFont("Segoe UI", 10))
    app.setStyleSheet(LIGHT_THEME)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())
