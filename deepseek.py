import sys
from collections import Counter
from datetime import datetime
from PyQt6.QtWidgets import (QApplication, QMainWindow, QWidget, QStackedWidget, QTableWidget,
    QPushButton, QLabel, QVBoxLayout, QHBoxLayout, QTableWidgetItem, QFileDialog,
    QFrame, QLineEdit, QComboBox, QFormLayout, QListWidget, QHeaderView, QDialog, QProgressBar,
    QMessageBox, QSizePolicy, QHBoxLayout, QSpacerItem, QToolBar, QGroupBox, QPlainTextEdit, QScrollArea)
from PyQt6.QtGui import QFont, QIcon, QAction, QActionGroup
from PyQt6 import QtCore, QtGui, QtWidgets
from PyQt6.QtCore import Qt, QDate, QDateTime, QTimer
import sqlite3
import json
from openpyxl import load_workbook
import pandas as pd
from pandas._libs.tslibs.timestamps import Timestamp
from openpyxl.utils import quote_sheetname
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from dateutil import parser
from datetime import datetime
import calendar
import itertools
import os
import re
from dateutil.parser import parse
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.merge import MergeCell, MergeCells
import numpy as np


# ======================
# THEME DEFINITIONS
# ======================
DARK_THEME = """
QWidget {
    background-color: #2b2b2b;
    color: #ffffff;
    selection-background-color: #3a3a3a;
}

QPushButton {
    background-color: #353535;
    border: 1px solid #454545;
    border-radius: 4px;
    padding: 8px 12px;
    min-width: 80px;
    font-size: 10px;
}

QPushButton:hover {
    background-color: #454545;
}

QPushButton:checked {
    background-color: #007acc;
    border-color: #006ab3;
}

QLineEdit, QComboBox {
    background-color: #353535;
    border: 1px solid #454545;
    border-radius: 4px;
    padding: 6px 8px;
    min-height: 28px;
}

QFrame#sidebar {
    background-color: #252526;
    border-right: 1px solid #353535;
}

QLabel#title {
    font-size: 20px;
    font-weight: 600;
    color: #007acc;
    padding-bottom: 15px;
}

QLabel#subtitle {
    font-size: 14px;
    color: #888;
}

QFormLayout QLabel {
    font-weight: 500;
    padding-bottom: 3px;
}
/* Holiday Group Box Buttons */
#format_info_btn {
    background-color: #6c5ce7;  /* Soft purple */
    color: white;
    border: 1px solid #5d4ec9;
}

#load_holiday_btn {
    background-color: #00b894;  /* Teal */
    color: white;
    border: 1px solid #00a383;
}

#view_holiday_btn {
    background-color: #e66767;  /* Soft coral */
    color: white;
    border: 1px solid #d35454;
}

QPushButton:hover {
    background-color: rgba(108, 92, 231, 0.8);  /* Purple hover */
}

#load_holiday_btn:hover {
    background-color: rgba(0, 184, 148, 0.8);  /* Teal hover */
}

#view_holiday_btn:hover {
    background-color: rgba(230, 103, 103, 0.8);  /* Coral hover */
}
"""

LIGHT_THEME = """
QWidget {
    background-color: #ffffff;
    color: #333333;
    selection-background-color: #e0e0e0;
}

QPushButton {
    background-color: #f5f5f5;
    border: 1px solid #cccccc;
    border-radius: 4px;
    padding: 8px 12px;
    min-width: 80px;
    font-size: 10px;
}

QPushButton:hover {
    background-color: #e8e8e8;
}

QPushButton:checked {
    background-color: #007acc;
    color: white;
    border-color: #006ab3;
}

QLineEdit, QComboBox {
    background-color: #ffffff;
    border: 1px solid #cccccc;
    border-radius: 4px;
    padding: 6px 8px;
    min-height: 28px;
}

QFrame#sidebar {
    background-color: #f8f8f8;
    border-right: 1px solid #e0e0e0;
}

QLabel#title {
    font-size: 20px;
    font-weight: 600;
    color: #007acc;
    padding-bottom: 15px;
}

QLabel#subtitle {
    font-size: 14px;
    color: #666;
}

QFormLayout QLabel {
    font-weight: 500;
    padding-bottom: 3px;
}
/* Holiday Group Box Buttons */
#format_info_btn {
    background-color: #a8a4e6;  /* Light purple */
    color: #2d3436;
    border: 1px solid #8f8bd9;
}

#load_holiday_btn {
    background-color: #55efc4;  /* Light teal */
    color: #2d3436;
    border: 1px solid #48cfad;
}

#view_holiday_btn {
    background-color: #ff7675;  /* Soft red */
    color: #2d3436;
    border: 1px solid #ff6564;
}

QPushButton:hover {
    background-color: rgba(168, 164, 230, 0.8);  /* Light purple hover */
}

#load_holiday_btn:hover {
    background-color: rgba(85, 239, 196, 0.8);  /* Light teal hover */
}

#view_holiday_btn:hover {
    background-color: rgba(255, 118, 117, 0.8);  /* Soft red hover */
}
"""


# =======================
#  GENERIC FUNCTION
#========================

TOTAL_WORKING_DAY = 0


def clean_string(s):
    # Remove text inside brackets (e.g., "[C]", "[c]", "[text]"), commas, and hyphens
    cleaned = re.sub(r"\[.*?\]|[, -]", " ", s)
    return cleaned.lower().strip()

def coverage_percentage(str1, str2):
    words1 = set(clean_string(str1).split())
    words2 = set(clean_string(str2).split())
    common_words = words1 & words2
    coverage = (len(common_words) / max(len(words1), len(words2))) * 100 if max(len(words1), len(words2)) > 0 else 0
    return coverage

def read_file(file_path):
    try:
        # Check file extension to determine how to read the file
        if file_path.endswith(".csv"):
            df = pd.read_csv(file_path)
        elif file_path.endswith(".xlsx"):
            df = pd.read_excel(file_path)
        else:
            raise ValueError("Unsupported file format. Please provide a CSV or Excel file.")

        # Return the DataFrame if reading is successful
        return df.to_dict("records")

    except FileNotFoundError:
        print(f"Error: File not found at '{file_path}'")
    except Exception as e:
        print(f"Error reading file '{file_path}': {e}")

    # Return None if there is an error
    return None

def sort_list_of_dicts(data):
    # Separate the dict with 'Name' equal to 'Total'
    total_dict = [d for d in data if d.get('Name') == 'Total']

    # Sort the remaining dicts by 'Name'
    sorted_data = sorted([d for d in data if d.get('Name') != 'Total'], key=lambda x: x['Name'])

    # Append the 'Total' dict at the end
    if total_dict:
        sorted_data.extend(total_dict)

    return sorted_data

def get_month_details(month_name, year):
    # Get the month number from the month name
    month_number = list(calendar.month_name).index(month_name.capitalize())

    # Get the calendar for the specified month and year
    cal = calendar.monthcalendar(year, month_number)

    # Dictionary to map weekday number to weekday name
    weekdays = {0: "Monday", 1: "Tuesday", 2: "Wednesday", 3: "Thursday", 4: "Friday", 5: "Saturday", 6: "Sunday", }

    month_details = []

    # Iterate through each week in the month
    for week in cal:
        week_details = []
        # Iterate through each day in the week
        for day in week:
            # If the day is zero, it means it's part of the previous or next month
            if day == 0:
                week_details.append(None)
            else:
                # Get the day name using the weekdays dictionary
                day_name = weekdays[calendar.weekday(year, month_number, day)]
                # Check if it's a weekend (Saturday or Sunday)
                is_weekend = day_name in ["Saturday", "Sunday"]
                # Append the day details to the week_details list
                week_details.append({"day": day, "day_name": day_name, "is_weekend": is_weekend})
        # Append the week_details list to the month_details list
        month_details.append(week_details)

    return month_details, month_number

def date_calculation(date):
    # Check if the date is already a datetime object
    if isinstance(date, datetime):
        date_obj = date
    else:
        # Parse the date string or timestamp
        date_obj = parser.parse(str(date))

    # Extract the day, month, and year
    day = date_obj.day
    month = date_obj.month
    year = date_obj.year

    return day, month, year

# add validation of month, like if use data is for april and month is may
# :: TODO calculation for individual user - billable days - fixed - done
# :: TODO exclude date cal for holiday days if user has marked the attendance - done
def generate_excel(month, year, output_file_name, selected_row, holiday_list, name_mapping, name_order_list, progress_bar):
    global TOTAL_WORKING_DAY
    sheets_name = []
    try:
        user_data = list()
        non_complaince_user = []
        month_name = month
        holiday_list = holiday_list

        year = int(year)
        month_details, month_number = get_month_details(month_name, year)
        month_number = f"{month_number:02}" if month_number < 10 else month_number
        month_holiday_list = [x for x in holiday_list if re.findall(f"\d+-{month_number}-\d+", x)]
        # print("month_holiday_list ===>", month_holiday_list)
        month_day_holiday_list = [k.split("-")[0] for k in month_holiday_list]
        # print("month_day_holiday_list ===>", month_day_holiday_list)
        df_sheets = dict()
        excel_file_path = output_file_name

        # Creating a mapping from Rsname to their positions in the custom order
        order_map = {preprocess_name(name): index for index, name in enumerate(name_order_list)}

        # Sorting the list of dictionaries by the custom order
        selected_row = sorted(selected_row, key=lambda x: order_map.get(preprocess_name(x["Rsname"]), float('inf')))

        attendance_len = len(selected_row)
        progress_step = 0
        step = 100 / attendance_len
        for new_data in selected_row:
            start_date, end_date, sd, sm, sy, ed, em, ey, sm_flag, em_flag = (None,) * 10
            billable_days = 0
            weekends = 0
            total_working_days = 0
            data_model = []
            leave_taken = 0
            public_holiday = 0
            mismatch_date = []
            name = preprocess_name(new_data.get("Rsname"))
            start_date, end_date = name_mapping[name][2:] if name_mapping.get(name) else (None,) * 2

            if start_date:
                sd, sm, sy = date_calculation(start_date)

                sm = f"{sm:02}" if sm < 10 else sm

                sm_flag = sm == month_number

            if end_date:
                ed, em, ey = date_calculation(end_date)
                em = f"{em:02}" if em < 10 else em
                em_flag = em == month_number

            for week in month_details:
                for day in week:
                    if day:
                        date = day["day"]
                        is_weekend_or_leave = "Weekend" if day["is_weekend"] else ""
                        if day["is_weekend"]:
                            weekends += 1
                        else:
                            total_working_days += 1
                        dt = f"{date}-{month_name[:3]}"
                        day_name = f"{day['day_name']}"

                        key = (f"{day_name[:3]}, {f'{date:02}' if date < 10 else date}-"
                               f"{month_name[:3].title()}")
                        # print("DAY ===>", f"{date:02}" if date < 10 else date)
                        # holiday_list = [j.split("-")[0] for j in holiday_list ]
                        calculated_date = f"{date:02}" if date < 10 else f"{date}"
                        if day["is_weekend"]:
                            dt_status = 0
                        else:
                            rounded_values = list(np.around(np.arange(2.5, 4, 0.1), decimals=1))

                            match new_data.get(key):
                                case 8:
                                    dt_status = 1
                                    # As some user can mark attendace on holiday date
                                    if calculated_date in month_day_holiday_list:
                                        mismatch_date.append(calculated_date)
                                        public_holiday += 1
                                        dt_status = 0

                                        # for Holiday keyword to be added
                                        is_weekend_or_leave = "Holiday"
                                case 4:
                                    dt_status = 0.5
                                    leave_taken += 0.5
                                case value if value in rounded_values:
                                    dt_status = 0.25
                                    leave_taken += 0.25
                                case 0:
                                    if calculated_date in month_day_holiday_list:
                                        public_holiday += 1
                                        dt_status = 0

                                        # for Holiday keyword to be added
                                        is_weekend_or_leave = "Holiday"
                                    else:
                                        leave_taken += 1
                                        dt_status = 0

                                        # for leave keyword to be added
                                        is_weekend_or_leave = "Leave"
                                case _:
                                    leave_taken += 1
                                    dt_status = 0
                        billable_days += dt_status

                        """
                            (dt, day_name[:3], dt_status, is_weekend_or_leave, "", "", "", "")
                            ('1-Jun', 'Thu', 1, '', '', '', '', '')
                        """
                        if sm_flag and not em_flag:
                            if date < sd:
                                dt_status = 0
                                total_working_days -= 1

                        if em_flag and not sm_flag:
                            if date > ed:
                                dt_status = 0
                                total_working_days -= 1

                        if em_flag and sm_flag:
                            if date > ed or date < sd:
                                dt_status = 0
                                total_working_days -= 1

                        data_model.append((dt, day_name[:3], dt_status, is_weekend_or_leave, "", "", "", "",))

            billable_days = total_working_days - leave_taken
            # print("USER =====>", new_data.get("Rsname"))
            # print("BILLABLE ====>", billable_days)
            # print("WEEKENDS ====>", weekends)
            # print("TOTAL WORKING DAYS ====>", total_working_days)
            # print("LEAVE TAKEN ====>", leave_taken)
            # print("-"*20)
            # print("PUBLIC HOLIDAY ====>", public_holiday)
            point_of_contact = (name_mapping[name][1] if name_mapping.get(name) else "xxxxxxx")
            ID_521 = name_mapping[name][0] if name_mapping.get(name) else "xxxxxxx"
            if mismatch_date:
                non_complaince_user.append(
                    {"Name": new_data.get("Rsname"), "Month": month, "Listed Month Holiday": month_day_holiday_list,
                        "Attendance Marked on Holiday": mismatch_date, })
            data = {"Vendor Organization": ["Resource Name", "Month", "Date"],
                "Hitachi Vantara": [f"{new_data.get('Rsname')}", f"{month_name}", "Day", ],
                "Point of Contact": ["5-2-1", "Working Days", "Working Status"],
                f"{point_of_contact}": [f"{ID_521}", f"{total_working_days}", "Remarks", ],
                "Adjustments from Last Month": ["", "", ""], "0": ["", "", ""], "": ["", "", ""],
                "Week Off": ["Personal/Sick Leave", "", ""], }
            df = pd.DataFrame(data)

            # Create a new sheet or get the existing one
            sheet_name = new_data.get("Rsname")

            for i in data_model:
                df.loc[len(df)] = i

            df.loc[len(df)] = ["Leaves Taken", f"{leave_taken}", "Billable Days", f"{billable_days}", "", "", "", "", ]

            df.loc[len(df)] = ["Weekends", f"{weekends}", "", "", "", "", "", ""]
            df.loc[len(df)] = ["Public Holidays", f"{public_holiday}", "", "", "", "", "", "", ]

            # print(df)
            df_sheets.update({sheet_name: df})
            user_data.append(
                {"Name": sheet_name, #   "Total Billable Time": (total_working_days-leave_taken-public_holiday) * 8 ,
                    "Billable Time (Hours)": (total_working_days - leave_taken) * 8,
                    #    "Weekends": weekends, "Public Holidays": public_holiday,
                    "Total Number of Billable Days": total_working_days - leave_taken,
                    "Service Credit Pool Days": leave_taken, })
            progress_step += int(step)
            progress_bar.setValue(progress_step)
        else:
            TOTAL_WORKING_DAY = total_working_days

            # Create a Pandas Excel writer using XlsxWriter as the engine
            with pd.ExcelWriter(excel_file_path, engine="xlsxwriter") as writer:

                for key, value in df_sheets.items():
                    # Write each dataframe to a different sheet
                    value.to_excel(writer, sheet_name=key, index=False)

                # Access the XlsxWriter workbook and worksheet objects
                workbook = writer.book
                worksheets = writer.sheets

                # Access each worksheet and modify the formatting if needed
                for sheet_name, worksheet in worksheets.items():
                    # Example: set column width of the first column to 20
                    worksheet.set_column(0, 0, 20)
                    worksheet.set_column(1, 1, 20)
                    worksheet.set_column(2, 2, 20)
                    worksheet.set_column(3, 3, 15)
                    worksheet.set_column(4, 4, 30)
                    worksheet.set_column(7, 7, 17)

                    sheets_name.append(sheet_name)

            wb_style = load_workbook(excel_file_path)
            border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"),
                bottom=Side(border_style="thin"), )
            for i in sheets_name:
                sheet = wb_style[i]

                for row in sheet.iter_rows():
                    for cell in row:
                        cell.border = border
                        if cell.value == "Weekend":
                            numbers = re.findall(r"\d+", cell.coordinate)[0]
                            cell_list = [f"B{numbers}", f"C{numbers}", f"D{numbers}", ]
                            for k in cell_list:
                                cell_bold = sheet[k]
                                cell_bold.fill = PatternFill(start_color="b6bbbf", end_color="b6bbbf",
                                    fill_type="solid", )  # grey color
                        if cell.value == "Leave":
                            cell_bold = sheet[cell.coordinate]
                            cell_bold.fill = PatternFill(start_color="fce1dc", end_color="fce1dc",
                                fill_type="solid", )  # red color

                        if cell.value == "Holiday":
                            cell_bold = sheet[cell.coordinate]
                            cell_bold.fill = PatternFill(start_color="cffccf", end_color="cffccf",
                                fill_type="solid", )  # green color

                        if cell.value in ["Leaves Taken", "Weekends", "Public Holidays", "Billable Days", ]:
                            cell_bold = sheet[cell.coordinate]
                            cell_bold.font = Font(bold=True, color="FFFFFF")
                            cell_bold.fill = PatternFill(start_color="4d6c82", end_color="4d6c82", fill_type="solid", )

                for j in ["B1", "D1", "H1", "F1"]:
                    cell_bold = sheet[j]
                    cell_bold.font = Font(bold=False)

                for j in ["A1", "C1", "E1", "A2", "C2", "A3", "C3", "A4", "B4", "C4", "D4", "G1", "G2", ]:
                    cell_bold = sheet[j]
                    cell_bold.font = Font(bold=True)
                    if j == "G2":
                        cell_bold.fill = PatternFill(start_color="fce1dc", end_color="fce1dc",
                            fill_type="solid", )  # red color
                    else:
                        cell_bold.fill = PatternFill(start_color="b6bbbf", end_color="b6bbbf",
                            fill_type="solid", )  # grey color

                for j in ["A4", "B4", "C4", "D4", "E4", "F4"]:
                    cell_bold = sheet[j]
                    cell_bold.font = Font(bold=True, color="FFFFFF")
                    cell_bold.fill = PatternFill(start_color="4d6c82", end_color="4d6c82", fill_type="solid")

                for k in [chr(i) + f"{j}" for i in range(65, 73) for j in range(1, 5)]:
                    cell = sheet[k]
                    cell.alignment = Alignment(horizontal="left")

                for k in [f"A{j}" for j in range(5, 40)]:
                    cell = sheet[k]

                    cell.alignment = Alignment(horizontal="center")

                for k in [f"C{j}" for j in range(5, 40)]:
                    cell = sheet[k]

                    if cell.value == "Billable Days":
                        global range_limit
                        range_limit = int(re.findall(r"\d+", cell.coordinate)[0])
                        break

                    cell.alignment = Alignment(horizontal="center")

                for k in [i + f"{j}" for i in ["B", "D"] for j in range(5, 40)]:
                    cell = sheet[k]

                    if cell.coordinate == f"D{range_limit}":
                        cell.font = Font(bold=True)

                    cell.alignment = Alignment(horizontal="center")

            else:
                wb_style.save(excel_file_path)

        return [200, "Report Generated Successfully.", user_data, non_complaince_user]

    except Exception as e:
        # Log the error
        print(f"An error occurred: {str(e)}")
        exc_type, exc_obj, exc_tb = sys.exc_info()
        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        print(exc_type, fname, exc_tb.tb_lineno)

        return [500, str(e), None, None]

def format_date(date_str):
    if isinstance(date_str, Timestamp):
        date_str = date_str.strftime("%Y-%m-%d")  # Convert Timestamp to string
    try:
        # Attempt to parse the date string with various formats
        date_obj = parse(date_str)
        formatted_date = date_obj.strftime("%d-%m-%Y")
        return formatted_date
    except (ValueError, TypeError):
        # If parsing fails or the input is not a string, return the original value
        return str(date_str)

# Preprocess the name and item['Name'] strings
def preprocess_name(input_str):
    # Split the string into words, remove spaces, convert to lowercase, and sort
    return "".join(sorted(input_str.replace(",", "").lower().split()))


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Billing Report Generator")
        self.setGeometry(100, 100, 900, 600)
        self.setFixedSize(900, 600)
        self.current_theme = "Light"
        self.themes = ["Dark", "Light", "System"]
        self.db_connection = None
        self.current_table = None
        self.current_year = str(datetime.now().year)  # Add this line
        print(f"Current year : {self.current_year}")

        self.init_ui()
        self.initialize_database()

    def initialize_database(self):
        """Initialize database connection and required tables"""
        try:
            self.db_connection = sqlite3.connect('billing.db')
            cursor = self.db_connection.cursor()

            # Check existing tables
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
            existing_tables = [table[0].lower() for table in cursor.fetchall()]

            # Table creation queries with corrected syntax
            tables = {'holiday': '''
                    CREATE TABLE IF NOT EXISTS holiday (
                        year TEXT PRIMARY KEY,
                        holidays TEXT
                    )
                ''', 'user': '''
                    CREATE TABLE IF NOT EXISTS user (
                        name TEXT,
                        month TEXT,
                        year TEXT,
                        attendance_report TEXT,
                        PRIMARY KEY (name, month, year)
                    )
                ''', 'user_leave': '''
                    CREATE TABLE IF NOT EXISTS user_leave (
                        name TEXT,
                        year TEXT,
                        month TEXT,
                        leave_days TEXT,
                        PRIMARY KEY (name, year, month)
                    )
                '''}

            # Create missing tables
            created_tables = []
            for table_name, query in tables.items():
                if table_name not in existing_tables:
                    try:
                        cursor.execute(query)
                        created_tables.append(table_name)
                    except sqlite3.Error as e:
                        print(f"Error creating table {table_name}: {e}")

            self.db_connection.commit()

            # Update database page with table list
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' ORDER BY name;")
            all_tables = [table[0] for table in cursor.fetchall()]
            self.update_database_page(all_tables, created_tables)

            # Check for current year holidays
            cursor.execute("SELECT year FROM holiday WHERE year = ?", (self.current_year,))
            if not cursor.fetchone():
                self.show_holiday_import_dialog()

        except sqlite3.Error as e:
            error_msg = f"Database error: {str(e)}"
            print(error_msg)
            if hasattr(self, 'db_status_label'):
                self.db_status_label.setText(error_msg)
        finally:
            if cursor:
                cursor.close()

    def show_holiday_import_dialog(self):
        """Show holiday import prompt and handle file selection"""
        reply = QMessageBox.question(self, "Holiday Data Required",
            f"No holidays found for {self.current_year}. Would you like to import from Excel?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)

        if reply == QMessageBox.StandardButton.Yes:
            self.import_holidays_from_excel()

    def import_holidays_from_excel(self):
        """Handle Excel import with datetime-formatted cells"""
        file_path, _ = QFileDialog.getOpenFileName(self, "Select Holiday File", "", "Excel Files (*.xlsx *.xls)")

        if not file_path:
            return

        try:
            wb = load_workbook(filename=file_path)
            sheet = wb.active

            # Get year from first cell
            year_cell = sheet['A1'].value
            if not isinstance(year_cell, int) or len(str(year_cell)) != 4:
                raise ValueError("First cell must contain a 4-digit year (e.g., 2025)")
            excel_year = str(year_cell)

            # Check for existing year in database
            cursor = self.db_connection.cursor()
            cursor.execute("SELECT year FROM holiday WHERE year = ?", (excel_year,))
            if cursor.fetchone():
                QMessageBox.warning(self, "Data Exists", f"Holidays for {excel_year} already exist")
                return

            # Process date cells
            holidays = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                cell_value = row[0]
                if not cell_value:break

                # Handle different cell types
                if isinstance(cell_value, datetime):
                    date_obj = cell_value
                else:
                    try:
                        # Try parsing string format
                        date_obj = datetime.strptime(str(cell_value), "%Y-%m-%d %H:%M:%S")
                    except ValueError:
                        raise ValueError(f"Invalid date format: {cell_value}")

                # Validate year match
                if str(date_obj.year) != excel_year:
                    raise ValueError(f"Date {date_obj} doesn't match Excel year {excel_year}")

                holidays.append(date_obj.strftime("%Y-%m-%d"))

            # Insert into database
            cursor.execute("INSERT INTO holiday (year, holidays) VALUES (?, ?)", (excel_year, json.dumps(holidays)))
            self.db_connection.commit()

            QMessageBox.information(self, "Import Successful", f"Added {len(holidays)} holidays for {excel_year}")

        except Exception as e:
            QMessageBox.critical(self, "Import Error", f"Failed to import holidays:\n{str(e)}")
        finally:
            if cursor:
                cursor.close()

    def update_database_page(self, all_tables, new_tables=None):
        """Update the database page with current table information"""
        self.db_table_list.clear()
        self.db_table_list.addItems(all_tables)

        status_text = f"Loaded {len(all_tables)} tables"
        if new_tables:
            status_text += f"\nCreated new tables: {', '.join(new_tables)}"

        # Show status message in home page for demo
        if hasattr(self, 'home_status_label'):
            self.home_status_label.setText(status_text)

    def init_ui(self):
        # Create custom title bar controls
        title_bar_widget = QWidget()
        title_bar_layout = QHBoxLayout(title_bar_widget)
        title_bar_layout.setContentsMargins(0, 0, 10, 0)

        # Theme toggle button
        self.theme_btn = QPushButton()
        self.theme_btn.setFixedHeight(32)  # Fixed height only
        self.theme_btn.setMinimumWidth(120)  # Minimum width to prevent cutting
        self.theme_btn.clicked.connect(self.cycle_theme)

        # Add vertical spacing around the button
        btn_container = QVBoxLayout()
        btn_container.setContentsMargins(0, 10, 0, 0)  # Top margin for the button
        btn_container.addWidget(self.theme_btn)

        # Set button styling
        self.theme_btn.setStyleSheet("""
                QPushButton {
                    margin-top: 20px; /* Increase top margin */
                    padding: 5px 10px;
                    border-radius: 10px;
                    border: 1px solid #606060;
                    font-weight: 500;
                }
            """)
        self.update_theme_button()

        title_bar_layout.addSpacerItem(QSpacerItem(40, 20, QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Minimum))
        title_bar_layout.addLayout(btn_container)  # Add the container instead of direct button

        # Add title bar widget to main window
        self.setMenuWidget(title_bar_widget)

        # Main layout
        main_widget = QWidget()
        main_layout = QHBoxLayout(main_widget)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)

        # Sidebar (updated to 3 buttons)
        self.sidebar = QFrame()
        self.sidebar.setObjectName("sidebar")
        self.sidebar.setFixedWidth(150)

        sidebar_layout = QVBoxLayout()
        sidebar_layout.setContentsMargins(8, 20, 8, 20)
        sidebar_layout.setSpacing(8)

        self.btn_home = QPushButton("🏠 Home")
        self.btn_database = QPushButton("📁 Database")
        self.btn_about = QPushButton("ℹ️ About")

        for btn in [self.btn_home, self.btn_database, self.btn_about]:
            btn.setCheckable(True)
            btn.setFixedSize(130, 40)
            btn.setFont(QFont("Segoe UI", 10))
            btn.setCursor(Qt.CursorShape.PointingHandCursor)
            sidebar_layout.addWidget(btn)

        sidebar_layout.addStretch()
        self.sidebar.setLayout(sidebar_layout)

        # Main content area
        self.stacked_widget = QStackedWidget()
        self.init_pages()

        main_layout.addWidget(self.sidebar)
        main_layout.addWidget(self.stacked_widget)
        self.setCentralWidget(main_widget)

        # Connect navigation
        self.btn_home.clicked.connect(lambda: self.switch_page(0))
        self.btn_database.clicked.connect(lambda: self.switch_page(1))
        self.btn_about.clicked.connect(lambda: self.switch_page(2))

    def init_pages(self):
        # ======================
        # PAGE CREATION
        # ======================
        self.stacked_widget.addWidget(self.create_home_page())
        self.stacked_widget.addWidget(self.create_database_page())
        self.stacked_widget.addWidget(self.create_about_page())

    def create_home_page(self):
        page = QWidget()
        main_layout = QVBoxLayout(page)
        main_layout.setContentsMargins(20, 15, 20, 15)
        main_layout.setSpacing(15)

        # Year/Month Selection Group
        year_month_group = QGroupBox("Select Month & Year")
        year_month_group.setStyleSheet("""
                QGroupBox {
                    border: 1px solid #e0e0e0;
                    border-radius: 6px;
                    margin-top: 8px;
                    padding-top: 12px;
                }
                QGroupBox::title {
                    subcontrol-origin: margin;
                    left: 10px;
                    color: #118370;
                    font-weight: 500;
                }
            """)

        # Layout for month and year
        date_layout = QHBoxLayout(year_month_group)
        date_layout.setContentsMargins(10, 15, 10, 15)
        date_layout.setSpacing(15)

        # Month ComboBox
        self.month_combo = QComboBox()
        self.month_combo.addItems(
            ["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November",
             "December"])
        self.month_combo.setFixedHeight(35)

        # Year ComboBox
        self.year_combo = QComboBox()
        current_year = QDate.currentDate().year()
        self.year_combo.addItems([str(y) for y in range(2022, 2051)])
        self.year_combo.setFixedHeight(35)

        # Set current date
        current_month = QDate.currentDate().month()
        self.month_combo.setCurrentIndex(current_month - 1)
        self.year_combo.setCurrentText(str(current_year))

        # Function to update month combo box based on selected year
        def update_month_combo():
            selected_year = int(self.year_combo.currentText())
            is_current_year = selected_year == current_year

            # If switching back to the current year, reset month to the current month
            if is_current_year:
                self.month_combo.setCurrentIndex(current_month - 1)

            for idx in range(self.month_combo.count()):
                month_enabled = (idx + 1) <= current_month if is_current_year else True
                self.month_combo.model().item(idx).setEnabled(month_enabled)

        # Disable future years
        for idx in range(self.year_combo.count()):
            year = int(self.year_combo.itemText(idx))
            self.year_combo.model().item(idx).setEnabled(year <= current_year)

        # Connect the year selection change to update months dynamically
        self.year_combo.currentTextChanged.connect(update_month_combo)

        # Initial check on startup
        update_month_combo()

        # Add widgets to layout
        date_layout.addWidget(self.month_combo, 70)  # 70% width
        date_layout.addWidget(self.year_combo, 30)  # 30% width

        # Styling
        combo_style = """
            QComboBox {
                padding: 8px;
                border: 1px solid #cccccc;
                border-radius: 4px;
                background-color: white;
            }
            QComboBox::drop-down {
                subcontrol-origin: padding;
                subcontrol-position: top right;
                width: 25px;
            }
        """
        self.month_combo.setStyleSheet(combo_style)
        self.year_combo.setStyleSheet(combo_style)

        # Holiday Information Section
        holiday_group = QGroupBox("Holiday Management")
        holiday_group.setStyleSheet("""
                QGroupBox {
                    border: 1px solid #e0e0e0;
                    border-radius: 6px;
                    margin-top: 8px;
                    padding-top: 12px;
                }
                QGroupBox::title {
                    color: #118370;
                    font-weight: 500;
                }
            """)
        holiday_layout = QVBoxLayout(holiday_group)
        holiday_layout.setContentsMargins(10, 15, 10, 10)
        holiday_layout.setSpacing(8)

        # Input row with buttons
        input_row = QHBoxLayout()
        input_row.setSpacing(8)

        # Input field
        self.holiday_input = QPlainTextEdit()
        self.holiday_input.setFixedHeight(35)
        self.holiday_input.setReadOnly(True)
        self.holiday_input.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)

        # Buttons
        self.load_holiday_btn = QPushButton("Load Holiday")
        self.load_holiday_btn.setToolTip("Please Upload Excel file as per the format")
        self.load_holiday_btn.setFixedSize(30, 30)

        self.view_holiday_btn = QPushButton("View Holidays")
        self.view_holiday_btn.setFixedSize(120, 35)

        # New action button
        self.format_info_btn = QPushButton("Format")
        self.format_info_btn.setToolTip("Holiday Excel Format")
        self.format_info_btn.setFixedSize(100, 35)

        input_row.addWidget(self.holiday_input)
        input_row.addWidget(self.load_holiday_btn)
        input_row.addWidget(self.view_holiday_btn)
        input_row.addWidget(self.format_info_btn)  # Added new button

        # Connect buttons
        self.format_info_btn.clicked.connect(self.show_format_guide)
        self.load_holiday_btn.clicked.connect(self.load_holidays_from_db)
        self.view_holiday_btn.clicked.connect(self.show_holiday_viewer)

        # When creating buttons
        self.format_info_btn.setObjectName("format_info_btn")
        self.load_holiday_btn.setObjectName("load_holiday_btn")
        self.view_holiday_btn.setObjectName("view_holiday_btn")

        holiday_layout.addLayout(input_row)

        # Error message label
        self.holiday_error_label = QLabel()
        self.holiday_error_label.setStyleSheet("color: red; font-size: 11px;")
        self.holiday_error_label.setWordWrap(True)
        holiday_layout.addWidget(self.holiday_error_label)

        # Styling for new button
        button_style = """
                QPushButton {
                    border-radius: 4px;
                    padding: 6px;
                    font-weight: 500;
                }
            """
        self.format_info_btn.setStyleSheet(button_style + "background-color: #aae0a4; color: black;")

        # Category Information Section
        category_group = QGroupBox("User Categories")
        category_group.setStyleSheet("""
                QGroupBox {
                    border: 1px solid #e0e0e0;
                    border-radius: 6px;
                    margin-top: 8px;
                    padding-top: 12px;
                }
                QGroupBox::title {
                    color: #118370;
                    font-weight: 500;
                }
            """)
        category_layout = QVBoxLayout(category_group)
        category_layout.setContentsMargins(10, 15, 10, 10)
        category_layout.setSpacing(8)

        # Input row with buttons
        catergory_input_row = QHBoxLayout()
        catergory_input_row.setSpacing(8)

        # Input field
        self.category_input = QPlainTextEdit()
        self.category_input.setFixedHeight(35)
        self.category_input.setReadOnly(True)
        self.category_input.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)

        # Buttons
        self.category_btn = QPushButton("Load Category")
        self.category_btn.setToolTip("Please Upload Excel file as per the format")
        self.category_btn.setFixedSize(30, 30)

        catergory_input_row.addWidget(self.category_input)
        catergory_input_row.addWidget(self.category_btn)

        # Connect buttons
        self.category_btn.clicked.connect(self.select_category)

        # When creating buttons
        self.category_btn.setObjectName("category_btn")
        category_layout.addLayout(catergory_input_row)

        # Main group box
        groupBox = QGroupBox("Please enter raw attendance file to generate Attendance Billing Report")
        groupBox.setStyleSheet("""
                QGroupBox {
                border: 1px solid #e0e0e0;
                border-radius: 6px;
                margin-top: 8px;
                padding-top: 12px;
            }
            QGroupBox::title {
                color: #118370;
                font-weight: 500;
            }
            # QGroupBox::title {
            #     subcontrol-origin: margin;
            #     left: 10px;
            #     padding: 0 3px;
            # }
        """)
        group_layout = QVBoxLayout(groupBox)
        group_layout.setContentsMargins(10, 15, 10, 10)  # Reduced top/bottom margins (was 10,15,10,10)
        group_layout.setSpacing(8)  # Reduced spacing between widgets

        # File input box
        main_input_row = QHBoxLayout()
        main_input_row.setSpacing(8)

        # Input field
        self.main_input = QPlainTextEdit()
        self.main_input.setFixedHeight(35)
        self.main_input.setReadOnly(True)
        self.main_input.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)

        # Upload button
        self.upload_button = QPushButton("Upload")
        self.upload_button.setIcon(QIcon.fromTheme("folder"))
        self.upload_button.setToolTip("Select file")

        main_input_row.addWidget(self.main_input)
        main_input_row.addWidget(self.upload_button)

        # Connect buttons
        self.upload_button.clicked.connect(self.upload_file)

        # Add to main group
        # When creating buttons
        self.upload_button.setObjectName("upload_button")
        group_layout.addLayout(main_input_row)

        # Status message system
        # Create the GroupBox
        statusBox = QGroupBox("Status Message")
        statusBox.setStyleSheet("""
                        QGroupBox {
                        border: 1px solid #e0e0e0;
                        border-radius: 6px;
                        margin-top: 8px;
                        padding-top: 12px;
                    }
                    QGroupBox::title {
                        color: #118370;
                        font-weight: 500;
                    }
                    # QGroupBox::title {
                    #     subcontrol-origin: margin;
                    #     left: 10px;
                    #     padding: 0 3px;
                    # }
                """)
        status_group_box = QVBoxLayout(statusBox)
        status_group_box.setContentsMargins(10, 15, 10, 10)  # Reduced top/bottom margins (was 10,15,10,10)
        status_group_box.setSpacing(8)  # Reduced spacing between widgets

        # File input box
        status_input_row = QHBoxLayout()
        status_input_row.setSpacing(8)

        self.msg_label = QLabel()
        self.msg_label.setObjectName("statusMsg")
        self.msg_label.setStyleSheet("""
                QLabel {
                    padding: 10px;
                    border-radius: 4px;
                    margin: 8px 0;
                    font-size: 12px;
                    min-height: 40px;
                }
                QLabel[messageType="error"] {
                    color: #dc3545;
                    background-color: #f8d7da;
                    border: 1px solid #f5c6cb;
                }
                QLabel[messageType="success"] {
                    color: #28a745;
                    background-color: #d4edda;
                    border: 1px solid #c3e6cb;
                }
                QLabel[messageType="info"] {
                    color: #004085;
                    background-color: #cce5ff;
                    border: 1px solid #b8daff;
                }
            """)
        self.msg_label.setWordWrap(True)
        self.msg_label.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        self.msg_label.hide()

        status_input_row.addWidget(self.msg_label)
        status_group_box.addLayout(status_input_row)

        # # Show error message (red)
        # self.show_message("Error: File not found!", "error", 5000)
        #
        # # Show success message (green)
        # self.show_message("Data loaded successfully!", "success", 3000)
        #
        # # Show info message (blue)
        # self.show_message("Processing completed", "info", 2000)

        # Progress Bar
        # progress_input_row = QHBoxLayout()
        # progress_input_row.setSpacing(8)
        progress_group = QWidget()
        progress_layout = QVBoxLayout(progress_group)
        progress_layout.setContentsMargins(0, 10, 0, 0)
        
        self.progress_bar = QProgressBar()
        self.progress_bar.setObjectName("progress_bar")
        self.progress_bar.setValue(0)
        self.progress_bar.setTextVisible(True)
        self.progress_bar.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.progress_bar.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        self.progress_bar.setStyleSheet("""
                QProgressBar {
                    height: 20px;
                    border: 1px solid #cccccc;
                    border-radius: 4px;
                    text-align: center;
                }
                QProgressBar::chunk {
                    background-color: #118370;
                    border-radius: 3px;
                }
            """)

        # progress_input_row.addWidget(self.progress_bar)
        progress_layout.addWidget(self.progress_bar)


        # Generate Button
        generate_container = QHBoxLayout()
        generate_container.addStretch()

        self.generateButton = QPushButton("Generate Report")
        self.generateButton.setObjectName("generateButton")
        self.generateButton.setFixedHeight(35)
        self.generateButton.setStyleSheet("""
                QPushButton {
                    background-color: #118370;
                    color: white;
                    border-radius: 4px;
                    padding: 8px;
                    font-weight: 500;
                }
                QPushButton:hover {
                    background-color: #0f7460;
                }
                QPushButton:disabled {
                    background-color: #cccccc;
                    color: #666666;
                }
            """)
        generate_container.addWidget(self.generateButton)

        progress_layout.addLayout(generate_container)

        # Create button container for alignment
        button_container = QHBoxLayout()
        button_container.addStretch()
        button_container.addWidget(self.generateButton)

        # Add to main layout without stretching
        main_layout.addWidget(year_month_group)
        main_layout.addWidget(holiday_group)
        main_layout.addWidget(category_group)
        # Add main group to page
        main_layout.addWidget(groupBox)

        main_layout.addWidget(self.progress_bar)
        main_layout.addLayout(button_container)
        # main_layout.addWidget(self.msg_label)
        main_layout.addWidget(statusBox)

        main_layout.addStretch()
        # Update progress
        self.progress_bar.setValue(50)  # 0-100%
        # Set fixed height for message label to prevent layout shift
        self.msg_label.setFixedHeight(0)

        def adjust_message_height():
            if self.msg_label.isVisible():
                self.msg_label.setFixedHeight(self.msg_label.sizeHint().height())
            else:
                self.msg_label.setFixedHeight(0)

        self.msg_label.showEvent = lambda e: adjust_message_height()
        self.msg_label.hideEvent = lambda e: adjust_message_height()

        return page

    def show_message(self, text, msg_type="info", timeout=5000):
        """Show dynamic message with auto-hide"""
        self.msg_label.setProperty("messageType", msg_type)
        self.msg_label.setText(text)
        self.msg_label.style().polish(self.msg_label)

        # Adjust height dynamically
        self.msg_label.setFixedHeight(self.msg_label.sizeHint().height())

        self.msg_label.show()

        # Auto-hide after timeout
        if timeout > 0:
            QTimer.singleShot(timeout, self.clear_message)

    def clear_message(self):
        """Clear the status message"""
        self.msg_label.hide()
        self.msg_label.setText("")
    def upload_file(self):
        file_dialog = QFileDialog(self)
        filepath, _ = file_dialog.getOpenFileName(self, "Open Excel File", "", "Excel Files (*.xlsx *.xls, *.csv)")

        if filepath:
            fileInfo = QtCore.QFileInfo(filepath)
            file_name = fileInfo.fileName()
            file_size = fileInfo.size()  # in bytes
            # Convert file size to kilobytes
            file_size_kb = file_size / 1024.0
            print(f"File Name: {file_name}, File Size: {file_size_kb:.2f} KB")
            self.main_input.setPlainText(f"{file_name} ({file_size_kb:.2f} KB)")

            # Create DataFrame with Pandas
            self.df = read_file(filepath)

            # Adding Validation to file upload
            try:
                dict_value = dict(Counter(list(itertools.chain.from_iterable(
                    [[item.split("-")[-1] for item in j] for j in [list(i.keys())[4:-2] for i in self.df]]))))
                value = max(dict_value, key=dict_value.get)

                self.selected_month = self.month_combo.currentText()

                if value == self.selected_month[:3]:
                    self.show_message("Valid Raw Excel Loaded", "info", 2000)
                else:
                    self.show_message("Invalid Raw excel, please check the file or selected month.", "error", 5000)
                    self.df = None
                # Update progress
                self.progress_bar.setValue(50)  # 0-100%

                # Show/hide when needed
                self.progress_bar.setVisible(True)

                # Reset on completion
                # self.progress_bar.reset()
            except Exception as e:
                print(e)
                self.show_message("Invalid Raw excel, please check the file.", "error", 5000)
                self.df = None

    def select_category(self):
        file_dialog = QFileDialog(self)
        filepath, _ = file_dialog.getOpenFileName(self, "Open Excel File", "", "Excel Files (*.xlsx *.xls, *.csv)")

        self.raw_category_list, self.name_order_list = ([],) * 2
        self.category = dict()
        self.name_mapping = dict()

        try:
            if filepath:
                fileInfo = QtCore.QFileInfo(filepath)
                file_name = fileInfo.fileName()
                file_size = fileInfo.size()  # in bytes
                # Convert file size to kilobytes
                file_size_kb = file_size / 1024.0
                print(f"File Name: {file_name}, File Size: {file_size_kb:.2f} KB")

                self.category_input.setPlainText(f"{file_name} ({file_size_kb:.2f} KB)")

                sheet_name = "PublicCloudResourceList"
                # Read the Excel file into a DataFrame
                df = pd.read_excel(filepath, sheet_name=sheet_name)
                df.replace({np.nan: None}, inplace=True)

                # Convert the DataFrame to a list of dictionaries
                self.raw_category_list = df.to_dict(orient="records")

                # Standardize the "Full Name" field by removing commas and spaces
                for item in self.raw_category_list:
                    name  = clean_string(item['Full Name'])
                    team = item["Team"]

                    # Check if the team is already in the dictionary
                    if team in self.category:
                        self.category[team].append(name)
                    else:
                        self.category[team] = [name]

                    self.name_mapping.update({name:[item["521 ID"], item["Point of Contact"], item["Start Date"],
                        item["End Date"]]})



                # Create a mapping from Full Name to 521 ID for quick lookup
                # self.name_mapping = {
                #     preprocess_name(item["Full Name"]): [item["521 ID"], item["Point of Contact"], item["Start Date"],
                #         item["End Date"], ] for item in self.raw_category_list}

                for k, v in self.category.items():
                    temp_list = sorted(v)
                    self.category[k] = temp_list
                    self.name_order_list.extend(temp_list)
                QMessageBox.information(self, "Success", "Category data successfully created!",
                    QMessageBox.StandardButton.Ok)

        except Exception as e:
            self.raw_category_list, self.name_order_list = ([],) * 2
            self.category, self.name_mapping = (dict(),) * 2
    def show_format_guide(self):
        """Show Excel format requirements"""
        guide_text = """
                <html>
                    <head>
                        <style>
                            body {
                                font-family: Arial, sans-serif;
                            }
                            table {
                                font-family: Arial, sans-serif;
                                border-collapse: collapse;
                                width: 100%;
                                margin-top: 10px;
                            }
                            td, th {
                                border: 1px solid #dddddd;
                                text-align: left;
                                padding: 8px;
                            }
                            tr:nth-child(even) {
                                background-color: #f2f2f2;
                            }
                            .info {
                                font-size: 14px;
                                margin-bottom: 10px;
                            }
                            .title {
                                font-size: 18px;
                                font-weight: bold;
                                margin-bottom: 10px;
                            }
                        </style>
                    </head>
                    <body>
                        <div class="title">Excel Format Requirements</div>
                        <div class="info">
                            <ol>
                                <li>First cell must contain the 4-digit year (e.g., <b>2025</b>).</li>
                                <li>Subsequent cells should contain dates in any valid Excel date format.</li>
                                <li>Dates must belong to the specified year.</li>
                            </ol>
                        </div>

                        <h2>Holiday List</h2>
                        <table>
                            <tr>
                                <th>Year</th>
                            </tr>
                            <tr>
                                <td>DD-MM-YYYY</td>
                            </tr>
                            <tr>
                                <td>DD-MM-YYYY</td>
                            </tr>
                            <tr>
                                <td>etc ...</td>
                            </tr>
                        </table>
                    </body>
                </html>
            """

        msg_box = QMessageBox(self)
        msg_box.setWindowTitle("Format Guide")
        msg_box.setTextFormat(Qt.TextFormat.RichText)  # Enable HTML rendering
        msg_box.setText(guide_text)
        msg_box.exec()

    def load_holidays_from_db(self):
        """Open file dialog and import holidays from Excel"""
        file_path, _ = QFileDialog.getOpenFileName(self, "Select Holiday Excel File", "", "Excel Files (*.xlsx *.xls)")

        if not file_path:
            return

        try:
            # Get file info before processing
            file_name = os.path.basename(file_path)
            file_size = os.path.getsize(file_path)

            # Convert bytes to human-readable format
            def sizeof_fmt(num, suffix='B'):
                for unit in ['', 'K', 'M', 'G', 'T', 'P', 'E', 'Z']:
                    if abs(num) < 1024.0:
                        return f"{num:3.1f} {unit}{suffix}"
                    num /= 1024.0
                return f"{num:.1f} Y{suffix}"

            size_str = sizeof_fmt(file_size)
            display_text = f"{file_name} ({size_str})"
            self.holiday_input.setPlainText(display_text)

            # Read Excel file
            df = pd.read_excel(file_path, header=None)

            # Get year from first cell
            excel_year = str(df.iloc[0, 0])
            if not excel_year.isdigit() or len(excel_year) != 4:
                raise ValueError("First cell must contain a 4-digit year")

            # Process dates
            holidays = []
            for idx in range(1, len(df)):
                date_val = df.iloc[idx, 0]

                if pd.isna(date_val):
                    continue

                try:
                    # Handle different date formats
                    if isinstance(date_val, str):
                        # Parse string date
                        date_obj = datetime.strptime(date_val, "%Y-%m-%d")
                    elif isinstance(date_val, pd.Timestamp):
                        # Convert pandas Timestamp to datetime
                        date_obj = date_val.to_pydatetime()
                    else:
                        # Assume native datetime object
                        date_obj = date_val

                    # Validate year match
                    if str(date_obj.year) != excel_year:
                        # raise ValueError(f"Date {date_obj.date()} doesn't match Excel year {excel_year}")
                        QMessageBox.critical(self, "Invalid Date",
                            f"Error in row {idx + 1}:\nDate {date_obj.date()} doesn't match Excel year {excel_year}")
                        return  # Stop the operation immediately

                    holidays.append(date_obj.strftime("%Y-%m-%d"))

                except Exception as e:
                    QMessageBox.critical(self, "Invalid Date", f"Error in row {idx + 1}:\n{str(e)}")
                    return

            # Check for existing entry
            cursor = self.db_connection.cursor()
            cursor.execute("SELECT year FROM holiday WHERE year = ?", (excel_year,))
            exists = cursor.fetchone()

            if exists:
                confirm = QMessageBox.question(self, "Override Confirmation",
                    f"Holidays for {excel_year} already exist. Override?",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
                if confirm != QMessageBox.StandardButton.Yes:
                    return

            # Insert/Update data
            cursor.execute("INSERT OR REPLACE INTO holiday (year, holidays) VALUES (?, ?)",
                (excel_year, json.dumps(holidays)))
            self.db_connection.commit()

            # Update UI
            current_year = QDate.currentDate().year()
            if not self.year_combo.findText(excel_year):
                self.year_combo.addItem(excel_year)
            self.year_combo.setCurrentText(excel_year)

            # self.holiday_input.setPlainText("\n".join(holidays))
            QMessageBox.information(self, "Success", f"Loaded {len(holidays)} holidays for {excel_year}")

        except Exception as e:
            QMessageBox.critical(self, "Import Error", f"Failed to load holidays:\n{str(e)}")
        finally:
            if cursor:
                cursor.close()

    def show_holiday_viewer(self):
        """Show holiday viewer dialog with table display"""
        dialog = QDialog(self)
        dialog.setWindowTitle("Holiday List")
        dialog.setMinimumSize(500, 400)  # Increased width and height for better visibility

        layout = QVBoxLayout(dialog)

        # Year selection
        year_layout = QHBoxLayout()
        year_label = QLabel("Select Year:")

        self.viewer_year_combo = QComboBox()
        self.viewer_year_combo.setFixedWidth(100)  # Increased width for better readability
        self.viewer_year_combo.addItems([str(y) for y in range(2024, 2051)])
        # Set the current year as the default selection
        current_year = str(QDate.currentDate().year())
        self.viewer_year_combo.setCurrentText(current_year)

        year_layout.addWidget(year_label)
        year_layout.addWidget(self.viewer_year_combo)
        year_layout.addStretch()

        # Table setup with scroll area
        self.holiday_table = QTableWidget()
        self.holiday_table.setColumnCount(2)
        self.holiday_table.setHorizontalHeaderLabels(["Date", "Day"])
        self.holiday_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)

        # Wrap the table in a scroll area
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)

        table_container = QWidget()
        table_layout = QVBoxLayout(table_container)
        table_layout.addWidget(self.holiday_table)
        scroll_area.setWidget(table_container)

        # Refresh button
        refresh_btn = QPushButton("Refresh")
        refresh_btn.setFixedHeight(35)  # Adjust button height for a better look
        refresh_btn.clicked.connect(lambda: self.populate_holiday_table(self.viewer_year_combo.currentText()))

        layout.addLayout(year_layout)
        layout.addWidget(scroll_area)  # Add scrollable table
        layout.addWidget(refresh_btn)

        # Initial population
        self.viewer_year_combo.currentTextChanged.connect(self.populate_holiday_table)
        self.populate_holiday_table(self.viewer_year_combo.currentText())

        dialog.exec()

    def populate_holiday_table(self, year):
        """Populate table with holidays for selected year"""
        try:
            cursor = self.db_connection.cursor()
            cursor.execute("SELECT holidays FROM holiday WHERE year = ?", (year,))
            result = cursor.fetchone()

            self.holiday_table.setRowCount(0)

            if result:
                holidays = json.loads(result[0])
                self.holiday_table.setRowCount(len(holidays))

                for row, date_str in enumerate(holidays):
                    date = QDateTime.fromString(date_str, "yyyy-MM-dd")

                    self.holiday_table.setItem(row, 0, QTableWidgetItem(date_str))
                    self.holiday_table.setItem(row, 1, QTableWidgetItem(date.toString("dddd")))

        except sqlite3.Error as e:
            QMessageBox.critical(self, "Database Error", f"Failed to load holidays:\n{str(e)}")

    def create_config_page(self):
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(40, 40, 40, 40)

        title = QLabel("Configuration Settings")
        title.setObjectName("title")

        form = QFormLayout()
        form.setVerticalSpacing(15)
        form.setHorizontalSpacing(20)

        self.username_input = QLineEdit()
        self.theme_combo = QComboBox()
        self.theme_combo.addItems(["Dark", "Light", "System"])
        self.theme_combo.currentTextChanged.connect(self.change_theme)

        form.addRow(QLabel("Username:"), self.username_input)
        form.addRow(QLabel("Interface Theme:"), self.theme_combo)

        layout.addWidget(title)
        layout.addSpacing(20)
        layout.addLayout(form)
        layout.addStretch()
        return page

    def create_database_page(self):
        """Create the database management page"""
        page = QWidget()
        layout = QHBoxLayout(page)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(20)

        # Table list panel
        left_panel = QWidget()
        left_layout = QVBoxLayout(left_panel)

        self.db_table_list = QListWidget()
        self.db_table_list.itemClicked.connect(self.show_table_contents)
        left_layout.addWidget(QLabel("Database Tables:"))
        left_layout.addWidget(self.db_table_list)

        # Table controls
        self.delete_btn = QPushButton("🗑️ Delete Table")
        self.delete_btn.clicked.connect(self.delete_current_table)
        self.delete_btn.setEnabled(False)

        left_layout.addWidget(self.delete_btn)
        left_layout.addStretch()

        # Table view panel
        right_panel = QWidget()
        right_layout = QVBoxLayout(right_panel)

        self.table_view = QTableWidget()
        self.table_view.setSortingEnabled(True)
        self.table_view.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        right_layout.addWidget(self.table_view)

        layout.addWidget(left_panel, 1)
        layout.addWidget(right_panel, 3)

        return page

    def show_table_contents(self, item):
        """Display contents of selected table"""
        self.current_table = item.text()
        self.delete_btn.setEnabled(True)

        try:
            cursor = self.db_connection.cursor()
            cursor.execute(f"SELECT * FROM {self.current_table}")
            rows = cursor.fetchall()

            # Get column names
            cursor.execute(f"PRAGMA table_info({self.current_table})")
            columns = [col[1] for col in cursor.fetchall()]

            # Configure table view
            self.table_view.setRowCount(len(rows))
            self.table_view.setColumnCount(len(columns))
            self.table_view.setHorizontalHeaderLabels(columns)

            # Populate data
            for row_idx, row in enumerate(rows):
                for col_idx, value in enumerate(row):
                    item = QTableWidgetItem(str(value))
                    self.table_view.setItem(row_idx, col_idx, item)

        except sqlite3.Error as e:
            QMessageBox.critical(self, "Database Error", f"Failed to load table:\n{str(e)}")
        finally:
            if cursor:
                cursor.close()

    def delete_current_table(self):
        """Delete currently selected table with confirmation"""
        if not self.current_table:
            return

        confirm = QMessageBox.question(self, "Confirm Delete",
            f"Are you sure you want to delete table '{self.current_table}'?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)

        if confirm == QMessageBox.StandardButton.Yes:
            try:
                cursor = self.db_connection.cursor()
                cursor.execute(f"DROP TABLE {self.current_table}")
                self.db_connection.commit()

                # Refresh UI
                self.initialize_database()
                self.table_view.clear()
                self.current_table = None
                self.delete_btn.setEnabled(False)

            except sqlite3.Error as e:
                QMessageBox.critical(self, "Database Error", f"Failed to delete table:\n{str(e)}")
            finally:
                if cursor:
                    cursor.close()

    def closeEvent(self, event):
        """Close database connection when window closes"""
        if self.db_connection:
            self.db_connection.close()
        event.accept()
    def create_about_page(self):
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setAlignment(Qt.AlignmentFlag.AlignCenter)

        title = QLabel("About This Application")
        title.setObjectName("title")

        content = QLabel("Proprietary Software - Hitachi Digital Solutions\n\n"
                         "Version 1.2.0\n"
                         "Build Date: 2023-07-20\n\n"
                         "Developed using:\n"
                         "• Python 3.11\n"
                         "• PyQt6 Framework\n"
                         "• SQLite Database\n"
                         "** Developed by : Vimit **")
        content.setAlignment(Qt.AlignmentFlag.AlignCenter)

        layout.addWidget(title)
        layout.addSpacing(20)
        layout.addWidget(content)
        layout.addStretch()
        return page

    # ======================
    # CORE FUNCTIONALITY
    # ======================
    def switch_page(self, index):
        self.stacked_widget.setCurrentIndex(index)
        buttons = self.sidebar.findChildren(QPushButton)
        for i, btn in enumerate(buttons):
            btn.setChecked(i == index)

    def cycle_theme(self):
        current_index = self.themes.index(self.current_theme)
        new_index = (current_index + 1) % len(self.themes)
        self.current_theme = self.themes[new_index]
        self.update_theme_button()
        self.change_theme(self.current_theme)

    def update_theme_button(self):
        theme_data = {"Dark": {"icon": "🌙", "text": "Dark"}, "Light": {"icon": "☀️", "text": "Light"},
            "System": {"icon": "⚙️", "text": "System"}}

        current_data = theme_data[self.current_theme]
        btn_text = f"{current_data['icon']} {current_data['text']}"

        # Calculate required width based on text content
        fm = self.theme_btn.fontMetrics()
        text_width = fm.horizontalAdvance(btn_text) + 20  # Add padding

        # Set dynamic width constraints
        self.theme_btn.setMinimumWidth(text_width)
        self.theme_btn.setMaximumWidth(text_width + 20)

        self.theme_btn.setText(btn_text)

        # Update button colors based on theme
        if self.current_theme == "Dark":
            self.theme_btn.setStyleSheet("""
                QPushButton {
                    background-color: #353535;
                    color: white;
                    border-color: #454545;
                }
            """)
        elif self.current_theme == "Light":
            self.theme_btn.setStyleSheet("""
                QPushButton {
                    background-color: #f0f0f0;
                    color: #333333;
                    border-color: #cccccc;
                }
            """)
        else:  # System
            self.theme_btn.setStyleSheet("""
                QPushButton {
                    background-color: #e0e0e0;
                    color: #000000;
                    border-color: #a0a0a0;
                }
            """)

    def change_theme(self, theme_name):
        app = QApplication.instance()
        if theme_name == "Dark":
            app.setStyleSheet(DARK_THEME)
        elif theme_name == "Light":
            app.setStyleSheet(LIGHT_THEME)
        else:
            app.setStyleSheet("")
        self.current_theme = theme_name


if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setFont(QFont("Segoe UI", 10))
    app.setStyleSheet(LIGHT_THEME)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())